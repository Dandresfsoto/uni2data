from __future__ import absolute_import, unicode_literals

import math

from config.celery import app
import openpyxl
from io import BytesIO
from usuarios.models import User
from django.conf import settings
from openpyxl.drawing.image import Image
from django.core.files import File
from mail_templated import send_mail
from direccion_financiera import models
from config.functions import construir_reporte
from reportes import models as models_reportes
from usuarios.models import Notifications
from desplazamiento.models import Solicitudes, Desplazamiento

@app.task
def build_orden_compra(id,email):
    purchase = models.PurchaseOrders.objects.get(id=id)
    products = models.Products.objects.filter(purchase_order = purchase)
    counts = math.ceil(products.count() / 37)


    if counts == 1:
        usuario = User.objects.get(email = email)
        cantidad = models.Products.objects.filter(purchase_order = purchase).count()

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/orden_compra.xlsx')
        ws = wb.get_sheet_by_name('Orden de Compra')
        logo_sican = Image(purchase.enterprise.logo)

        logo_sican.width = 120
        logo_sican.height = 80
        logo_sican.drawing = 100





        ws.add_image(logo_sican, 'C2')
        ws['E2'] = str(purchase.enterprise.name)

        ws['H6'] = str(purchase.enterprise.code) +'-'+ str(purchase.consecutive)
        ws['H7'] = purchase.pretty_date_datetime()


        ws['D6'] = purchase.third.get_full_name()
        ws['D7'] = purchase.third.get_cedula()


        ws['D10'] = purchase.department.nombre
        ws['D11'] = purchase.municipality.nombre
        ws['D12'] = purchase.beneficiary.nombre
        ws['D13'] = purchase.project_order.name




        ws['I63'] = purchase.pretty_date_datetime()

        i = 15

        for product in products:
            ws['B' + str(i)] = product.codigo.upper()
            ws['C' + str(i)] = purchase.subbeneficiary.nombre
            ws['E' + str(i)] = product.name.upper()
            ws['G' + str(i)] = product.stock
            ws['I' + str(i)] = product.pretty_print_price()
            ws['J' + str(i)] = product.pretty_print_total_price()
            i += 1

        ws['I53'] = purchase.pretty_print_subtotal()
        ws['I56'] = purchase.pretty_print_total()

        ws['B58'] = purchase.observation


        ws['G58'] = str(purchase.departure)+ ' %'
        ws['G59'] = str(purchase.counterpart)+ ' %'


        ws['I58'] = purchase.pretty_print_total_percentage_enterprise()
        ws['I59'] = purchase.pretty_print_total_percentage_project()
        filename = str(purchase.id) + '.xlsx'
        wb.save(output)
        purchase.file_purchase_order.save(filename, File(output))
    if counts == 2:
        usuario = User.objects.get(email=email)
        cantidad = models.Products.objects.filter(purchase_order=purchase).count()

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/orden_compra_2.xlsx')
        ws = wb.get_sheet_by_name('Orden de Compra')
        logo_sican = Image(purchase.enterprise.logo)
        logo_sican_2 = Image(purchase.enterprise.logo)

        logo_sican.width = 120
        logo_sican.height = 80
        logo_sican.drawing = 100

        logo_sican_2.width = 120
        logo_sican_2.height = 80
        logo_sican_2.drawing = 100

        ws.add_image(logo_sican, 'C2')
        ws.add_image(logo_sican_2, 'C67')

        ws['E2'] = str(purchase.enterprise.name)
        ws['E67'] = str(purchase.enterprise.name)

        ws['H6'] = str(purchase.enterprise.code) + '-' + str(purchase.consecutive)
        ws['H71'] = str(purchase.enterprise.code) + '-' + str(purchase.consecutive)
        ws['H7'] = purchase.pretty_date_datetime()
        ws['H72'] = purchase.pretty_date_datetime()

        ws['D6'] = purchase.third.get_full_name()
        ws['D71'] = purchase.third.get_full_name()

        ws['D7'] = purchase.third.get_cedula()
        ws['D72'] = purchase.third.get_cedula()

        ws['D10'] = purchase.department.nombre
        ws['D75'] = purchase.department.nombre
        ws['D11'] = purchase.municipality.nombre
        ws['D76'] = purchase.municipality.nombre
        ws['D12'] = purchase.beneficiary.nombre
        ws['D77'] = purchase.beneficiary.nombre
        ws['D13'] = purchase.project_order.name
        ws['D78'] = purchase.project_order.name

        ws['I128'] = purchase.pretty_date_datetime()

        i = 15

        for product in products:
            ws['B' + str(i)] = product.codigo.upper()
            ws['C' + str(i)] = purchase.subbeneficiary.nombre
            ws['E' + str(i)] = product.name.upper()
            ws['G' + str(i)] = product.stock
            ws['I' + str(i)] = product.pretty_print_price()
            ws['J' + str(i)] = product.pretty_print_total_price()
            if i <51 or i>79:
                i += 1
            else:
                i=80

        ws['I118'] = purchase.pretty_print_subtotal()
        ws['I121'] = purchase.pretty_print_total()

        ws['B123'] = purchase.observation

        ws['G123'] = str(purchase.departure) + ' %'
        ws['G124'] = str(purchase.counterpart) + ' %'

        ws['I123'] = purchase.pretty_print_total_percentage_enterprise()
        ws['I124'] = purchase.pretty_print_total_percentage_project()
        filename = str(purchase.id) + '.xlsx'
        wb.save(output)
        purchase.file_purchase_order.save(filename, File(output))
    if counts == 3:
        usuario = User.objects.get(email=email)
        cantidad = models.Products.objects.filter(purchase_order=purchase).count()

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/orden_compra_3.xlsx')
        ws = wb.get_sheet_by_name('Orden de Compra')
        logo_sican = Image(purchase.enterprise.logo)
        logo_sican_2 = Image(purchase.enterprise.logo)
        logo_sican_3 = Image(purchase.enterprise.logo)

        logo_sican.width = 120
        logo_sican.height = 80
        logo_sican.drawing = 100

        logo_sican_2.width = 120
        logo_sican_2.height = 80
        logo_sican_2.drawing = 100

        logo_sican_3.width = 120
        logo_sican_3.height = 80
        logo_sican_3.drawing = 100

        ws.add_image(logo_sican, 'C2')
        ws.add_image(logo_sican_2, 'C67')
        ws.add_image(logo_sican_3, 'C132')


        ws['E2'] = str(purchase.enterprise.name)
        ws['E67'] = str(purchase.enterprise.name)
        ws['E132'] = str(purchase.enterprise.name)

        ws['H6'] = str(purchase.enterprise.code) + '-' + str(purchase.consecutive)
        ws['H71'] = str(purchase.enterprise.code) + '-' + str(purchase.consecutive)
        ws['H136'] = str(purchase.enterprise.code) + '-' + str(purchase.consecutive)

        ws['H7'] = purchase.pretty_date_datetime()
        ws['H72'] = purchase.pretty_date_datetime()
        ws['H137'] = purchase.pretty_date_datetime()

        ws['D6'] = purchase.third.get_full_name()
        ws['D71'] = purchase.third.get_full_name()
        ws['D136'] = purchase.third.get_full_name()

        ws['D7'] = purchase.third.get_cedula()
        ws['D72'] = purchase.third.get_cedula()
        ws['D137'] = purchase.third.get_cedula()

        ws['D10'] = purchase.department.nombre
        ws['D75'] = purchase.department.nombre
        ws['D140'] = purchase.department.nombre

        ws['D11'] = purchase.municipality.nombre
        ws['D76'] = purchase.municipality.nombre
        ws['D141'] = purchase.municipality.nombre

        ws['D12'] = purchase.beneficiary.nombre
        ws['D77'] = purchase.beneficiary.nombre
        ws['D142'] = purchase.beneficiary.nombre

        ws['D13'] = purchase.project_order.name
        ws['D78'] = purchase.project_order.name
        ws['D143'] = purchase.project_order.name

        ws['I193'] = purchase.pretty_date_datetime()

        i = 15

        for product in products:
            ws['B' + str(i)] = product.codigo.upper()
            ws['C' + str(i)] = purchase.subbeneficiary.nombre
            ws['E' + str(i)] = product.name.upper()
            ws['G' + str(i)] = product.stock
            ws['I' + str(i)] = product.pretty_print_price()
            ws['J' + str(i)] = product.pretty_print_total_price()
            if i <51:
                i += 1
            elif i == 51:
                i=80
            elif i > 79 and i < 115:
                i += 1
            elif i == 115:
                i=145
            elif i > 144:
                i += 1

        ws['I183'] = purchase.pretty_print_subtotal()
        ws['I186'] = purchase.pretty_print_total()

        ws['B188'] = purchase.observation

        ws['G188'] = str(purchase.departure) + ' %'
        ws['G189'] = str(purchase.counterpart) + ' %'

        ws['I188'] = purchase.pretty_print_total_percentage_enterprise()
        ws['I189'] = purchase.pretty_print_total_percentage_project()
        filename = str(purchase.id) + '.xlsx'
        wb.save(output)
        purchase.file_purchase_order.save(filename, File(output))
    if counts == 4:
        usuario = User.objects.get(email=email)
        cantidad = models.Products.objects.filter(purchase_order=purchase).count()

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/orden_compra_4.xlsx')
        ws = wb.get_sheet_by_name('Orden de Compra')
        logo_sican = Image(purchase.enterprise.logo)
        logo_sican_2 = Image(purchase.enterprise.logo)
        logo_sican_3 = Image(purchase.enterprise.logo)
        logo_sican_4 = Image(purchase.enterprise.logo)

        logo_sican.width = 120
        logo_sican.height = 80
        logo_sican.drawing = 100

        logo_sican_2.width = 120
        logo_sican_2.height = 80
        logo_sican_2.drawing = 100

        logo_sican_3.width = 120
        logo_sican_3.height = 80
        logo_sican_3.drawing = 100

        logo_sican_4.width = 120
        logo_sican_4.height = 80
        logo_sican_4.drawing = 100

        ws.add_image(logo_sican, 'C2')
        ws.add_image(logo_sican_2, 'C67')
        ws.add_image(logo_sican_3, 'C132')
        ws.add_image(logo_sican_4, 'C198')

        ws['E2'] = str(purchase.enterprise.name)
        ws['E67'] = str(purchase.enterprise.name)
        ws['E133'] = str(purchase.enterprise.name)
        ws['E198'] = str(purchase.enterprise.name)

        ws['H6'] = str(purchase.enterprise.code) + '-' + str(purchase.consecutive)
        ws['H71'] = str(purchase.enterprise.code) + '-' + str(purchase.consecutive)
        ws['H136'] = str(purchase.enterprise.code) + '-' + str(purchase.consecutive)
        ws['H202'] = str(purchase.enterprise.code) + '-' + str(purchase.consecutive)

        ws['H7'] = purchase.pretty_date_datetime()
        ws['H72'] = purchase.pretty_date_datetime()
        ws['H137'] = purchase.pretty_date_datetime()
        ws['H203'] = purchase.pretty_date_datetime()

        ws['D6'] = purchase.third.get_full_name()
        ws['D71'] = purchase.third.get_full_name()
        ws['D136'] = purchase.third.get_full_name()
        ws['D202'] = purchase.third.get_full_name()

        ws['D7'] = purchase.third.get_cedula()
        ws['D72'] = purchase.third.get_cedula()
        ws['D137'] = purchase.third.get_cedula()
        ws['D203'] = purchase.third.get_cedula()

        ws['D10'] = purchase.department.nombre
        ws['D75'] = purchase.department.nombre
        ws['D140'] = purchase.department.nombre
        ws['D206'] = purchase.department.nombre

        ws['D11'] = purchase.municipality.nombre
        ws['D76'] = purchase.municipality.nombre
        ws['D141'] = purchase.municipality.nombre
        ws['D207'] = purchase.municipality.nombre

        ws['D12'] = purchase.beneficiary.nombre
        ws['D77'] = purchase.beneficiary.nombre
        ws['D142'] = purchase.beneficiary.nombre
        ws['D208'] = purchase.beneficiary.nombre

        ws['D13'] = purchase.project_order.name
        ws['D78'] = purchase.project_order.name
        ws['D143'] = purchase.project_order.name
        ws['D209'] = purchase.project_order.name

        ws['I259'] = purchase.pretty_date_datetime()

        i = 15

        for product in products:
            ws['B' + str(i)] = product.codigo.upper()
            ws['C' + str(i)] = purchase.subbeneficiary.nombre
            ws['E' + str(i)] = product.name.upper()
            ws['G' + str(i)] = product.stock
            ws['I' + str(i)] = product.pretty_print_price()
            ws['J' + str(i)] = product.pretty_print_total_price()
            if i < 51:
                i += 1
            elif i == 51:
                i = 80
            elif i > 79 and i < 115:
                i += 1
            elif i == 115:
                i = 145
            elif i > 144 and i < 183:
                i += 1
            elif i == 183:
                i = 211
            if i < 210:
                i += 1

        ws['I249'] = purchase.pretty_print_subtotal()
        ws['I252'] = purchase.pretty_print_total()

        ws['B254'] = purchase.observation

        ws['G254'] = str(purchase.departure) + ' %'
        ws['G255'] = str(purchase.counterpart) + ' %'

        ws['I254'] = purchase.pretty_print_total_percentage_enterprise()
        ws['I255'] = purchase.pretty_print_total_percentage_project()
        filename = str(purchase.id) + '.xlsx'
        wb.save(output)
        purchase.file_purchase_order.save(filename, File(output))



    return "Reporte generado"

@app.task
def build_reporte_interno(id, email):
    reporte = models.Reportes.objects.get(id=id)
    pagos = models.Pagos.objects.filter(reporte = reporte)


    if pagos.count() > 0:

        usuario = User.objects.get(email = email)
        cantidad = models.Pagos.objects.filter(reporte = reporte).count()

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/reporte_'+ str(cantidad) +'.xlsx')
        ws = wb.get_sheet_by_name('REPORTE')
        logo_sican = Image(settings.STATICFILES_DIRS[0] + '/img/andes-logo.png')
        #logo_sican.drawing.top = 46
        #logo_sican.drawing.left = 66

        logo_sican.width = 200
        logo_sican.height = 170

        ws.add_image(logo_sican, 'C2')

        ws['F4'] = str(reporte.enterprise.report_code) + ' - ' + str(reporte.consecutive)
        ws['I4'] = reporte.reporte_update_datetime()
        ws['O4'] = usuario.get_full_name_string()

        if reporte.servicio.descontable:
            ws['F5'] = reporte.servicio.nombre + ' - DESCONTABLE'
        else:
            ws['F5'] = reporte.servicio.nombre
        ws['O5'] = reporte.proyecto.nombre


        if reporte.efectivo:
            ws['J6'] = 'PAGO EN EFECTIVO'
        else:
            ws['J6'] = 'CARGAR A LA CUENTA {0}'.format(reporte.proyecto.cuenta)

        ws['F6'] = reporte.tipo_soporte.nombre

        ws['F7'] = str(reporte.id)

        ws['G8'] = cantidad
        ws['K8'] = reporte.reporte_inicio_datetime()
        ws['O8'] = reporte.reporte_fin_datetime()

        i = 12

        for pago in pagos:
            ws['C'+str(i)] = 'ACTIVO'
            ws['D' + str(i)] = pago.tercero.cargo.nombre.upper()
            ws['E' + str(i)] = pago.tercero.fullname().upper()
            ws['G' + str(i)] = pago.tercero.cedula

            if reporte.efectivo:
                ws['H' + str(i)] = ''
                ws['I' + str(i)] = ''
                ws['J' + str(i)] = ''
            else:
                if pago.tercero.first_active_account == True:
                    ws['H' + str(i)] = pago.tercero.banco.nombre.upper()
                    ws['I' + str(i)] = pago.tercero.tipo_cuenta.upper()
                    ws['J' + str(i)] = pago.tercero.cuenta
                elif pago.tercero.second_active_account == True:
                    ws['H' + str(i)] = pago.tercero.bank.nombre.upper()
                    ws['I' + str(i)] = pago.tercero.type.upper()
                    ws['J' + str(i)] = pago.tercero.account
            ws['L' + str(i)] = pago.valor_descuentos().amount
            ws['O' + str(i)] = pago.observacion_pretty().upper()
            i += 1

        wb.save(output)

        filename = str(reporte.id) + '.xlsx'
        reporte.file.save(filename, File(output))

    return "Reporte generado"


@app.task
def send_mail_templated_pago(id,template,dictionary,from_email,list_to_email):
    try:
        send_mail(template, dictionary, from_email, list_to_email)
    except:
        pass
    else:
        models.Pagos.objects.filter(id = id).update(notificado = True)
    return 'Pago reportado a ' + str(list_to_email)

@app.task
def send_mail_templated_reporte(template,dictionary,from_email,list_to_email,attachments):
    send_mail(template, dictionary, from_email, list_to_email,attachments = attachments)
    return 'Email enviado'

@app.task
def send_mail_templated_reporte_delete(template,dictionary,from_email,list_to_email):
    send_mail(template, dictionary, from_email, list_to_email)
    return 'Email enviado'

@app.task
def build_listado_terceros(id):
    from recursos_humanos.models import Contratistas
    reporte = models_reportes.Reportes.objects.get(id = id)
    proceso = "SICAN-LST-TERCEROS"


    titulos = ['Consecutive', 'Nombres', 'Apellidos', 'Tipo identificación', '# Documento', 'Cargo', 'Usuario', 'Celular',
               'Correo', 'Fecha de nacimiento', '# Cuenta', 'Banco', 'Tipo de cuenta']

    formatos = ['0', 'General', 'General', 'General', '0', 'General', 'General', 'General',
                'General', 'dd/mm/yy', '0', 'General', 'General']

    ancho_columnas = [20, 30, 30, 25, 25, 35, 40, 20,
                      40, 25, 30, 30, 30]

    contenidos = []

    i = 0
    for tercero in Contratistas.objects.all().order_by('nombres'):
        i += 1
        contenidos.append([
            int(i),
            tercero.nombres,
            tercero.apellidos,
            tercero.get_tipo_identificacion(),
            tercero.cedula,
            tercero.get_cargo_nombre(),
            tercero.get_usuario_asociado(),
            str(tercero.celular),
            tercero.email,
            tercero.birthday,
            tercero.cuenta,
            tercero.get_banco(),
            tercero.tipo_cuenta
        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))

    notificacion = Notifications.objects.create(
        user = reporte.usuario,
        title = 'Se ha construido un reporte',
        short_description = reporte.nombre,
        body = '<p style="display:inline;">Se ha construido el reporte <b>'+ reporte.nombre +
               '</b>, lo puedes decargar haciendo clic </p><a style="display:inline;" href="'+ reporte.url_file() +'">aqui</a>',
        icon = 'accessibility',
        color = 'orange-text text-darken-4'
    )

    return "Archivo paquete ID: " + filename

@app.task
def build_listado_tercero_especifico(id, tercero_id):

    reporte = models_reportes.Reportes.objects.get(id = id)

    proceso = "SICAN-PGS-TERCEROS"


    titulos = ['Consecutivo', 'Fecha', 'Usuario que reporta el pago', 'Consecutivo del reporte', 'Observación', 'Estado',
               'Valor inicial','Descuento 1','Descuento 2','Descuento 3',
               'Descuento 4','Descuento 5','Valor despues de descuentos']

    formatos = ['0', 'dd/mm/yy h:mm AM/PM', 'General', '0', 'General', 'General',
                '"$"#,##0_);("$"#,##0)', '"$"#,##0_);("$"#,##0)', '"$"#,##0_);("$"#,##0)', '"$"#,##0_);("$"#,##0)',
                '"$"#,##0_);("$"#,##0)','"$"#,##0_);("$"#,##0)','"$"#,##0_);("$"#,##0)','"$"#,##0_);("$"#,##0)']

    ancho_columnas = [20, 30, 30, 25, 30, 25,
                      20, 20, 20, 20,
                      20, 20, 30]

    contenidos = []

    pagos = models.Pagos.objects.filter(tercero__id = tercero_id).order_by('-creation')

    i = 0

    for pago in pagos:
        i += 1
        contenidos.append([
            int(i),
            pago.creation,
            pago.usuario_actualizacion.get_full_name_string(),
            pago.reporte.consecutive,
            pago.observacion,
            pago.estado,
            pago.valor.amount.__float__()] + pago.get_list_descuentos() + [
            pago.valor_descuentos().__float__()
        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))

    notificacion = Notifications.objects.create(
        user = reporte.usuario,
        title = 'Se ha completado un reporte',
        short_description = reporte.nombre,
        body = '<p style="display:inline;">Se ha construido el reporte <b>'+ reporte.nombre +
               '</b>, lo puedes decargar haciendo clic </p><a style="display:inline;" href="'+ reporte.url_file() +'">aqui</a>',
        icon = 'accessibility',
        color = 'orange-text text-darken-4'
    )

    return "Archivo paquete ID: " + filename

@app.task
def build_reporte_pagos(reporte_id,enterprise_id):
    reporte = models_reportes.Reportes.objects.get(id = reporte_id)
    enterprise = models.Enterprise.objects.get(id=enterprise_id)


    proceso = "SION-REPORTE-PAGOS"


    titulos = ['Consecutivo', 'Fecha creación', 'Reporte', 'Rubro presupuestal', 'Tipo','Nombre', 'Servicio', 'Proyecto', 'Contratista', 'Cedula', 'Valor inicial', 'Descuentos',
               'Valor','Estado', 'Tipo de cuenta', 'Banco', 'Cuenta', 'Cargo']

    formatos = ['0', 'dd/mm/yy', '0', 'General', 'General', 'General', 'General', 'General', 'General', '0', '"$"#,##0.00_);[Red]("$"#,##0.00)', '"$"#,##0.00_);[Red]("$"#,##0.00)',
                '"$"#,##0.00_);[Red]("$"#,##0.00)', 'General', 'General', 'General', 'General', 'General']

    ancho_columnas = [20, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30,
                      30, 30, 30,30, 30,30, 30]

    contenidos = []

    i = 0
    for pago in models.Pagos.objects.filter(reporte__enterprise=enterprise, reporte__activo=True).order_by('creation'):
        i += 1
        contenidos.append([
            int(i),
            pago.creation,
            pago.reporte.consecutive,
            pago.get_rubro(),
            'Efectivo' if pago.reporte.efectivo else 'Bancarizado',
            pago.reporte.nombre,
            pago.reporte.servicio.nombre,
            pago.reporte.proyecto.nombre,
            pago.tercero.get_full_name(),
            pago.tercero.cedula,
            pago.valor.amount,
            pago.valor_solo_descuentos_amount(),
            pago.valor_descuentos_amount(),
            pago.estado,
            pago.tipo_cuenta,
            pago.banco,
            pago.cuenta,
            pago.cargo
        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))


    return "Archivo paquete ID: " + filename

@app.task
def build_finantial_reports(reporte_id,enterprise_id):
    reporte = models_reportes.Reportes.objects.get(id = reporte_id)
    enterprise = models.Enterprise.objects.get(id=enterprise_id)


    proceso = "UNI2DATA-REPORTE-PAGOS"


    titulos = ['Consecutivo','Consecutivo del reporte', 'Línea presupuestal','Sub Línea presupuestal Nivel II','Sub Línea presupuestal Nivel III','Cuenta Contable', 'Contratista', 'Cedula','No. Contrato','Concepto del Pago',
               'No. Factura o Equivalente','No. Comprobante de Pago','Fecha de Pago','Valor del Gasto' ,'Deducciones y/o Rentenciones Practicdas','Valor del Gasto']

    formatos = ['0','General','General','General','General','General', '0', 'General', 'General', 'General', 'General', 'General','dd/mm/yy',
                '"$"#,##0.00_);[Red]("$"#,##0.00)','"$"#,##0.00_);[Red]("$"#,##0.00)','"$"#,##0.00_);[Red]("$"#,##0.00)']

    ancho_columnas = [20,20, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30]

    contenidos = []

    i = 0
    for pago in models.Pagos.objects.filter(reporte__enterprise=enterprise, reporte__activo=True).order_by('creation'):
        i += 1
        contenidos.append([
            int(i),
            pago.reporte.get_cod_consecutive(),
            pago.get_rubro(),
            pago.get_rubro_lvl2(),
            pago.get_rubro_lvl3(),
            pago.reporte.cuenta_contable,
            pago.tercero.get_full_name(),
            pago.tercero.cedula,
            pago.get_contrato(),
            pago.reporte.nombre,
            pago.reporte.numero_documento_equivalente,
            pago.reporte.numero_comprobante_pago,
            pago.reporte.fecha_pago,
            pago.valor.amount,
            pago.valor_solo_descuentos_amount(),
            pago.valor_descuentos_amount(),
        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))


    return "Archivo paquete ID: " + filename

@app.task
def build_listado_solicitudes(id):

    reporte = models_reportes.Reportes.objects.get(id = id)

    proceso = "SICAN-SOLICITUDES-DESPLAZAMIENTO"


    titulos = ['Consecutivo','Contratista', 'Cedula', 'Consecutivo solicitud', 'Nombre',
               'Fecha', 'Origen', 'Destino', 'Tipo transporte', 'Transportador', 'Telefono',
               'Valor', 'Observaciones', 'Estado', 'Valor original']

    formatos = ['0', 'General', '0', '0', 'General',
                'dd/mm/yy', 'General', 'General', 'General', 'General', 'General',
                '"$"#,##0_);("$"#,##0)', 'General', 'General', '"$"#,##0_);("$"#,##0)']

    ancho_columnas = [20, 40, 40, 40, 40,
                      30, 30, 30, 30, 30 ,30,
                      30, 60, 30, 30]

    contenidos = []


    i = 0

    for solicitud in Solicitudes.objects.all().order_by('creation'):
        for desplazamiento in Desplazamiento.objects.filter(solicitud = solicitud).order_by('creation'):
            i += 1
            contenidos.append([
                int(i),
                solicitud.get_contratista(),
                solicitud.get_contratista_cedula(),
                solicitud.consecutive,
                solicitud.nombre,
                desplazamiento.fecha,
                desplazamiento.origen,
                desplazamiento.destino,
                desplazamiento.tipo_transporte,
                desplazamiento.transportador,
                desplazamiento.telefono,
                desplazamiento.valor.amount.__float__(),
                desplazamiento.observaciones,
                desplazamiento.verificado,
                desplazamiento.valor_original.amount.__float__()
            ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))

    notificacion = Notifications.objects.create(
        user = reporte.usuario,
        title = 'Se ha completado un reporte',
        short_description = reporte.nombre,
        body = '<p style="display:inline;">Se ha construido el reporte <b>'+ reporte.nombre +
               '</b>, lo puedes decargar haciendo clic </p><a style="display:inline;" href="'+ reporte.url_file() +'">aqui</a>',
        icon = 'accessibility',
        color = 'orange-text text-darken-4'
    )

    return "Archivo paquete ID: " + filename
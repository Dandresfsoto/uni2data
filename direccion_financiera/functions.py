import io
from usuarios.models import User
from django.core.files import File
from direccion_financiera import models
import xlsxwriter
from desplazamiento.models import Solicitudes, Desplazamiento
import openpyxl
from openpyxl.drawing.image import Image
from django.conf import settings
from openpyxl.styles import Border, Side, Alignment

def removeNonAscii(s):
    return "".join(i for i in s if ord(i)<128)


def build_archivo_plano(id, email):

    reporte = models.Reportes.objects.get(id = id)
    usuario = User.objects.get(email = email)
    pagos = models.Pagos.objects.filter(reporte = reporte)

    if pagos.count() > 0:

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {'in_memory': True})
        ws = wb.add_worksheet()

        i = 2

        bold = wb.add_format({'bold': True})

        ws.write('A1','Tipo de identificación',bold)
        ws.write('B1', 'Número de identificación',bold)
        ws.write('C1', 'Nombre',bold)
        ws.write('D1', 'Apellido',bold)
        ws.write('E1', 'Código del banco',bold)
        ws.write('F1', 'Tipo de producto o servicio',bold)
        ws.write('G1', 'Número del producto o servicio',bold)
        ws.write('H1', 'Valor del pago o de la recarga',bold)


        for pago in pagos:
            ws.write('A'+str(i),pago.tercero.tipo_identificacion)
            ws.write('B' + str(i),pago.tercero.cedula)
            ws.write('C' + str(i), removeNonAscii(pago.tercero.nombres.upper()))
            ws.write('D' + str(i),removeNonAscii(pago.tercero.apellidos.upper()))
            ws.write('E' + str(i),pago.tercero.banco.codigo)
            ws.write('F' + str(i),'CA' if pago.tercero.tipo_cuenta == 'Ahorros' else 'CC')
            ws.write('G' + str(i),pago.tercero.cuenta)
            ws.write('H' + str(i),pago.valor_descuentos().amount.__float__())
            i += 1

        ws.set_column('A:B', 28)
        ws.set_column('C:D', 25)
        ws.set_column('E:E', 20)
        ws.set_column('F:H', 30)

        wb.close()

        filename = str(reporte.id) + '.xlsx'
        reporte.plano.save(filename, File(output))

    return "Reporte generado"

def build_desplazamiento_file(id):

    solicitud = Solicitudes.objects.get(id = id)
    desplazamientos = Desplazamiento.objects.filter(solicitud = solicitud)

    if desplazamientos.count() > 0:

        output = io.BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/reintegro_transporte.xlsx')
        ws = wb.get_sheet_by_name('REINTEGRO DE TRANSPORTE')

        logo_sican = Image(settings.STATICFILES_DIRS[0] + '/img/andes-logo.png')
        logo_sican.drawing.top = 58
        logo_sican.drawing.left = 160
        ws.add_image(logo_sican)

        ws['F5'] = 'F-GAF-17-T'
        ws['I5'] = solicitud.pretty_creation_datetime()
        ws['F6'] = solicitud.get_contratista()
        ws['I6'] = solicitud.get_contratista_cedula()
        ws['F7'] = str(solicitud.id)
        ws['F9'] = solicitud.estado

        i = 11

        border = Border(
            left=Side(border_style = 'dotted',color='FF000000'),
            right=Side(border_style='dotted', color='FF000000'),
            top=Side(border_style='dotted', color='FF000000'),
            bottom=Side(border_style='dotted', color='FF000000')
        )

        alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
            shrink_to_fit=False
        )

        for desplazamiento in desplazamientos:

            ws['B' + str(i)] = i - 10
            ws['B' + str(i)].border = border
            ws['B' + str(i)].alignment = alignment

            ws['C' + str(i)] = desplazamiento.fecha
            ws['C' + str(i)].border = border
            ws['C' + str(i)].alignment = alignment

            ws['D' + str(i)] = desplazamiento.origen
            ws['D' + str(i)].border = border
            ws['D' + str(i)].alignment = alignment

            ws['E' + str(i)] = desplazamiento.destino
            ws['E' + str(i)].border = border
            ws['E' + str(i)].alignment = alignment

            ws['F' + str(i)] = desplazamiento.tipo_transporte
            ws['F' + str(i)].border = border
            ws['F' + str(i)].alignment = alignment

            ws['G' + str(i)] = desplazamiento.transportador
            ws['G' + str(i)].border = border
            ws['G' + str(i)].alignment = alignment

            ws['H' + str(i)] = desplazamiento.telefono
            ws['H' + str(i)].border = border
            ws['H' + str(i)].alignment = alignment

            ws['I' + str(i)] = desplazamiento.valor.amount
            ws['I' + str(i)].border = border
            ws['I' + str(i)].alignment = alignment
            ws['I' + str(i)].number_format = '"$"#,##0_);("$"#,##0)'

            ws.row_dimensions[i].height = float(50)

            i += 1

        wb.save(output)

        filename = str(solicitud.id) + '.xlsx'
        solicitud.file.save(filename, File(output))

    return "Archivo generado"
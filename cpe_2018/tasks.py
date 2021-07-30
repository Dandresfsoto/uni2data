from __future__ import absolute_import, unicode_literals
from config.celery import app
import openpyxl
from io import BytesIO
from usuarios.models import User
from django.conf import settings
from openpyxl.drawing.image import Image
from django.core.files import File
from mail_templated import send_mail
from direccion_financiera import models
from config.functions import construir_reporte, construir_reporte_pagina
from reportes import models as models_reportes
from usuarios.models import Notifications
from cpe_2018 import models
from six.moves.urllib import parse as urlparse
import os
import ftplib
import io
import datetime
import json
import shutil
from django.utils import timezone

@app.task
def send_mail_templated_cuenta_cobro(template,dictionary,from_email,list_to_email):
    send_mail(template, dictionary, from_email, list_to_email)
    return 'Email enviado'

@app.task
def build_resultado_actualizacion_radicados(id):

    location = os.getenv('FTP_SICAN_2018')
    splitted_url = urlparse.urlparse(location)
    actualizacion = models.ActualizacionRadicados.objects.get(id=id)

    ftp = ftplib.FTP()

    input = io.BytesIO()

    try:
        ftp.connect(splitted_url.hostname,int(splitted_url.port))
        ftp.login(splitted_url.username,splitted_url.password)
        ftp.retrbinary('RETR /' + actualizacion.file.name, input.write)
        input.seek(0)
    except:
        return 'Error FTP'
    else:
        ftp.quit()
        wb = openpyxl.load_workbook(input, use_iterators=True)
        ws = wb.get_sheet_by_name('ACTUALIZACION RADICADOS')

        nuevos = 0
        modificados = 0
        rechazados = 0

        proceso = "SICAN-ACTUALIZACIÓN-RADICADOS"

        titulos = [
            'Código municipio', 'Numero de radicado', 'Nombre institución educativa', 'Dane sede educativa', 'Nombre sede educativa',
            'Tipologia sede educativa','Ubicación de la sede','Estado del radicado','Cantidad portatiles','Cantidad kvd',
            'Cantidad equipos escritorio','Cantidad tabletas','Matricula sede','Observaciones','Resultado'
        ]

        formatos = ['0', '0', 'General', '0', 'General',
                    'General', 'General', 'General', '0', '0',
                    '0', '0', '0', 'General', 'General']

        ancho_columnas = [20, 40, 40, 40, 40,
                          40, 40, 40, 40, 40,
                          40, 40, 40, 40, 40]

        contenidos = []

        for fila in ws.iter_rows(row_offset=1):

            resultado = ''

            if fila[0].value != '' and fila[0].value != None:

                try:
                    municipio = models.Municipios.objects.get(numero=fila[0].value)
                except:
                    resultado = 'No existen municipios con el código ingresado'
                    rechazados += 1
                else:
                    if fila[2].value != '' and fila[3].value != '' and fila[4].value != '' and fila[5].value != '' and fila[6].value != '' and fila[7].value != '' and fila[2].value != None and fila[3].value != None and fila[4].value != None and fila[5].value != None and fila[6].value != None and fila[7].value != None:

                        try:
                            numero = int(fila[1].value)
                        except:
                            resultado = 'Error con el numero de radicado'
                            rechazados += 1
                        else:

                            if fila[5].value == 'A' or fila[5].value == 'D' or fila[5].value == 'A,D':
                                if fila[6].value == 'Rural' or fila[6].value == 'Urbana':
                                    if fila[7].value == 'Aprobado' or fila[7].value == 'Sale de beneficio':
                                        if models.Radicados.objects.filter(numero = numero).count() == 1:
                                            radicado = models.Radicados.objects.get(numero = numero)
                                            radicado.municipio = municipio
                                            radicado.nombre_ie = fila[2].value
                                            radicado.dane_sede = fila[3].value
                                            radicado.nombre_sede = fila[4].value
                                            radicado.tipologia_sede = fila[5].value
                                            radicado.ubicacion = fila[6].value
                                            radicado.estado = fila[7].value
                                            radicado.save()

                                            resultado = 'Radicado actualizado'
                                            modificados += 1

                                        else:
                                            radicado = models.Radicados.objects.create(
                                                municipio = municipio,
                                                numero = numero,
                                                nombre_ie = fila[2].value,
                                                dane_sede = fila[3].value,
                                                nombre_sede = fila[4].value,
                                                tipologia_sede = fila[5].value,
                                                ubicacion = fila[6].value,
                                                estado = fila[7].value
                                            )

                                            resultado = 'Radicado creado'
                                            nuevos += 1

                                        try:
                                            portatiles = int(fila[8].value)
                                        except:
                                            portatiles = None

                                        try:
                                            kvd = int(fila[9].value)
                                        except:
                                            kvd = None

                                        try:
                                            equipos_escritorio = int(fila[10].value)
                                        except:
                                            equipos_escritorio = None

                                        try:
                                            tabletas = int(fila[11].value)
                                        except:
                                            tabletas = None

                                        try:
                                            matricula = int(fila[12].value)
                                        except:
                                            matricula = None

                                        radicado.portatiles = portatiles
                                        radicado.kvd = kvd
                                        radicado.equipos_escritorio = equipos_escritorio
                                        radicado.tabletas = tabletas
                                        radicado.matricula = matricula
                                        radicado.observaciones = fila[13].value
                                        radicado.save()

                                    else:
                                        resultado = 'El estado del radicado debe ser Aprobado o Sale de beneficio'
                                        rechazados += 1
                                else:
                                    resultado = 'La ubicación de la sede debe ser Rural o Urbana'
                                    rechazados += 1

                            else:
                                resultado = 'La tipologia de la sede no corresponde'
                                rechazados += 1

                    else:
                        resultado = 'Información obligatoria incompleta'
                        rechazados += 1


            contenidos.append([
                fila[0].value,
                fila[1].value,
                fila[2].value,
                fila[3].value,
                fila[4].value,
                fila[5].value,
                fila[6].value,
                fila[7].value,
                fila[8].value,
                fila[9].value,
                fila[10].value,
                fila[11].value,
                fila[12].value,
                fila[13].value,
                resultado
            ])

        output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, 'Actualiación de radicados', actualizacion.creation,
                                   actualizacion.usuario_creacion, proceso)

        filename = str(actualizacion.id) + '.xlsx'
        actualizacion.result.save(filename, File(output))

        actualizacion.nuevos = nuevos
        actualizacion.modificados = modificados
        actualizacion.rechazados = rechazados
        actualizacion.save()

        return "Resultado actualizacion completo"

@app.task
def build_resultado_actualizacion_docentes(id):

    location = os.getenv('FTP_SICAN_2018')
    splitted_url = urlparse.urlparse(location)
    actualizacion = models.ActualizacionDocentes.objects.get(id=id)

    ftp = ftplib.FTP()

    input = io.BytesIO()

    try:
        ftp.connect(splitted_url.hostname,int(splitted_url.port))
        ftp.login(splitted_url.username,splitted_url.password)
        ftp.retrbinary('RETR /' + actualizacion.file.name, input.write)
        input.seek(0)
    except:
        return 'Error FTP'
    else:
        ftp.quit()
        wb = openpyxl.load_workbook(input, use_iterators=True)
        ws = wb.get_sheet_by_name('DOCENTES')

        nuevos = 0
        modificados = 0
        rechazados = 0

        proceso = "SICAN-ACTUALIZACIÓN-DOCENTES"

        titulos = [
            'Departamento', 'Cupos', 'Utilizados', 'Registrados', 'Disponibles', 'Cierre', 'Diplomado', 'Municipio',
            'Estado', 'Sede', 'Docente', 'Telefono', 'Registro', 'Aprobado', 'Resultado'
        ]

        formatos = [
            'General', 'General', 'General', 'General', 'General', 'General', 'General', 'General',
            'General', 'General', 'General', 'General', 'General', 'General', 'General'
        ]

        ancho_columnas = [20, 40, 40, 40, 40, 40, 40, 40,
                          40, 40, 40, 40, 40, 40, 40,]

        contenidos = []

        for fila in ws.iter_rows(row_offset=1):

            resultado = ''

            if fila[0].value != '' and fila[0].value != None:

                try:
                    departamento = models.Departamentos.objects.get(alias_simec = fila[0].value)
                except:
                    resultado = 'No existen departamentos con el nombre ingresado'
                    rechazados += 1
                else:

                    if fila[7].value != '' and fila[7].value != None:
                        try:
                            municipio = models.Municipios.objects.filter(departamento = departamento).get(alias_simec = fila[7].value)
                        except:
                            resultado = 'No existen municipios con el nombre ingresado'
                            rechazados += 1
                        else:
                            if fila[10].value != '' and fila[10].value != None:

                                cedula = fila[10].value.split(' - ')[0]
                                nombre = fila[10].value.split(' - ')[1]

                                if models.Docentes.objects.filter(cedula = cedula).count() == 1:
                                    models.Docentes.objects.filter(cedula = cedula).update(
                                        municipio = municipio,
                                        estrategia = models.Estrategias.objects.get(nombre = fila[6].value),
                                        sede = fila[9].value,
                                        nombre = nombre,
                                        telefono = fila[11].value,
                                        registro = fila[12].value,
                                        estado = fila[13].value
                                    )

                                    resultado = 'Docente actualizado'
                                    modificados += 1
                                else:
                                    models.Docentes.objects.create(
                                        municipio=municipio,
                                        estrategia=models.Estrategias.objects.get(nombre=fila[6].value),
                                        sede=fila[9].value,
                                        cedula=cedula,
                                        nombre=nombre,
                                        telefono=fila[11].value,
                                        registro=fila[12].value,
                                        estado=fila[13].value
                                    )
                                    resultado = 'Docente creado'
                                    nuevos += 1

                            else:
                                resultado = 'No hay información del docente'
                                rechazados += 1





            contenidos.append([
                fila[0].value,
                fila[1].value,
                fila[2].value,
                fila[3].value,
                fila[4].value,
                fila[5].value,
                fila[6].value,
                fila[7].value,
                fila[8].value,
                fila[9].value,
                fila[10].value,
                fila[11].value,
                fila[12].value,
                fila[13].value,
                resultado
            ])

        output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, 'Actualiación de docentes', actualizacion.creation,
                                   actualizacion.usuario_creacion, proceso)

        filename = str(actualizacion.id) + '.xlsx'
        actualizacion.result.save(filename, File(output))

        actualizacion.nuevos = nuevos
        actualizacion.modificados = modificados
        actualizacion.rechazados = rechazados
        actualizacion.save()

        return "Resultado actualizacion completo"

@app.task
def build_ruteo(id, region_id):
    reporte = models_reportes.Reportes.objects.get(id = id)
    proceso = "SICAN-RUTEO"
    region = models.Regiones.objects.get(id = region_id)

    titulos = ['Consecutivo', 'Ruta', 'Contratista', 'Cedula', 'Telefono', 'Email', 'Radicados','Valor promedio Radicado',
               'Docentes', 'Valor promedio Docente','Retoma', 'Valor promedio Retoma', 'Valor entregables ruta']

    formatos = ['0', 'General', 'General','0', 'General', 'General', '0', '"$"#,##0_);("$"#,##0)',
                '0', '"$"#,##0_);("$"#,##0)', '0', '"$"#,##0_);("$"#,##0)', '"$"#,##0_);("$"#,##0)']

    ancho_columnas = [20, 30, 50, 30, 30, 30, 30, 30,
                      30, 30, 30, 30, 30]

    contenidos = []

    i = 0
    for ruta in models.Rutas.objects.filter(region=region).order_by('nombre'):
        i += 1
        actividades_json = json.loads(ruta.actividades_json)
        retoma = 0

        for entregable in models.Entregables.objects.filter(tipo='ruta&estrategia'):
            retoma = actividades_json['entregable_{0}'.format(entregable.id)]

        docentes = actividades_json['componente_' + str(models.Componentes.objects.get(numero='2').id)]

        contenidos.append([
            int(i),
            ruta.nombre,
            ruta.contrato.contratista.get_full_name(),
            ruta.contrato.contratista.cedula,
            str(ruta.contrato.contratista.celular),
            ruta.contrato.contratista.email,
            models.Radicados.objects.filter(ruta = ruta).count(),
            ruta.get_valor_promedio_radicados(),
            docentes,
            ruta.get_valor_promedio_docentes(),
            retoma,
            ruta.get_valor_promedio_retomas(),
            ruta.get_valor_entregables_ruta()
        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)




    titulos = ['Consecutivo', 'Ruta', 'Contratista', 'Cedula', 'Radicado', 'Departamento', 'Cod Departamento', 'Municipio', 'Cod Municipio',
               'Nombre IE', 'Codigo Dane Sede', 'Nombre Sede', 'Ubicación sede']

    formatos = ['0', 'General', 'General', '0', '0', 'General', '0','General', '0',
                'General', '0', 'General', 'General']

    ancho_columnas = [20, 30, 50, 30, 30, 30, 30, 30, 30,
                      30, 30, 30, 30]

    contenidos = []

    i = 0
    for radicado in models.Radicados.objects.filter(ruta__region = region).exclude(ruta = None).order_by('ruta__nombre'):
        i += 1

        contenidos.append([
            int(i),
            radicado.ruta.nombre,
            radicado.ruta.contrato.contratista.get_full_name(),
            radicado.ruta.contrato.contratista.cedula,
            radicado.numero,
            radicado.municipio.departamento.nombre,
            radicado.municipio.departamento.numero,
            radicado.municipio.nombre,
            radicado.municipio.numero,
            radicado.nombre_ie,
            radicado.dane_sede,
            radicado.nombre_sede,
            radicado.ubicacion
        ])



    output2 = construir_reporte_pagina(output,'Hoja2',titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)



    titulos = ['Consecutivo', 'Ruta', 'Contratista', 'Cedula', 'Cedula Docente', 'Nombre', 'Municipio', 'Departamento',
               'Diplomado', 'Grupo']

    formatos = ['0', 'General', 'General', '0', '0', 'General', 'General', 'General',
                'General', 'General']

    ancho_columnas = [20, 30, 50, 30, 30, 30, 30, 30,
                      30, 30]

    contenidos = []

    i = 0
    for docente in models.Docentes.objects.filter(grupo__ruta__region=region):
        i += 1

        contenidos.append([
            int(i),
            docente.grupo.ruta.nombre,
            docente.grupo.ruta.contrato.contratista.get_full_name(),
            docente.grupo.ruta.contrato.contratista.cedula,
            docente.cedula,
            docente.nombre,
            docente.municipio.nombre,
            docente.municipio.departamento.nombre,
            docente.estrategia.nombre,
            docente.grupo.get_nombre_grupo()
        ])

    output = construir_reporte_pagina(output2, 'Hoja3', titulos, contenidos, formatos, ancho_columnas, reporte.nombre,
                                       reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))


    return "Ruteo generado: " + filename

@app.task
def build_red_acceso(red_id):

    red = models.Red.objects.get(id = red_id)

    output = BytesIO()
    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/red_acceso.xlsx')
    ws = wb.get_sheet_by_name('RED')

    i = 1

    for taller in models.RelatoriaTallerAdministratic.objects.filter(red=red):

        if taller.tipo == 'Sican':

            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo = 'relatoria_taller_administratic').id,
                taller.id
            )
        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='relatoria_taller_administratic').id,
                taller.id
            )

        ws['B' + str(i+2)] = red.consecutivo
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format(red.region.numero)
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Taller Administratic'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1

    for taller in models.RelatoriaTallerContenidosEducativos.objects.filter(red=red):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='relatoria_taller_contenidos_educativos').id,
                taller.id
            )

        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='relatoria_taller_contenidos_educativos').id,
                taller.id
            )

        ws['B' + str(i+2)] = red.consecutivo
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format(red.region.numero)
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Taller de Contenidos'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1

    for taller in models.RelatoriaTallerRAEE.objects.filter(red=red):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='relatoria_taller_raee').id,
                taller.id
            )

        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='relatoria_taller_raee').id,
                taller.id
            )

        ws['B' + str(i+2)] = red.consecutivo
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format(red.region.numero)
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Taller RAEE'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1

    for taller in models.RelatoriaTallerApertura.objects.filter(red=red):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='relatoria_taller_apertura').id,
                taller.id
            )
        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='relatoria_taller_apertura').id,
                taller.id
            )

        ws['B' + str(i+2)] = red.consecutivo
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format(red.region.numero)
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Taller de Apertura'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1

    for taller in models.DocumentoLegalizacionTerminales.objects.filter(red=red):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales').id,
                taller.id
            )
        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales').id,
                taller.id
            )

        ws['B' + str(i+2)] = red.consecutivo
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format(red.region.numero)
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Legalización Estudiantes'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1



    for taller in models.DocumentoLegalizacionTerminalesValle1.objects.filter(red=red):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales_v1').id,
                taller.id
            )
        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales_v1').id,
                taller.id
            )

        ws['B' + str(i+2)] = red.consecutivo
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format(red.region.numero)
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Legalización Estudiantes (Valle 1)'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1





    for taller in models.DocumentoLegalizacionTerminalesValle2.objects.filter(red=red):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales_v2').id,
                taller.id
            )
        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales_v2').id,
                taller.id
            )

        ws['B' + str(i+2)] = red.consecutivo
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format(red.region.numero)
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Legalización Estudiantes (Valle 2)'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1



    for encuesta in models.EncuestaMonitoreo.objects.filter(red=red):

        if encuesta.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='encuesta_monitoreo').id,
                encuesta.id
            )
        elif encuesta.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='encuesta_monitoreo').id,
                encuesta.id
            )

        ws['B' + str(i+2)] = red.consecutivo
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format(red.region.numero)
        ws['E' + str(i + 2)] = encuesta.radicado.numero
        ws['H' + str(i + 2)] = 'Encuesta de Monitoreo'
        ws['I' + str(i + 2)] = encuesta.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = encuesta.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = encuesta.get_observaciones()

        i += 1



    for retoma in models.Retoma.objects.filter(red=red):
        if retoma.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='retoma').id,
                retoma.id
            )
        elif retoma.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                red.region.id,
                red.id,
                models.Entregables.objects.get(modelo='retoma').id,
                retoma.id
            )


        try:
            radicado_int = int(retoma.radicado)
        except:
            radicado_int = retoma.radicado

        ws['B' + str(i+2)] = red.consecutivo
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format(red.region.numero)
        ws['E' + str(i + 2)] = radicado_int
        ws['H' + str(i + 2)] = 'Legalización Retoma'
        ws['I' + str(i + 2)] = retoma.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = retoma.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = retoma.get_observaciones()

        i += 1


    filename = str(red.id) + '.xlsx'
    wb.save(output)
    red.file.save(filename, File(output))


    return "Red acceso: " + filename


@app.task
def build_red_acceso_informacion(reporte_id):


    output = BytesIO()
    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/red_acceso.xlsx')
    ws = wb.get_sheet_by_name('RED')

    i = 1

    for taller in models.RelatoriaTallerAdministratic.objects.filter(red=None):

        if taller.tipo == 'Sican':

            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo = 'relatoria_taller_administratic').id,
                taller.id
            )
        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='relatoria_taller_administratic').id,
                taller.id
            )

        ws['B' + str(i+2)] = ""
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format("")
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Taller Administratic'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1

    for taller in models.RelatoriaTallerContenidosEducativos.objects.filter(red=None):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='relatoria_taller_contenidos_educativos').id,
                taller.id
            )

        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='relatoria_taller_contenidos_educativos').id,
                taller.id
            )

        ws['B' + str(i+2)] = ""
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format("")
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Taller de Contenidos'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1

    for taller in models.RelatoriaTallerRAEE.objects.filter(red=None):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='relatoria_taller_raee').id,
                taller.id
            )

        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='relatoria_taller_raee').id,
                taller.id
            )

        ws['B' + str(i+2)] = ""
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format("")
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Taller RAEE'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1

    for taller in models.RelatoriaTallerApertura.objects.filter(red=None):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='relatoria_taller_apertura').id,
                taller.id
            )
        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='relatoria_taller_apertura').id,
                taller.id
            )

        ws['B' + str(i+2)] = ""
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format("")
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Taller de Apertura'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1

    for taller in models.DocumentoLegalizacionTerminales.objects.filter(red=None):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales').id,
                taller.id
            )
        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales').id,
                taller.id
            )

        ws['B' + str(i+2)] = ""
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format("")
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Legalización Estudiantes'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1



    for taller in models.DocumentoLegalizacionTerminalesValle1.objects.filter(red=None):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales_v1').id,
                taller.id
            )
        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales_v1').id,
                taller.id
            )

        ws['B' + str(i+2)] = ""
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format("")
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Legalización Estudiantes (Valle 1)'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1





    for taller in models.DocumentoLegalizacionTerminalesValle2.objects.filter(red=None):

        if taller.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales_v2').id,
                taller.id
            )
        elif taller.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='documento_legalizacion_terminales_v2').id,
                taller.id
            )

        ws['B' + str(i+2)] = ""
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format("")
        ws['E' + str(i + 2)] = taller.radicado.numero
        ws['H' + str(i + 2)] = 'Legalización Estudiantes (Valle 2)'
        ws['I' + str(i + 2)] = taller.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = taller.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = taller.get_observaciones()

        i += 1



    for encuesta in models.EncuestaMonitoreo.objects.filter(red=None):

        if encuesta.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='encuesta_monitoreo').id,
                encuesta.id
            )
        elif encuesta.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='encuesta_monitoreo').id,
                encuesta.id
            )

        ws['B' + str(i+2)] = ""
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format("")
        ws['E' + str(i + 2)] = encuesta.radicado.numero
        ws['H' + str(i + 2)] = 'Encuesta de Monitoreo'
        ws['I' + str(i + 2)] = encuesta.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = encuesta.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = encuesta.get_observaciones()

        i += 1



    for retoma in models.Retoma.objects.filter(red=None):
        if retoma.tipo == 'Sican':
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='retoma').id,
                retoma.id
            )
        elif retoma.tipo == 'Lupaap':
            link = 'Lupaap'

        else:
            link = 'https://sican.asoandes.org/cpe_2018/red/{0}/ver/{1}/actividades/{2}/calificar/{3}/'.format(
                "",
                "",
                models.Entregables.objects.get(modelo='retoma').id,
                retoma.id
            )


        try:
            radicado_int = int(retoma.radicado)
        except:
            radicado_int = retoma.radicado

        ws['B' + str(i+2)] = ""
        ws['D' + str(i + 2)] = 'Asoandes RG{0}'.format("")
        ws['E' + str(i + 2)] = radicado_int
        ws['H' + str(i + 2)] = 'Legalización Retoma'
        ws['I' + str(i + 2)] = retoma.fecha.strftime('%d/%m/%Y')
        ws['K' + str(i + 2)] = link
        ws['L' + str(i + 2)] = retoma.ruta.contrato.contratista.get_full_name()
        ws['M' + str(i + 2)] = retoma.get_observaciones()

        i += 1


    filename = str(reporte_id) + '.xlsx'
    wb.save(output)

    reporte = models_reportes.Reportes.objects.get(id=reporte_id)

    reporte.file.save(filename, File(output))


    return "Informe RED: " + filename




@app.task
def build_red_formacion(red_id):

    red = models.Red.objects.get(id = red_id)

    output = BytesIO()
    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/red_formacion.xlsx')
    ws_lupaap = wb.get_sheet_by_name('LUPAAP')
    ws_sican = wb.get_sheet_by_name('SICAN')

    imagen = Image(settings.STATICFILES_DIRS[0] + '/img/cabecera_red.png')
    imagen.drawing.top = 0
    imagen.drawing.left = 0
    ws_lupaap.add_image(imagen)
    ws_sican.add_image(imagen)


    # ----------------------------------------------- INNOVATIC LUPAAP ---------------------------------------------------------

    asistencias_innovatic = models.ListadoAsistencia.objects.filter(red=red, grupo__estrategia__nombre = 'InnovaTIC', tipo = 'Lupaap')
    ple_innovatic = models.ProductoFinalPle.objects.filter(red=red, grupo__estrategia__nombre = 'InnovaTIC', tipo = 'Lupaap')

    ids_docentes_innovatic = list(asistencias_innovatic.values_list('docentes__id',flat=True).distinct()) + \
                             list(ple_innovatic.values_list('docentes__id',flat=True).distinct())


    entregables_innovatic = {

        'K': {'numero':3, 'query': asistencias_innovatic},
        'L': {'numero':4, 'query': asistencias_innovatic},
        'O': {'numero':5, 'query': asistencias_innovatic},
        'P': {'numero':6, 'query': asistencias_innovatic},
        'Q': {'numero':7, 'query': asistencias_innovatic},
        'R': {'numero':8, 'query': asistencias_innovatic},
        'S': {'numero':9, 'query': asistencias_innovatic},
        'T': {'numero':10, 'query': asistencias_innovatic},
        'U': {'numero':11, 'query': asistencias_innovatic},

        'V': {'numero':15, 'query': asistencias_innovatic},
        'W': {'numero':16, 'query': asistencias_innovatic},
        'X': {'numero':17, 'query': asistencias_innovatic},
        'AB': {'numero':18, 'query': asistencias_innovatic},
        'AC': {'numero':19, 'query': asistencias_innovatic},

        'AD': {'numero':23, 'query': asistencias_innovatic},
        'AE': {'numero':24, 'query': asistencias_innovatic},
        'AF': {'numero':25, 'query': asistencias_innovatic},
        'AG': {'numero':26, 'query': asistencias_innovatic},
        'AH': {'numero':27, 'query': asistencias_innovatic},
        'AI': {'numero':28, 'query': asistencias_innovatic},

        'AJ': {'numero':32, 'query': asistencias_innovatic},
        'AK': {'numero':33, 'query': asistencias_innovatic},
        'AO': {'numero':34, 'query': asistencias_innovatic},
        'AP': {'numero':35, 'query': asistencias_innovatic},
        'AQ': {'numero':36, 'query': asistencias_innovatic},
        'AR': {'numero':37, 'query': asistencias_innovatic},
        'AS': {'numero':38, 'query': asistencias_innovatic},
        'BQ': {'numero':30, 'query': ple_innovatic},
    }

    i = 1

    for docente in models.Docentes.objects.filter(id__in = ids_docentes_innovatic):

        ws_lupaap['A' + str(i + 9)] = i
        ws_lupaap['B' + str(i + 9)] = docente.grupo.ruta.region.nombre if docente.grupo != None else ''
        ws_lupaap['C' + str(i + 9)] = docente.municipio.departamento.nombre
        ws_lupaap['D' + str(i + 9)] = docente.municipio.nombre
        ws_lupaap['E' + str(i + 9)] = docente.cedula
        ws_lupaap['F' + str(i + 9)] = docente.nombre
        ws_lupaap['G' + str(i + 9)] = docente.telefono
        ws_lupaap['H' + str(i + 9)] = docente.grupo.estrategia.nombre
        ws_lupaap['I' + str(i + 9)] = docente.grupo.ruta.contrato.contratista.cedula if docente.grupo != None else ''
        ws_lupaap['J' + str(i + 9)] = docente.grupo.ruta.contrato.contratista.get_full_name() if docente.grupo != None else ''

        for key in entregables_innovatic.keys():
            entregable = models.Entregables.objects.get(momento__estrategia__nombre = 'InnovaTIC',numero = entregables_innovatic[key]['numero'])

            if entregables_innovatic[key]['numero'] == 30:
                ws_lupaap['BP' + str(i + 9)] = 'PLE'

            try:
                ws_lupaap[str(key) + str(i + 9)] = entregables_innovatic[key]['query'].filter(entregable = entregable).get(docentes__id = docente.id).fecha.strftime('%d/%m/%Y')
            except:
                pass
        i += 1

    # ------------------------------------------------ RURALTIC LUPAAP --------------------------------------------------------

    asistencias_ruraltic = models.ListadoAsistencia.objects.filter(red=red, grupo__estrategia__nombre='RuralTIC',tipo='Lupaap')
    repositorio_ruraltic = models.RepositorioActividades.objects.filter(red=red,grupo__estrategia__nombre='RuralTIC',tipo='Lupaap')

    ids_docentes_ruraltic = list(asistencias_ruraltic.values_list('docentes__id', flat=True).distinct()) + \
                            list(repositorio_ruraltic.values_list('docentes__id', flat=True).distinct())

    entregables_ruraltic = {

        'K': {'numero': 3, 'query': asistencias_ruraltic},
        'L': {'numero': 4, 'query': asistencias_ruraltic},

        'O': {'numero': 7, 'query': asistencias_ruraltic},
        'P': {'numero': 8, 'query': asistencias_ruraltic},
        'Q': {'numero': 9, 'query': asistencias_ruraltic},
        'R': {'numero': 10, 'query': asistencias_ruraltic},
        'S': {'numero': 11, 'query': asistencias_ruraltic},
        'T': {'numero': 12, 'query': asistencias_ruraltic},
        'U': {'numero': 13, 'query': asistencias_ruraltic},
        'V': {'numero': 14, 'query': asistencias_ruraltic},
        'W': {'numero': 15, 'query': asistencias_ruraltic},
        'X': {'numero': 16, 'query': asistencias_ruraltic},

        'AB': {'numero': 20, 'query': asistencias_ruraltic},
        'AC': {'numero': 21, 'query': asistencias_ruraltic},
        'AD': {'numero': 22, 'query': asistencias_ruraltic},
        'AE': {'numero': 23, 'query': asistencias_ruraltic},
        'AF': {'numero': 24, 'query': asistencias_ruraltic},
        'AG': {'numero': 25, 'query': asistencias_ruraltic},
        'AH': {'numero': 26, 'query': asistencias_ruraltic},
        'AI': {'numero': 27, 'query': asistencias_ruraltic},
        'AJ': {'numero': 28, 'query': asistencias_ruraltic},
        'AK': {'numero': 29, 'query': asistencias_ruraltic},

        'AO': {'numero': 33, 'query': asistencias_ruraltic},
        'AP': {'numero': 34, 'query': asistencias_ruraltic},
        'AQ': {'numero': 35, 'query': asistencias_ruraltic},
        'AR': {'numero': 36, 'query': asistencias_ruraltic},
        'AS': {'numero': 37, 'query': asistencias_ruraltic},
        'AT': {'numero': 38, 'query': asistencias_ruraltic},
        'AU': {'numero': 39, 'query': asistencias_ruraltic},
        'AV': {'numero': 40, 'query': asistencias_ruraltic},
        'AW': {'numero': 41, 'query': asistencias_ruraltic},
        'AX': {'numero': 42, 'query': asistencias_ruraltic},
        'AY': {'numero': 43, 'query': asistencias_ruraltic},
        'AZ': {'numero': 44, 'query': asistencias_ruraltic},
        'BA': {'numero': 45, 'query': asistencias_ruraltic},
        'BB': {'numero': 46, 'query': asistencias_ruraltic},
        'BC': {'numero': 47, 'query': asistencias_ruraltic},
        'BD': {'numero': 48, 'query': asistencias_ruraltic},

        'BH': {'numero': 53, 'query': asistencias_ruraltic},
        'BI': {'numero': 54, 'query': asistencias_ruraltic},

        'BQ': {'numero': 50, 'query': repositorio_ruraltic},
    }

    for docente in models.Docentes.objects.filter(id__in=ids_docentes_ruraltic):

        ws_lupaap['A' + str(i + 9)] = i
        ws_lupaap['B' + str(i + 9)] = docente.grupo.ruta.region.nombre if docente.grupo != None else ''
        ws_lupaap['C' + str(i + 9)] = docente.municipio.departamento.nombre
        ws_lupaap['D' + str(i + 9)] = docente.municipio.nombre
        ws_lupaap['E' + str(i + 9)] = docente.cedula
        ws_lupaap['F' + str(i + 9)] = docente.nombre
        ws_lupaap['G' + str(i + 9)] = docente.telefono
        ws_lupaap['H' + str(i + 9)] = docente.grupo.estrategia.nombre if docente.grupo != None else ''
        ws_lupaap['I' + str(i + 9)] = docente.grupo.ruta.contrato.contratista.cedula if docente.grupo != None else ''
        ws_lupaap['J' + str(i + 9)] = docente.grupo.ruta.contrato.contratista.get_full_name() if docente.grupo != None else ''

        for key in entregables_ruraltic.keys():
            entregable = models.Entregables.objects.get(momento__estrategia__nombre='RuralTIC',
                                                        numero=entregables_ruraltic[key]['numero'])

            if entregables_ruraltic[key]['numero'] == 50:
                ws_lupaap['BP' + str(i + 9)] = 'Repositorio APA'

            try:
                ws_lupaap[str(key) + str(i + 9)] = entregables_ruraltic[key]['query'].filter(
                    entregable=entregable).get(docentes__id=docente.id).fecha.strftime('%d/%m/%Y')
            except:
                pass
        i += 1

    # ----------------------------------------------- INNOVATIC SICAN ---------------------------------------------------------

    asistencias_innovatic = models.ListadoAsistencia.objects.filter(red=red, grupo__estrategia__nombre='InnovaTIC',tipo__in=['Sican',''])
    ple_innovatic = models.ProductoFinalPle.objects.filter(red=red, grupo__estrategia__nombre='InnovaTIC',tipo__in=['Sican',''])

    ids_docentes_innovatic = list(asistencias_innovatic.values_list('docentes__id', flat=True).distinct()) + \
                             list(ple_innovatic.values_list('docentes__id', flat=True).distinct())

    entregables_innovatic = {

        'K': {'numero': 3, 'query': asistencias_innovatic},
        'L': {'numero': 4, 'query': asistencias_innovatic},
        'O': {'numero': 5, 'query': asistencias_innovatic},
        'P': {'numero': 6, 'query': asistencias_innovatic},
        'Q': {'numero': 7, 'query': asistencias_innovatic},
        'R': {'numero': 8, 'query': asistencias_innovatic},
        'S': {'numero': 9, 'query': asistencias_innovatic},
        'T': {'numero': 10, 'query': asistencias_innovatic},
        'U': {'numero': 11, 'query': asistencias_innovatic},

        'V': {'numero': 15, 'query': asistencias_innovatic},
        'W': {'numero': 16, 'query': asistencias_innovatic},
        'X': {'numero': 17, 'query': asistencias_innovatic},
        'AB': {'numero': 18, 'query': asistencias_innovatic},
        'AC': {'numero': 19, 'query': asistencias_innovatic},

        'AD': {'numero': 23, 'query': asistencias_innovatic},
        'AE': {'numero': 24, 'query': asistencias_innovatic},
        'AF': {'numero': 25, 'query': asistencias_innovatic},
        'AG': {'numero': 26, 'query': asistencias_innovatic},
        'AH': {'numero': 27, 'query': asistencias_innovatic},
        'AI': {'numero': 28, 'query': asistencias_innovatic},

        'AJ': {'numero': 32, 'query': asistencias_innovatic},
        'AK': {'numero': 33, 'query': asistencias_innovatic},
        'AO': {'numero': 34, 'query': asistencias_innovatic},
        'AP': {'numero': 35, 'query': asistencias_innovatic},
        'AQ': {'numero': 36, 'query': asistencias_innovatic},
        'AR': {'numero': 37, 'query': asistencias_innovatic},
        'AS': {'numero': 38, 'query': asistencias_innovatic},
        'BQ': {'numero': 30, 'query': ple_innovatic},
    }

    i = 1

    for docente in models.Docentes.objects.filter(id__in=ids_docentes_innovatic):

        ws_sican['A' + str(i + 9)] = i
        ws_sican['B' + str(i + 9)] = docente.grupo.ruta.region.nombre if docente.grupo != None else ''
        ws_sican['C' + str(i + 9)] = docente.municipio.departamento.nombre
        ws_sican['D' + str(i + 9)] = docente.municipio.nombre
        ws_sican['E' + str(i + 9)] = docente.cedula
        ws_sican['F' + str(i + 9)] = docente.nombre
        ws_sican['G' + str(i + 9)] = docente.telefono
        ws_sican['H' + str(i + 9)] = docente.grupo.estrategia.nombre if docente.grupo != None else ''
        ws_sican['I' + str(i + 9)] = docente.grupo.ruta.contrato.contratista.cedula if docente.grupo != None else ''
        ws_sican['J' + str(i + 9)] = docente.grupo.ruta.contrato.contratista.get_full_name() if docente.grupo != None else ''

        for key in entregables_innovatic.keys():
            if entregables_innovatic[key]['numero'] == 30:
                ws_lupaap['BP' + str(i + 9)] = 'PLE'
            entregable = models.Entregables.objects.get(momento__estrategia__nombre='InnovaTIC',
                                                        numero=entregables_innovatic[key]['numero'])
            try:
                ws_sican[str(key) + str(i + 9)] = entregables_innovatic[key]['query'].filter(
                    entregable=entregable).get(docentes__id=docente.id).fecha.strftime('%d/%m/%Y')
            except:
                pass
        i += 1

    # ------------------------------------------------ RURALTIC SICAN --------------------------------------------------------

    asistencias_ruraltic = models.ListadoAsistencia.objects.filter(red=red, grupo__estrategia__nombre='RuralTIC',tipo__in=['Sican',''])
    repositorio_ruraltic = models.RepositorioActividades.objects.filter(red=red,grupo__estrategia__nombre='RuralTIC',tipo__in=['Sican',''])

    ids_docentes_ruraltic = list(asistencias_ruraltic.values_list('docentes__id', flat=True).distinct()) + \
                            list(repositorio_ruraltic.values_list('docentes__id', flat=True).distinct())

    entregables_ruraltic = {

        'K': {'numero': 3, 'query': asistencias_ruraltic},
        'L': {'numero': 4, 'query': asistencias_ruraltic},

        'O': {'numero': 7, 'query': asistencias_ruraltic},
        'P': {'numero': 8, 'query': asistencias_ruraltic},
        'Q': {'numero': 9, 'query': asistencias_ruraltic},
        'R': {'numero': 10, 'query': asistencias_ruraltic},
        'S': {'numero': 11, 'query': asistencias_ruraltic},
        'T': {'numero': 12, 'query': asistencias_ruraltic},
        'U': {'numero': 13, 'query': asistencias_ruraltic},
        'V': {'numero': 14, 'query': asistencias_ruraltic},
        'W': {'numero': 15, 'query': asistencias_ruraltic},
        'X': {'numero': 16, 'query': asistencias_ruraltic},

        'AB': {'numero': 20, 'query': asistencias_ruraltic},
        'AC': {'numero': 21, 'query': asistencias_ruraltic},
        'AD': {'numero': 22, 'query': asistencias_ruraltic},
        'AE': {'numero': 23, 'query': asistencias_ruraltic},
        'AF': {'numero': 24, 'query': asistencias_ruraltic},
        'AG': {'numero': 25, 'query': asistencias_ruraltic},
        'AH': {'numero': 26, 'query': asistencias_ruraltic},
        'AI': {'numero': 27, 'query': asistencias_ruraltic},
        'AJ': {'numero': 28, 'query': asistencias_ruraltic},
        'AK': {'numero': 29, 'query': asistencias_ruraltic},

        'AO': {'numero': 33, 'query': asistencias_ruraltic},
        'AP': {'numero': 34, 'query': asistencias_ruraltic},
        'AQ': {'numero': 35, 'query': asistencias_ruraltic},
        'AR': {'numero': 36, 'query': asistencias_ruraltic},
        'AS': {'numero': 37, 'query': asistencias_ruraltic},
        'AT': {'numero': 38, 'query': asistencias_ruraltic},
        'AU': {'numero': 39, 'query': asistencias_ruraltic},
        'AV': {'numero': 40, 'query': asistencias_ruraltic},
        'AW': {'numero': 41, 'query': asistencias_ruraltic},
        'AX': {'numero': 42, 'query': asistencias_ruraltic},
        'AY': {'numero': 43, 'query': asistencias_ruraltic},
        'AZ': {'numero': 44, 'query': asistencias_ruraltic},
        'BA': {'numero': 45, 'query': asistencias_ruraltic},
        'BB': {'numero': 46, 'query': asistencias_ruraltic},
        'BC': {'numero': 47, 'query': asistencias_ruraltic},
        'BD': {'numero': 48, 'query': asistencias_ruraltic},

        'BH': {'numero': 53, 'query': asistencias_ruraltic},
        'BI': {'numero': 54, 'query': asistencias_ruraltic},

        'BQ': {'numero': 50, 'query': repositorio_ruraltic},
    }


    for docente in models.Docentes.objects.filter(id__in=ids_docentes_ruraltic):

        ws_sican['A' + str(i + 9)] = i
        ws_sican['B' + str(i + 9)] = docente.grupo.ruta.region.nombre if docente.grupo != None else ''
        ws_sican['C' + str(i + 9)] = docente.municipio.departamento.nombre
        ws_sican['D' + str(i + 9)] = docente.municipio.nombre
        ws_sican['E' + str(i + 9)] = docente.cedula
        ws_sican['F' + str(i + 9)] = docente.nombre
        ws_sican['G' + str(i + 9)] = docente.telefono
        ws_sican['H' + str(i + 9)] = docente.grupo.estrategia.nombre if docente.grupo != None else ''
        ws_sican['I' + str(i + 9)] = docente.grupo.ruta.contrato.contratista.cedula if docente.grupo != None else ''
        ws_sican['J' + str(i + 9)] = docente.grupo.ruta.contrato.contratista.get_full_name() if docente.grupo != None else ''

        for key in entregables_ruraltic.keys():
            entregable = models.Entregables.objects.get(momento__estrategia__nombre='RuralTIC',
                                                        numero=entregables_ruraltic[key]['numero'])
            if entregables_ruraltic[key]['numero'] == 50:
                ws_lupaap['BP' + str(i + 9)] = 'Repositorio APA'
            try:
                ws_sican[str(key) + str(i + 9)] = entregables_ruraltic[key]['query'].filter(
                    entregable=entregable).get(docentes__id=docente.id).fecha.strftime('%d/%m/%Y')
            except:
                pass
        i += 1


    filename = str(red.id) + '.xlsx'
    wb.save(output)
    red.file.save(filename, File(output))


    return "Red formacion: " + filename

@app.task
def calcular_actualizacion(actualizacion_id):

    actualizacion = models.ActualizacionLupaap.objects.get(id = actualizacion_id)
    output = BytesIO()

    contador = 0
    soportes = {}

    if not bool(actualizacion.tablero_control_json):

        wb_tablero = openpyxl.load_workbook(filename=actualizacion.tablero_control.path)
        ws_tablero = wb_tablero.get_sheet_by_name('Docentes')

        entregables_innovatic = {

            'N': {'numero': 3},
            'O': {'numero': 4},
            'P': {'numero': 5},
            'Q': {'numero': 6},
            'R': {'numero': 7},
            'S': {'numero': 8},
            'T': {'numero': 9},
            'U': {'numero': 10},
            'V': {'numero': 11},

            'Z': {'numero': 15},
            'AA': {'numero': 16},
            'AB': {'numero': 17},
            'AC': {'numero': 18},
            'AD': {'numero': 19},

            'AM': {'numero': 23},
            'AN': {'numero': 24},
            'AO': {'numero': 25},
            'AP': {'numero': 26},
            'AQ': {'numero': 27},
            'AR': {'numero': 28},

            'AZ': {'numero': 32},
            'BA': {'numero': 33},
            'BB': {'numero': 34},
            'BC': {'numero': 35},
            'BD': {'numero': 36},
            'BE': {'numero': 37},
            'BF': {'numero': 38}
        }

        entregables_ruraltic = {

            'N': {'numero': 3},
            'O': {'numero': 4},

            'Z': {'numero': 7},
            'AA': {'numero': 8},
            'AB': {'numero': 9},
            'AC': {'numero': 10},
            'AD': {'numero': 11},
            'AE': {'numero': 12},
            'AF': {'numero': 13},
            'AG': {'numero': 14},
            'AH': {'numero': 15},
            'AI': {'numero': 16},

            'AM': {'numero': 20},
            'AN': {'numero': 21},
            'AO': {'numero': 22},
            'AP': {'numero': 23},
            'AQ': {'numero': 24},
            'AR': {'numero': 25},
            'AS': {'numero': 26},
            'AT': {'numero': 27},
            'AU': {'numero': 28},
            'AV': {'numero': 29},

            'AZ': {'numero': 33},
            'BA': {'numero': 34},
            'BB': {'numero': 35},
            'BC': {'numero': 36},
            'BD': {'numero': 37},
            'BE': {'numero': 38},
            'BF': {'numero': 39},
            'BG': {'numero': 40},
            'BH': {'numero': 41},
            'BI': {'numero': 42},
            'BJ': {'numero': 43},
            'BK': {'numero': 44},
            'BL': {'numero': 45},
            'BM': {'numero': 46},
            'BN': {'numero': 47},
            'BO': {'numero': 48},

            'BS': {'numero': 53},
            'BT': {'numero': 54},
        }

        data = {
            'InnovaTIC': {},
            'RuralTIC': {}
        }

        for i in range(2,ws_tablero.max_row+1):
            if ws_tablero['H'+str(i)].value == 'InnovaTIC':
                entregables = entregables_innovatic
                diplomado = 'InnovaTIC'
            elif ws_tablero['H' + str(i)].value == 'RuralTIC':
                entregables = entregables_ruraltic
                diplomado = 'RuralTIC'
            else:
                entregables = {}
                diplomado = ''

            for key in entregables.keys():
                entregable = models.Entregables.objects.get(momento__estrategia__nombre=diplomado,numero=entregables[key]['numero'])
                nivel = entregable.momento.nombre
                numero = entregables[key]['numero']

                if nivel not in data[diplomado].keys():
                    data[diplomado][nivel] = {}

                if numero not in data[diplomado][nivel].keys():
                    data[diplomado][nivel][numero] = {}

                value = ws_tablero['{0}{1}'.format(key, str(i))].value
                cedula = ws_tablero['E' + str(i)].value

                if value == '' or value == None:
                    pass
                else:
                    if cedula not in data[diplomado][nivel].keys():
                        data[diplomado][nivel][numero][cedula] = {}

                    data[diplomado][nivel][numero][cedula] = str(value)

        json_file_tablero = BytesIO(json.dumps(data).encode())


        tablero_json_filename = str(actualizacion.id) + '.json'
        actualizacion.tablero_control_json.save(tablero_json_filename, File(json_file_tablero))

    else:
        with open(actualizacion.tablero_control_json.path) as f:
            data = json.load(f)


    if not bool(actualizacion.informe_lupaap_json):

        wb_lupaap = openpyxl.load_workbook(filename=actualizacion.informe_lupaap.path)
        ws_lupaap = wb_lupaap.get_sheet_by_name('Reporte')
        letras = ['Q','U','Y','AC','AG','AK','AO','AS','AW','BA','BE','BI','BM','BQ','BU','BY','CC','CG','CK','CO','CS',
                  'CW','DA','DE','DI','DM','DQ','DU','DY','EC','EG','EK','EO','ES','EW']

        modelos = {
            'InnovaTIC': {
                '1':{
                    '1': '3',
                    '2': '4',
                    '3': '5',
                    '4': '6',
                    '5': '7',
                    '6': '8',
                    '7': '9',
                    '8': '10',
                    '9': '11',
                },
                '2': {
                    '1': '15',
                    '2': '16',
                    '3': '17',
                    '4': '18',
                    '5': '19',
                },
                '3': {
                    '1': '23',
                    '2': '24',
                    '3': '25',
                    '4': '26',
                    '5': '27',
                    '6': '28',
                },
                '4': {
                    '1': '32',
                    '2': '33',
                    '3': '34',
                    '4': '35',
                    '5': '36',
                    '6': '37',
                    '7': '38',
                },
            },
            'RuralTIC': {
                '1': {
                    '1': '3',
                    '2': '4',
                },
                '2': {
                    '1': '7',
                    '2': '8',
                    '3': '9',
                    '4': '10',
                    '5': '11',
                    '6': '12',
                    '7': '13',
                    '8': '14',
                    '9': '15',
                    '10': '16',
                },
                '3': {
                    '1': '20',
                    '2': '21',
                    '3': '22',
                    '4': '23',
                    '5': '24',
                    '6': '25',
                    '7': '26',
                    '8': '27',
                    '9': '28',
                    '10': '29',
                },
                '4': {
                    '1': '33',
                    '2': '34',
                    '3': '35',
                    '4': '36',
                    '5': '37',
                    '6': '38',
                    '7': '39',
                    '8': '40',
                    '9': '41',
                    '10': '42',
                    '11': '43',
                    '12': '44',
                    '13': '45',
                    '14': '46',
                    '15': '47',
                    '16': '48',
                },
                '5': {
                    '1': '53',
                    '2': '54',
                },
            }
        }
        data_lupaap = {
            'InnovaTIC': {
                'Nivel 1':{
                    '3': {},
                    '4': {},
                    '5': {},
                    '6': {},
                    '7': {},
                    '8': {},
                    '9': {},
                    '10': {},
                    '11': {},
                },
                'Nivel 2': {
                    '15': {},
                    '16': {},
                    '17': {},
                    '18': {},
                    '19': {},
                },
                'Nivel 3': {
                    '23': {},
                    '24': {},
                    '25': {},
                    '26': {},
                    '27': {},
                    '28': {},
                },
                'Nivel 4': {
                    '32': {},
                    '33': {},
                    '34': {},
                    '35': {},
                    '36': {},
                    '37': {},
                    '38': {},
                },
            },
            'RuralTIC': {
                'Módulo 1': {
                    '3': {},
                    '4': {},
                },
                'Módulo 2': {
                    '7': {},
                    '8': {},
                    '9': {},
                    '10': {},
                    '11': {},
                    '12': {},
                    '13': {},
                    '14': {},
                    '15': {},
                    '16': {},
                },
                'Módulo 3': {
                    '20': {},
                    '21': {},
                    '22': {},
                    '23': {},
                    '24': {},
                    '25': {},
                    '26': {},
                    '27': {},
                    '28': {},
                    '29': {},
                },
                'Módulo 4': {
                    '33': {},
                    '34': {},
                    '35': {},
                    '36': {},
                    '37': {},
                    '38': {},
                    '39': {},
                    '40': {},
                    '41': {},
                    '42': {},
                    '43': {},
                    '44': {},
                    '45': {},
                    '46': {},
                    '47': {},
                    '48': {},
                },
                'Módulo 5': {
                    '53': {},
                    '54': {},
                },
            }
        }

        for i in range(4,ws_lupaap.max_row+1):
            if ws_lupaap['L'+str(i)].value == 'InnovaTIC':
                diplomado = 'InnovaTIC'
                nivel = ws_lupaap['N'+str(i)].value.replace('Nivel ','')
                sesion = ws_lupaap['O' + str(i)].value.replace('Sesión ', '')
                prev = 'Nivel '
                cedula_lider = ws_lupaap['C' + str(i)].value
            elif ws_lupaap['L' + str(i)].value == 'RuralTIC':
                diplomado = 'RuralTIC'
                nivel = ws_lupaap['N' + str(i)].value.replace('Nivel ', '')
                sesion = ws_lupaap['O' + str(i)].value.replace('Sesión ', '')
                prev = 'Módulo '
                cedula_lider = ws_lupaap['C' + str(i)].value
            else:
                diplomado = ''
                nivel = ''
                sesion = ''
                prev = ''
                cedula_lider = ''

            numero = modelos[diplomado][str(nivel)][str(sesion)]

            for letra in letras:
                cedula = ws_lupaap['{0}{1}'.format(letra,str(i))].value
                if cedula != None and cedula != '':
                    if str(cedula_lider) not in data_lupaap[diplomado][prev + str(nivel)][numero]:
                        data_lupaap[diplomado][prev + str(nivel)][numero][str(cedula_lider)] = []
                    data_lupaap[diplomado][prev + str(nivel)][numero][str(cedula_lider)].append(cedula)


        json_file_lupaap = BytesIO(json.dumps(data_lupaap).encode())

        informe_lupaap_json_filename = str(actualizacion.id) + '.json'
        actualizacion.informe_lupaap_json.save(informe_lupaap_json_filename, File(json_file_lupaap))

    else:
        with open(actualizacion.informe_lupaap_json.path) as f:
            data_lupaap = json.load(f)


    soportes = {}

    titulos = ['Cedula', 'Docente', 'Grupo', 'Lider', 'Cedula', 'Nivel', 'Sesión']

    formatos = ['0', 'General', 'General', 'General', '0', 'General', 'General']

    ancho_columnas = [20, 30, 50, 30, 30, 50, 30]

    contenidos = []

    for diplomado in data.keys():
        for nivel in data[diplomado].keys():
            if nivel == 'Socialización':
                nivel_update = 'Nivel 4'
            else:
                nivel_update = nivel
            for consecutivo_entregable in data[diplomado][nivel].keys():

                entregable = models.Entregables.objects.filter(momento__estrategia__nombre = diplomado).get(numero = consecutivo_entregable)

                for cedula in data[diplomado][nivel][consecutivo_entregable].keys():
                    fecha = datetime.datetime.strptime(data[diplomado][nivel][consecutivo_entregable][cedula],'%Y-%m-%d %H:%M:%S').date()

                    try:
                        docente = models.Docentes.objects.get(cedula = cedula)
                    except:
                        pass
                    else:
                        numero_1 = 0
                        numero_2 = 0
                        if diplomado == 'InnovaTIC':
                            numero_1 = consecutivo_entregable
                            numero_2 = 0
                        elif diplomado == 'RuralTIC':
                            numero_1 = 0
                            numero_2 = consecutivo_entregable
                        else:
                            raise NotImplementedError("Problemas con el nombre del diplomado")

                        estado = docente.get_estado_tablero_listado_texto(numero_1,numero_2)

                        if estado in ['Rechazado','',None]:
                            if docente.grupo != None:

                                print("Diplomado: {0} - Nivel: [1] - Entregable: {2}".format(
                                    diplomado,
                                    nivel_update,
                                    consecutivo_entregable
                                ))

                                cedulas = data_lupaap[diplomado][nivel_update][consecutivo_entregable].get(str(docente.grupo.ruta.contrato.contratista.cedula))

                                if cedulas != None:

                                    if int(cedula) in cedulas:
                                        nombre_grupo = docente.grupo.get_nombre_grupo()
                                        try:
                                            objeto = models.EntregableRutaObject.objects.get(ruta = docente.grupo.ruta,entregable = entregable,docente=docente,estado='asignado')
                                        except:
                                            #print('No existe el objeto')
                                            pass
                                        else:

                                            if nombre_grupo not in soportes.keys():
                                                soportes[nombre_grupo] = {}

                                            if consecutivo_entregable not in soportes[nombre_grupo].keys():

                                                if docente.grupo.ruta.region.nombre == 'Región 2':
                                                    red = actualizacion.red_r2
                                                elif docente.grupo.ruta.region.nombre == 'Región 3':
                                                    red = actualizacion.red_r3
                                                else:
                                                    red = None

                                                soporte = models.ListadoAsistencia.objects.create(
                                                    red=red,
                                                    grupo=docente.grupo,
                                                    entregable=entregable,
                                                    fecha=fecha,
                                                    estado='Aprobado'
                                                )

                                                soportes[nombre_grupo][consecutivo_entregable] = soporte
                                                soporte_filename = str(actualizacion.id) + '.xlsx'
                                                soporte.file.save(soporte_filename, File(actualizacion.informe_lupaap))

                                            soporte = soportes[nombre_grupo][consecutivo_entregable]

                                            soporte.docentes.add(docente)
                                            objeto.estado = 'Reportado'
                                            objeto.soporte = '{0}&{1}'.format(entregable.modelo,soporte.id)
                                            objeto.save()
                                            #print('Cedula: {0} - Lider: {1} - Consecutivo: {2}'.format(str(cedula),str(docente.grupo.ruta.contrato.contratista.cedula),consecutivo_entregable))

                                            contenidos.append([docente.cedula, docente.nombre, nombre_grupo, str(docente.grupo.ruta.contrato.contratista.get_full_name()), str(docente.grupo.ruta.contrato.contratista.cedula), nivel, entregable.nombre])

                                    else:
                                        pass
                                else:
                                    pass

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, 'Actualización Lupaap', datetime.datetime.now(), User.objects.get(email = 'sistemas@asoandes.org'), 'Actualización')


    filename = str(actualizacion.id) + '.xlsx'
    actualizacion.resultado.save(filename, File(output))

    return "Actualización Lupaap: " + filename

@app.task
def calcular_actualizacion_autoreporte_evaluacion(actualizacion_id):

    actualizacion = models.ActualizacionAutoreporteEvaluacion.objects.get(id = actualizacion_id)
    output = BytesIO()

    contador = 0
    soportes = {}

    if not bool(actualizacion.tablero_control_json):

        wb_tablero = openpyxl.load_workbook(filename=actualizacion.tablero_control.path)
        ws_tablero = wb_tablero.get_sheet_by_name('Docentes')

        entregables_innovatic = {

            'M': {'numero': 12},
            'Y': {'numero': 13},

            'AL': {'numero': 21},


            'AY': {'numero': 29},


            'BR': {'numero': 39},
            'CE': {'numero': 40},
        }

        entregables_ruraltic = {

            'M': {'numero': 5},

            'AL': {'numero': 18},

            'AY': {'numero': 31},

            'BR': {'numero': 51},

            'CD': {'numero': 56},
            'CE': {'numero': 57},
        }

        data = {
            'InnovaTIC': {},
            'RuralTIC': {}
        }

        for i in range(2,ws_tablero.max_row+1):
            if ws_tablero['H'+str(i)].value == 'InnovaTIC':
                entregables = entregables_innovatic
                diplomado = 'InnovaTIC'
            elif ws_tablero['H' + str(i)].value == 'RuralTIC':
                entregables = entregables_ruraltic
                diplomado = 'RuralTIC'
            else:
                entregables = {}
                diplomado = ''

            for key in entregables.keys():
                entregable = models.Entregables.objects.get(momento__estrategia__nombre=diplomado,numero=entregables[key]['numero'])
                nivel = entregable.momento.nombre
                numero = entregables[key]['numero']

                if nivel not in data[diplomado].keys():
                    data[diplomado][nivel] = {}

                if numero not in data[diplomado][nivel].keys():
                    data[diplomado][nivel][numero] = {}

                value = ws_tablero['{0}{1}'.format(key, str(i))].value
                cedula = ws_tablero['E' + str(i)].value

                if value == '' or value == None:
                    pass
                else:
                    if cedula not in data[diplomado][nivel].keys():
                        data[diplomado][nivel][numero][cedula] = {}

                    data[diplomado][nivel][numero][cedula] = str(value)

        json_file_tablero = BytesIO(json.dumps(data).encode())


        tablero_json_filename = str(actualizacion.id) + '.json'
        actualizacion.tablero_control_json.save(tablero_json_filename, File(json_file_tablero))

    else:
        with open(actualizacion.tablero_control_json.path) as f:
            data = json.load(f)


    soportes = {}

    titulos = ['Cedula', 'Docente', 'Grupo', 'Lider', 'Cedula', 'Nivel', 'Sesión', 'Resultado']

    formatos = ['0', 'General', 'General', 'General', '0', 'General', 'General', 'General']

    ancho_columnas = [20, 30, 50, 30, 30, 50, 30, 30]

    contenidos = []

    for diplomado in data.keys():
        for nivel in data[diplomado].keys():
            for consecutivo_entregable in data[diplomado][nivel].keys():

                entregable = models.Entregables.objects.filter(momento__estrategia__nombre = diplomado).get(numero = consecutivo_entregable)

                for cedula in data[diplomado][nivel][consecutivo_entregable].keys():

                    resultado = ''

                    fecha = datetime.datetime.strptime(data[diplomado][nivel][consecutivo_entregable][cedula],'%Y-%m-%d %H:%M:%S').date()

                    try:
                        docente = models.Docentes.objects.get(cedula = cedula)
                    except:
                        resultado = 'No existe el docente'
                    else:
                        if docente.grupo != None:
                            nombre_grupo = docente.grupo.get_nombre_grupo()
                        else:
                            nombre_grupo = ''

                        if docente.grupo != None:
                            contratista = str(docente.grupo.ruta.contrato.contratista.get_full_name())
                            cedula_contratista = str(docente.grupo.ruta.contrato.contratista.cedula)
                        else:
                            contratista = ''
                            cedula_contratista = ''

                        numero_1 = 0
                        numero_2 = 0
                        if diplomado == 'InnovaTIC':
                            numero_1 = consecutivo_entregable
                            numero_2 = 0
                        elif diplomado == 'RuralTIC':
                            numero_1 = 0
                            numero_2 = consecutivo_entregable
                        else:
                            raise NotImplementedError("Problemas con el nombre del diplomado")

                        if entregable.modelo == "instrumento_autoreporte":
                            estado = docente.get_estado_tablero_autoreporte_texto(numero_1,numero_2)
                            modelo = models.InstrumentoAutoreporte
                        elif entregable.modelo == "instrumento_hagamos_memoria":
                            estado = docente.get_estado_tablero_evaluacion_texto(numero_1, numero_2)
                            modelo = models.InstrumentoHagamosMemoria
                        elif entregable.modelo == "instrumento_evaluacion":
                            estado = docente.get_estado_tablero_evaluacion_texto(numero_1, numero_2)
                            modelo = models.InstrumentoEvaluacion
                        else:
                            estado = 'N/A'
                            modelo = None


                        if estado in ['Rechazado','',None]:
                            if docente.grupo != None:
                                nombre_grupo = docente.grupo.get_nombre_grupo()
                                try:
                                    objeto = models.EntregableRutaObject.objects.get(ruta = docente.grupo.ruta,entregable = entregable,docente=docente,estado='asignado')
                                except:
                                    resultado = 'El objeto no esta disponible'
                                else:
                                    resultado = 'Soporte creado'
                                    if nombre_grupo not in soportes.keys():
                                        soportes[nombre_grupo] = {}

                                    if consecutivo_entregable not in soportes[nombre_grupo].keys():

                                        soporte = modelo.objects.create(
                                            grupo=docente.grupo,
                                            entregable=entregable,
                                            fecha=fecha,
                                            estado='Aprobado'
                                        )

                                        soportes[nombre_grupo][consecutivo_entregable] = soporte
                                        soporte_filename = str(actualizacion.id) + '.xlsx'
                                        soporte.file.save(soporte_filename, File(actualizacion.tablero_control))

                                    soporte = soportes[nombre_grupo][consecutivo_entregable]

                                    soporte.docentes.add(docente)
                                    objeto.estado = 'Reportado'
                                    objeto.soporte = '{0}&{1}'.format(entregable.modelo,soporte.id)
                                    objeto.save()
                                    print('Cedula: {0} - Lider: {1} - Consecutivo: {2}'.format(str(cedula),str(docente.grupo.ruta.contrato.contratista.cedula),consecutivo_entregable))

                        else:
                            resultado = estado

                        contenidos.append([docente.cedula, docente.nombre, nombre_grupo, contratista, cedula_contratista, nivel, entregable.nombre, resultado])


    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, 'Actualización Lupaap', datetime.datetime.now(), User.objects.get(email = 'sistemas@asoandes.org'), 'Actualización')


    filename = str(actualizacion.id) + '.xlsx'
    actualizacion.resultado.save(filename, File(output))

    return "Actualización Lupaap: " + filename

@app.task
def calcular_actualizacion_productos_finales(actualizacion_id):

    actualizacion = models.ActualizacionProductosFinales.objects.get(id = actualizacion_id)
    output = BytesIO()

    contador = 0
    soportes = {}

    if not bool(actualizacion.tablero_control_json):

        wb_tablero = openpyxl.load_workbook(filename=actualizacion.tablero_control.path)
        ws_tablero = wb_tablero.get_sheet_by_name('Docentes')

        entregables_innovatic = {

            'CH': [20,30],
        }

        entregables_ruraltic = {

            'CH': [17,30,49,50],
        }

        data = {
            'InnovaTIC': {},
            'RuralTIC': {}
        }

        for i in range(2,ws_tablero.max_row+1):
            if ws_tablero['H'+str(i)].value == 'InnovaTIC':
                entregables = entregables_innovatic
                diplomado = 'InnovaTIC'
            elif ws_tablero['H' + str(i)].value == 'RuralTIC':
                entregables = entregables_ruraltic
                diplomado = 'RuralTIC'
            else:
                entregables = {}
                diplomado = ''

            for key in entregables.keys():
                for numero_entregable in entregables[key]:
                    print(numero_entregable)
                    entregable = models.Entregables.objects.get(momento__estrategia__nombre=diplomado,numero=numero_entregable)
                    nivel = entregable.momento.nombre
                    numero = numero_entregable

                    if nivel not in data[diplomado].keys():
                        data[diplomado][nivel] = {}

                    if numero not in data[diplomado][nivel].keys():
                        data[diplomado][nivel][numero] = {}

                    value = ws_tablero['{0}{1}'.format(key, str(i))].value
                    cedula = ws_tablero['E' + str(i)].value

                    if value == '' or value == None:
                        pass
                    else:
                        if cedula not in data[diplomado][nivel].keys():
                            data[diplomado][nivel][numero][cedula] = {}

                        data[diplomado][nivel][numero][cedula] = str(value)

        json_file_tablero = BytesIO(json.dumps(data).encode())


        tablero_json_filename = str(actualizacion.id) + '.json'
        actualizacion.tablero_control_json.save(tablero_json_filename, File(json_file_tablero))

    else:
        with open(actualizacion.tablero_control_json.path) as f:
            data = json.load(f)


    soportes = {}

    titulos = ['Cedula', 'Docente', 'Grupo', 'Lider', 'Cedula', 'Nivel', 'Sesión', 'Resultado']

    formatos = ['0', 'General', 'General', 'General', '0', 'General', 'General', 'General']

    ancho_columnas = [20, 30, 50, 30, 30, 50, 30, 30]

    contenidos = []

    for diplomado in data.keys():
        for nivel in data[diplomado].keys():
            for consecutivo_entregable in data[diplomado][nivel].keys():

                entregable = models.Entregables.objects.filter(momento__estrategia__nombre = diplomado).get(numero = consecutivo_entregable)

                for cedula in data[diplomado][nivel][consecutivo_entregable].keys():

                    resultado = ''

                    fecha = datetime.datetime.strptime(data[diplomado][nivel][consecutivo_entregable][cedula],'%Y-%m-%d %H:%M:%S').date()

                    try:
                        docente = models.Docentes.objects.get(cedula = cedula)
                    except:
                        resultado = 'No existe el docente'
                    else:
                        if docente.grupo != None:
                            nombre_grupo = docente.grupo.get_nombre_grupo()
                        else:
                            nombre_grupo = ''

                        if docente.grupo != None:
                            contratista = str(docente.grupo.ruta.contrato.contratista.get_full_name())
                            cedula_contratista = str(docente.grupo.ruta.contrato.contratista.cedula)
                        else:
                            contratista = ''
                            cedula_contratista = ''

                        numero_1 = 0
                        numero_2 = 0
                        if diplomado == 'InnovaTIC':
                            numero_1 = consecutivo_entregable
                            numero_2 = 0
                        elif diplomado == 'RuralTIC':
                            numero_1 = 0
                            numero_2 = consecutivo_entregable
                        else:
                            raise NotImplementedError("Problemas con el nombre del diplomado")



                        if entregable.modelo == "instrumento_estructuracion_ple":
                            modelo = models.InstrumentoEstructuracionPle
                            estado = docente.get_estado_tablero_modelo_texto(numero_1, numero_2, modelo)


                        elif entregable.modelo == "producto_final_ple":
                            modelo = models.ProductoFinalPle
                            estado = docente.get_estado_tablero_modelo_texto(numero_1, numero_2, modelo)


                        elif entregable.modelo == "presentacion_apa":
                            modelo = models.PresentacionApa
                            estado = docente.get_estado_tablero_modelo_texto(numero_1, numero_2, modelo)


                        elif entregable.modelo == "presentacion_actividad_pedagogica":
                            modelo = models.PresentacionActividadPedagogica
                            estado = docente.get_estado_tablero_modelo_texto(numero_1, numero_2, modelo)



                        elif entregable.modelo == "repositorio_actividades":
                            modelo = models.RepositorioActividades
                            estado = docente.get_estado_tablero_modelo_texto(numero_1, numero_2, modelo)


                        else:
                            estado = 'N/A'
                            modelo = None




                        if estado in ['Rechazado','',None]:
                            if docente.grupo != None:
                                nombre_grupo = docente.grupo.get_nombre_grupo()
                                try:
                                    objeto = models.EntregableRutaObject.objects.get(ruta = docente.grupo.ruta,entregable = entregable,docente=docente,estado='asignado')
                                except:
                                    resultado = 'El objeto no esta disponible'
                                else:
                                    resultado = 'Soporte creado'
                                    if nombre_grupo not in soportes.keys():
                                        soportes[nombre_grupo] = {}

                                    if consecutivo_entregable not in soportes[nombre_grupo].keys():

                                        soporte = modelo.objects.create(
                                            grupo=docente.grupo,
                                            entregable=entregable,
                                            fecha=fecha,
                                            estado='Aprobado'
                                        )

                                        soportes[nombre_grupo][consecutivo_entregable] = soporte
                                        soporte_filename = str(actualizacion.id) + '.xlsx'
                                        soporte.file.save(soporte_filename, File(actualizacion.tablero_control))

                                    soporte = soportes[nombre_grupo][consecutivo_entregable]

                                    soporte.docentes.add(docente)
                                    objeto.estado = 'Reportado'
                                    objeto.soporte = '{0}&{1}'.format(entregable.modelo,soporte.id)
                                    objeto.save()
                                    print('Cedula: {0} - Lider: {1} - Consecutivo: {2}'.format(str(cedula),str(docente.grupo.ruta.contrato.contratista.cedula),consecutivo_entregable))

                        else:
                            resultado = estado

                        contenidos.append([docente.cedula, docente.nombre, nombre_grupo, contratista, cedula_contratista, nivel, entregable.nombre, resultado])


    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, 'Actualización Lupaap', datetime.datetime.now(), User.objects.get(email = 'sistemas@asoandes.org'), 'Actualización')


    filename = str(actualizacion.id) + '.xlsx'
    actualizacion.resultado.save(filename, File(output))

    return "Actualización Lupaap: " + filename

@app.task
def build_red_estado_formacion(id, red_id):
    reporte = models_reportes.Reportes.objects.get(id = id)
    proceso = "SICAN-ESTADO-RED"
    red = models.Red.objects.get(id = red_id)

    titulos = ['Nro', 'Región', 'Departamento', 'Municipio', 'Cedula del Docente', 'Nombre del Docente', 'No. RED','Diplomado',
               'No. Grupo', 'Nivel', 'Sesión', 'Fecha sesión', 'Estado', 'Fecha actualización estado', 'Observación']

    formatos = ['0', 'General', 'General', 'General', '0', 'General', '0', 'General',
                'General', 'General', 'General', 'dd/mm/yyyyy', 'General', 'dd/mm/yyyyy', 'General']

    ancho_columnas = [20, 30, 50, 30, 30, 30, 30, 30,
                      30, 30, 30, 30, 30, 30, 100]

    contenidos = []

    asistencias_innovatic = models.ListadoAsistencia.objects.filter(red=red, grupo__estrategia__nombre='InnovaTIC')
    ple_innovatic = models.ProductoFinalPle.objects.filter(red=red, grupo__estrategia__nombre='InnovaTIC')


    asistencias_ruraltic = models.ListadoAsistencia.objects.filter(red=red, grupo__estrategia__nombre='RuralTIC')
    repositorio_ruraltic = models.RepositorioActividades.objects.filter(red=red, grupo__estrategia__nombre='RuralTIC')



    i = 0

    for asistencia_innovatic in asistencias_innovatic:

        for docente in asistencia_innovatic.docentes.all():
            i += 1
            contenidos.append([
                int(i),
                asistencia_innovatic.grupo.ruta.region.nombre,
                docente.municipio.departamento.nombre,
                docente.municipio.nombre,
                docente.cedula,
                docente.nombre,
                red.consecutivo,
                asistencia_innovatic.grupo.estrategia.nombre,
                asistencia_innovatic.grupo.get_nombre_grupo(),
                asistencia_innovatic.entregable.momento.nombre,
                asistencia_innovatic.entregable.nombre,
                asistencia_innovatic.fecha.strftime('%d/%m/%Y'),
                asistencia_innovatic.estado,
                asistencia_innovatic.get_fecha_actualizacion_estado(),
                asistencia_innovatic.get_observaciones()
            ])

    for ple in ple_innovatic:

        for docente in ple.docentes.all():
            i += 1
            contenidos.append([
                int(i),
                ple.grupo.ruta.region.nombre,
                docente.municipio.departamento.nombre,
                docente.municipio.nombre,
                docente.cedula,
                docente.nombre,
                red.consecutivo,
                ple.grupo.estrategia.nombre,
                ple.grupo.get_nombre_grupo(),
                ple.entregable.momento.nombre,
                ple.entregable.nombre,
                ple.fecha.strftime('%d/%m/%Y'),
                ple.estado,
                ple.get_fecha_actualizacion_estado(),
                ple.get_observaciones()
            ])

    for asistencia_ruraltic in asistencias_ruraltic:

        for docente in asistencia_ruraltic.docentes.all():
            i += 1
            contenidos.append([
                int(i),
                asistencia_ruraltic.grupo.ruta.region.nombre,
                docente.municipio.departamento.nombre,
                docente.municipio.nombre,
                docente.cedula,
                docente.nombre,
                red.consecutivo,
                asistencia_ruraltic.grupo.estrategia.nombre,
                asistencia_ruraltic.grupo.get_nombre_grupo(),
                asistencia_ruraltic.entregable.momento.nombre,
                asistencia_ruraltic.entregable.nombre,
                asistencia_ruraltic.fecha.strftime('%d/%m/%Y'),
                asistencia_ruraltic.estado,
                asistencia_ruraltic.get_fecha_actualizacion_estado(),
                asistencia_ruraltic.get_observaciones()
            ])

    for repositorio in repositorio_ruraltic:

        for docente in repositorio.docentes.all():
            i += 1
            contenidos.append([
                int(i),
                repositorio.grupo.ruta.region.nombre,
                docente.municipio.departamento.nombre,
                docente.municipio.nombre,
                docente.cedula,
                docente.nombre,
                red.consecutivo,
                repositorio.grupo.estrategia.nombre,
                repositorio.grupo.get_nombre_grupo(),
                repositorio.entregable.momento.nombre,
                repositorio.entregable.nombre,
                repositorio.fecha.strftime('%d/%m/%Y'),
                repositorio.estado,
                repositorio.get_fecha_actualizacion_estado(),
                repositorio.get_observaciones()
            ])


    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)


    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))


    return "Ruteo generado: " + filename

@app.task
def build_ftp_formacion(red_id):

    path = 'D:\\FTP-FORMACION-2018\\'

    red = models.Red.objects.get(id = red_id)

    asistencias_innovatic = models.ListadoAsistencia.objects.filter(red=red, grupo__estrategia__nombre='InnovaTIC')
    ple_innovatic = models.ProductoFinalPle.objects.filter(red=red, grupo__estrategia__nombre='InnovaTIC')

    asistencias_ruraltic = models.ListadoAsistencia.objects.filter(red=red, grupo__estrategia__nombre='RuralTIC')
    repositorio_ruraltic = models.RepositorioActividades.objects.filter(red=red, grupo__estrategia__nombre='RuralTIC')


    for asistencia_innovatic in asistencias_innovatic:
        region = asistencia_innovatic.grupo.ruta.region
        estrategia = asistencia_innovatic.grupo.estrategia
        entregable = asistencia_innovatic.entregable
        grupo = asistencia_innovatic.grupo

        if not os.path.exists(path + region.nombre):
            os.makedirs(path + region.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo)):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo))

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo()):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo())


        path_final = path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo()

        try:
            shutil.copy(asistencia_innovatic.file.path,path_final)
        except:
            pass

    for ple in ple_innovatic:
        region = ple.grupo.ruta.region
        estrategia = ple.grupo.estrategia
        entregable = ple.entregable
        grupo = ple.grupo

        if not os.path.exists(path + region.nombre):
            os.makedirs(path + region.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo)):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo))

        if not os.path.exists(
                path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo()):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo())

        path_final = path + region.nombre + '\\' + 'RED {0}'.format(
            red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo()

        try:
            shutil.copy(ple.file.path, path_final)
        except:
            pass

    for asistencia_ruraltic in asistencias_ruraltic:
        region = asistencia_ruraltic.grupo.ruta.region
        estrategia = asistencia_ruraltic.grupo.estrategia
        entregable = asistencia_ruraltic.entregable
        grupo = asistencia_ruraltic.grupo

        if not os.path.exists(path + region.nombre):
            os.makedirs(path + region.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo)):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo))

        if not os.path.exists(
                path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo()):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo())

        path_final = path + region.nombre + '\\' + 'RED {0}'.format(
            red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo()

        try:
            shutil.copy(asistencia_ruraltic.file.path, path_final)
        except:
            pass

    for repositorio in repositorio_ruraltic:
        region = repositorio.grupo.ruta.region
        estrategia = repositorio.grupo.estrategia
        entregable = repositorio.entregable
        grupo = repositorio.grupo

        if not os.path.exists(path + region.nombre):
            os.makedirs(path + region.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo)):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo))

        if not os.path.exists(
                path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(red.consecutivo) + '\\' + estrategia.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre)

        if not os.path.exists(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo()):
            os.makedirs(path + region.nombre + '\\' + 'RED {0}'.format(
                red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo())

        path_final = path + region.nombre + '\\' + 'RED {0}'.format(
            red.consecutivo) + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo()

        try:
            shutil.copy(repositorio.file.path, path_final)
        except:
            pass


    return "FTP formación: RED {0}".format(red.consecutivo)

@app.task
def build_ftp_acceso():

    path = 'D:\\EXPORT_SICAN\\ACCESO\\'

    modelos = {
        'encuesta_monitoreo': {
            'modelo': models.EncuestaMonitoreo,
            'nombre': 'Encuesta monitoreo'
        },
        'documento_legalizacion_terminales': {
            'modelo': models.DocumentoLegalizacionTerminales,
            'nombre': 'Legalización'
        },
        'documento_legalizacion_terminales_v1': {
            'modelo': models.DocumentoLegalizacionTerminalesValle1,
            'nombre': 'Legalización'
        },
        'documento_legalizacion_terminales_v2': {
            'modelo': models.DocumentoLegalizacionTerminalesValle2,
            'nombre': 'Legalización'
        },
        'relatoria_taller_apertura': {
            'modelo': models.RelatoriaTallerApertura,
            'nombre': 'Taller Apertura'
        },
        'relatoria_taller_administratic': {
            'modelo': models.RelatoriaTallerAdministratic,
            'nombre': 'Taller Administratic'
        },
        'relatoria_taller_contenidos_educativos': {
            'modelo': models.RelatoriaTallerContenidosEducativos,
            'nombre': 'Taller de contenidos'
        },
        'relatoria_taller_raee': {
            'modelo': models.RelatoriaTallerRAEE,
            'nombre': 'Taller RAEE'
        }
    }


    for key in modelos.keys():
        modelo = modelos[key]
        objetos = modelo['modelo'].objects.filter(estado = 'Aprobado')
        for objeto in objetos:

            if not os.path.exists(path + str(objeto.radicado.municipio.departamento.nombre)):
                os.makedirs(path + str(objeto.radicado.municipio.departamento.nombre))

            if not os.path.exists(path + str(objeto.radicado.municipio.departamento.nombre) + '\\' + str(objeto.radicado.municipio.nombre)):
                os.makedirs(path + str(objeto.radicado.municipio.departamento.nombre) + '\\' + str(objeto.radicado.municipio.nombre))

            if not os.path.exists(path + str(objeto.radicado.municipio.departamento.nombre) + '\\' + str(objeto.radicado.municipio.nombre) + '\\' + str(objeto.radicado.numero)):
                os.makedirs(path + str(objeto.radicado.municipio.departamento.nombre) + '\\' + str(objeto.radicado.municipio.nombre) + '\\' + str(objeto.radicado.numero))

            if not os.path.exists(path + str(objeto.radicado.municipio.departamento.nombre) + '\\' + str(objeto.radicado.municipio.nombre) + '\\' + str(objeto.radicado.numero) + '\\' + modelo['nombre']):
                os.makedirs(path + str(objeto.radicado.municipio.departamento.nombre) + '\\' + str(objeto.radicado.municipio.nombre) + '\\' + str(objeto.radicado.numero) + '\\' + modelo['nombre'])

            try:
                shutil.copy(objeto.file.path, path + str(objeto.radicado.municipio.departamento.nombre) + '\\' + str(objeto.radicado.municipio.nombre) + '\\' + str(objeto.radicado.numero) + '\\' + modelo['nombre'])
            except:
                pass


    return "FTP Acceso"

@app.task
def build_ftp_formacion_region(departamento_id):

    path = 'D:\\EXPORT_SICAN\\FORMACION\\'

    ids = models.ListadoAsistencia.objects.filter(estado = 'Aprobado', docentes__municipio__departamento__id = departamento_id).values_list('id',flat = True)

    asistencias_innovatic = models.ListadoAsistencia.objects.filter(id__in = ids)

    for asistencia_innovatic in asistencias_innovatic:
        region = asistencia_innovatic.grupo.ruta.region
        estrategia = asistencia_innovatic.grupo.estrategia
        entregable = asistencia_innovatic.entregable
        grupo = asistencia_innovatic.grupo


        if not os.path.exists(path + region.nombre):
            os.makedirs(path + region.nombre)

        if not os.path.exists(path + region.nombre + '\\' + estrategia.nombre):
            os.makedirs(path + region.nombre + '\\' + estrategia.nombre)

        if not os.path.exists(path + region.nombre + '\\' + estrategia.nombre + '\\' + entregable.nombre):
            os.makedirs(path + region.nombre + '\\' + estrategia.nombre + '\\' + entregable.nombre)

        if not os.path.exists(path + region.nombre + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo()):
            os.makedirs(path + region.nombre + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo())

        path_final = path + region.nombre + '\\' + estrategia.nombre + '\\' + entregable.nombre + '\\' + grupo.get_nombre_grupo()

        try:
            shutil.copy(asistencia_innovatic.file.path, path_final)
        except:
            pass


    return "FTP formación"

@app.task
def build_ftp_formacion_region_pet_1(departamento_id):

    path = 'D:\\EXPORT_SICAN\\FORMACION\\PET_1\\'

    ids = models.ListadoAsistencia.objects.filter(estado = 'Aprobado', docentes__municipio__departamento__id = departamento_id).values_list('id',flat = True)

    asistencias_innovatic = models.ListadoAsistencia.objects.filter(id__in = ids)

    for asistencia_innovatic in asistencias_innovatic:
        region = asistencia_innovatic.grupo.ruta.region
        estrategia = asistencia_innovatic.grupo.estrategia
        entregable = asistencia_innovatic.entregable
        grupo = asistencia_innovatic.grupo
        municipios_id = asistencia_innovatic.docentes.all().values_list('municipio__id',flat = True).distinct()
        municipios = models.Municipios.objects.filter(id__in = municipios_id)
        departamento = models.Departamentos.objects.get(id = departamento_id)
        fecha_str = asistencia_innovatic.fecha.strftime('%d-%m-%Y')
        ext = asistencia_innovatic.file.path.split('.')[-1]
        filename = '{0}-{1}'.format(asistencia_innovatic.entregable.nombre,asistencia_innovatic.entregable.momento.nombre)

        for municipio in municipios:

            if not os.path.exists(path + departamento.nombre):
                os.makedirs(path + departamento.nombre)

            if not os.path.exists(path + departamento.nombre + '\\' + municipio.nombre):
                os.makedirs(path + departamento.nombre + '\\' + municipio.nombre)

            if not os.path.exists(path + departamento.nombre + '\\' + municipio.nombre + '\\' + fecha_str):
                os.makedirs(path + departamento.nombre + '\\' + municipio.nombre + '\\' + fecha_str)

            path_final = path + departamento.nombre + '\\' + municipio.nombre + '\\' + fecha_str

            try:
                shutil.copy(asistencia_innovatic.file.path, path_final + '\\' + filename + '.' + ext)
            except:
                pass


    return "FTP formación"


@app.task
def build_ftp_formacion_region_final():

    path = 'D:\\EXPORT_SICAN\\FORMACION_FINAL\\'

    ids = models.ListadoAsistencia.objects.filter(estado = 'Aprobado').values_list('id',flat = True)
    asistencias_innovatic = models.ListadoAsistencia.objects.filter(id__in = ids)

    ids_ple = models.ProductoFinalPle.objects.filter(estado='Aprobado').values_list('id', flat=True)
    ples = models.ProductoFinalPle.objects.filter(id__in=ids_ple)

    ids_apas = models.RepositorioActividades.objects.filter(estado='Aprobado').values_list('id', flat=True)
    apas = models.RepositorioActividades.objects.filter(id__in=ids_apas)

    for asistencia_innovatic in asistencias_innovatic:
        for docente in asistencia_innovatic.docentes.all():
            region = asistencia_innovatic.grupo.ruta.region
            estrategia = asistencia_innovatic.grupo.estrategia
            entregable = asistencia_innovatic.entregable


            if not os.path.exists(path + region.nombre):
                os.makedirs(path + region.nombre)


            if not os.path.exists(path + region.nombre + '\\' + estrategia.nombre):
                os.makedirs(path + region.nombre + '\\' + estrategia.nombre)

            if not os.path.exists(path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula)):
                os.makedirs(path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula))


            if not os.path.exists(path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula) + '\\' + entregable.nombre):
                os.makedirs(path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula) + '\\' + entregable.nombre)


            path_final = path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula) + '\\' + entregable.nombre

            try:
                shutil.copy(asistencia_innovatic.file.path, path_final)
            except:
                pass

    for ple in ples:
        for docente in ple.docentes.all():
            region = ple.grupo.ruta.region
            estrategia = ple.grupo.estrategia
            entregable = ple.entregable

            if not os.path.exists(path + region.nombre):
                os.makedirs(path + region.nombre)

            if not os.path.exists(path + region.nombre + '\\' + estrategia.nombre):
                os.makedirs(path + region.nombre + '\\' + estrategia.nombre)

            if not os.path.exists(path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula)):
                os.makedirs(path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula))

            if not os.path.exists(
                    path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula) + '\\' + entregable.nombre):
                os.makedirs(
                    path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula) + '\\' + entregable.nombre)

            path_final = path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula) + '\\' + entregable.nombre

            try:
                shutil.copy(ple.file.path, path_final)
            except:
                pass

    for apa in apas:
        for docente in apa.docentes.all():
            region = apa.grupo.ruta.region
            estrategia = apa.grupo.estrategia
            entregable = apa.entregable

            if not os.path.exists(path + region.nombre):
                os.makedirs(path + region.nombre)

            if not os.path.exists(path + region.nombre + '\\' + estrategia.nombre):
                os.makedirs(path + region.nombre + '\\' + estrategia.nombre)

            if not os.path.exists(path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula)):
                os.makedirs(path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula))

            if not os.path.exists(
                    path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula) + '\\' + entregable.nombre):
                os.makedirs(
                    path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula) + '\\' + entregable.nombre)

            path_final = path + region.nombre + '\\' + estrategia.nombre + '\\' + str(docente.cedula) + '\\' + entregable.nombre

            try:
                shutil.copy(apa.file.path, path_final)
            except:
                pass

    return "FTP formación"


@app.task
def build_excel_corte(id):
    corte = models.Cortes.objects.get(id = id)
    proceso = "REPORTE PAGO EVIDENCIAS"

    titulos = ['Consecutivo', 'Fecha', 'Ruta', 'Contratista', 'Cedula',
               'Radicado', 'Departamento radicado','Municipio radicado', 'Nombre Sede radicado','Ubicación sede radicado',
               'Docente','Cedula', 'Departamento docente', 'Municipio docente', 'Diplomado',
               'Producto', 'Presupuesto','Costo']

    formatos = ['0', 'dd/mm/yyyy', 'General','General', 'General',
                'General', 'General', 'General', 'General', 'General',
                'General', 'General', 'General', 'General', 'General',
                'General', 'General', '"$"#,##0_);("$"#,##0)']

    ancho_columnas = [20, 30, 50, 50, 30,
                      30, 40, 40, 40, 30,
                      50, 30, 30, 30, 30,
                      30, 30, 30]

    contenidos = []


    i = 0
    for entregable in models.EntregableRutaObject.objects.filter(corte = corte).order_by('ruta'):
        i += 1
        contenidos.append([
            int(i),
            corte.creation,
            entregable.ruta.nombre,
            entregable.ruta.contrato.contratista.get_full_name(),
            entregable.ruta.contrato.contratista.cedula,

            entregable.get_radicado_numero(),
            entregable.get_radicado_departamento(),
            entregable.get_radicado_municipio(),
            entregable.get_radicado_sede(),
            entregable.get_radicado_ubicacion(),

            entregable.get_docente_nombre(),
            entregable.get_docente_cedula(),
            entregable.get_docente_departamento(),
            entregable.get_docente_municipio(),
            entregable.get_docente_diplomado(),


            entregable.entregable.nombre if entregable.entregable != None else '',
            entregable.entregable.presupuesto if entregable.entregable != None else '',
            entregable.valor.amount
        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, 'Corte {0} - {1}'.format(corte.consecutivo,corte.region.nombre), corte.creation, corte.usuario_creacion, proceso)

    filename = str(corte.id) + '.xlsx'
    corte.file.save(filename, File(output))

    return "Reporte generado: " + filename



@app.task
def build_excel_liquidacion(id,user_id):
    liquidacion = models.Liquidaciones.objects.get(id = id)
    usuario = User.objects.get(id = user_id)
    proceso = "REPORTE LIQUIDACION"

    titulos = ['Consecutivo', 'Fecha', 'Ruta', 'Contratista', 'Cedula',
               'Radicado', 'Departamento radicado','Municipio radicado', 'Nombre Sede radicado','Ubicación sede radicado',
               'Docente','Cedula', 'Departamento docente', 'Municipio docente', 'Diplomado',
               'Producto', 'Presupuesto','Costo']

    formatos = ['0', 'dd/mm/yyyy', 'General','General', 'General',
                'General', 'General', 'General', 'General', 'General',
                'General', 'General', 'General', 'General', 'General',
                'General', 'General', '"$"#,##0_);("$"#,##0)']

    ancho_columnas = [20, 30, 50, 50, 30,
                      30, 40, 40, 40, 30,
                      50, 30, 30, 30, 30,
                      30, 30, 30]

    contenidos = []


    i = 0
    for entregable in models.EntregableRutaObject.objects.filter(liquidacion = liquidacion):
        i += 1
        contenidos.append([
            int(i),
            liquidacion.fecha_actualizacion,
            entregable.ruta.nombre,
            entregable.ruta.contrato.contratista.get_full_name(),
            entregable.ruta.contrato.contratista.cedula,

            entregable.get_radicado_numero(),
            entregable.get_radicado_departamento(),
            entregable.get_radicado_municipio(),
            entregable.get_radicado_sede(),
            entregable.get_radicado_ubicacion(),

            entregable.get_docente_nombre(),
            entregable.get_docente_cedula(),
            entregable.get_docente_departamento(),
            entregable.get_docente_municipio(),
            entregable.get_docente_diplomado(),


            entregable.entregable.nombre if entregable.entregable != None else '',
            entregable.entregable.presupuesto if entregable.entregable != None else '',
            entregable.valor.amount
        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, '', timezone.now(), usuario, proceso)

    filename = str(liquidacion.id) + '.xlsx'
    liquidacion.file3.save(filename, File(output))

    return "Reporte generado: " + filename



@app.task
def build_tablero_control(id,region):
    reporte = models_reportes.Reportes.objects.get(id = id)
    region = models.Regiones.objects.get(id = region)
    proceso = "TABLERO DE CONTROL"

    titulos = ['Radicado', 'Región', 'Departamento', 'CodMunicipio', 'Municipio', 'Nombre_IE', 'DANE_Sede', 'Nombre_Sede',
               'Legalizacion', 'Taller Apertura', 'Taller Administratic', 'Taller Contenidos', 'Taller RAEE']

    formatos = ['0', 'General', 'General','0', 'General', 'General', 'General', 'General',
                'dd/mm/yyyy', 'dd/mm/yyyy', 'dd/mm/yyyy', 'dd/mm/yyyy', 'dd/mm/yyyy']

    ancho_columnas = [20, 30, 50, 30, 30, 50, 30, 50,
                      30, 30, 30, 30, 30]

    contenidos = []

    for radicado in models.Radicados.objects.filter(municipio__departamento__region = region):
        contenidos.append([
            radicado.numero,
            radicado.municipio.departamento.region.nombre,
            radicado.municipio.departamento.nombre,
            radicado.municipio.numero,
            radicado.municipio.nombre,
            radicado.nombre_ie,
            radicado.dane_sede,
            radicado.nombre_sede,
            radicado.get_list_legalizacion(),
            radicado.get_list_taller_apertura(),
            radicado.get_list_taller_administratic(),
            radicado.get_list_taller_contenidos(),
            radicado.get_list_taller_raee()
        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)


    titulos = ['Radicado', 'Región', 'Municipio', 'Departamento', 'Fecha','CPU', 'CRT', 'LCD', 'PORTATIL', 'Tableta',
               'Impresora', 'Equipos Calculadora', 'Equipos Sican', 'Estado']

    formatos = ['0', 'General', 'General', 'General', 'dd/mm/yyyy', '0', '0', '0', '0', '0',
                '0', '0', '0', 'General']

    ancho_columnas = [20, 20, 30, 30, 30, 20, 20, 20, 20, 20,
                      20, 30, 30, 30]

    contenidos = []


    for retoma in models.Retoma.objects.filter(ruta__region = region):

        contenidos.append([
            retoma.radicado,
            retoma.ruta.region.nombre,
            retoma.municipio.nombre,
            retoma.municipio.departamento.nombre,
            retoma.fecha,
            retoma.cpu,
            retoma.trc,
            retoma.lcd,
            retoma.portatil,
            retoma.tableta,
            retoma.impresora,
            retoma.get_equipos_calculadora(),
            retoma.get_bolsas_calculadora_sican(),
            retoma.estado
        ])

    output = construir_reporte_pagina(output, 'Hoja2', titulos, contenidos, formatos, ancho_columnas, reporte.nombre,
                                       reporte.creation, reporte.usuario, proceso)


    titulos = ['Departamento', 'Municipio', 'Cedula', 'Nombre', 'Diplomado', 'Registro', 'AutoreporteIn',
               'N1S01', 'ESTADO','RED', 'N1S02', 'ESTADO','RED', 'N1S03', 'ESTADO','RED', 'N1S04', 'ESTADO','RED','N1S05', 'ESTADO','RED','N1S06','ESTADO','RED', 'N1S07','ESTADO','RED', 'N1S08','ESTADO','RED', 'N1S09','ESTADO','RED',
               'PruebaNivel1', 'NroIntentosN1', 'FechaPruebaNivel1',
               'N2S01','ESTADO','RED', 'N2S02','ESTADO','RED', 'N2S03','ESTADO','RED', 'N2S04','ESTADO','RED', 'N2S05','ESTADO','RED', 'N2S06','ESTADO','RED', 'N2S07','ESTADO','RED', 'N2S08','ESTADO','RED', 'N2S09','ESTADO','RED', 'N2S010','ESTADO','RED',
               'PruebaNivel2', 'NroIntentosN2', 'FechaPruebaNivel2',
               'N3S01','ESTADO','RED', 'N3S02','ESTADO','RED', 'N3S03','ESTADO','RED', 'N3S04','ESTADO','RED', 'N3S05','ESTADO','RED', 'N3S06','ESTADO','RED', 'N3S07','ESTADO','RED', 'N3S08','ESTADO','RED', 'N3S09','ESTADO','RED', 'N3S010','ESTADO','RED',
               'PruebaNivel3', 'NroIntentosN3', 'FechaPruebaNivel3',
               'N4S01','ESTADO','RED', 'N4S02','ESTADO','RED', 'N4S03','ESTADO','RED', 'N4S04','ESTADO','RED', 'N4S05','ESTADO','RED', 'N4S06','ESTADO','RED', 'N4S07','ESTADO','RED', 'N4S08','ESTADO','RED', 'N4S09','ESTADO','RED', 'N4S010','ESTADO','RED', 'N4S11','ESTADO','RED', 'N4S12','ESTADO','RED', 'N4S13','ESTADO','RED', 'N4S14','ESTADO','RED', 'N4S15','ESTADO','RED', 'N4S016','ESTADO','RED',
               'PruebaNivel4', 'NroIntentosN4', 'FechaPruebaNivel4',
               'N5S01','ESTADO','RED', 'N5S02','ESTADO','RED', 'N5S03','ESTADO','RED', 'N5S04','ESTADO','RED', 'N5S05','ESTADO','RED', 'N5S06','ESTADO','RED', 'N5S07','ESTADO','RED', 'N5S08','ESTADO','RED', 'N5S09','ESTADO','RED',
               'PruebaNivel5', 'NroIntentosN5', 'FechaPruebaNivel5', 'AutoreporteOut', 'Producto Final', 'ESTADO', 'RED', 'Listados Aprobados'
               ]

    formatos = ['General', 'General', '0', 'General', 'General', 'dd/mm/yyyy', 'dd/mm/yyyy',
                'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General',
                'dd/mm/yyyy', 'dd/mm/yyyy', 'dd/mm/yyyy',
                'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General',
                'dd/mm/yyyy', 'dd/mm/yyyy', 'dd/mm/yyyy',
                'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General',
                'dd/mm/yyyy', 'dd/mm/yyyy', 'dd/mm/yyyy',
                'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General',
                'dd/mm/yyyy', 'dd/mm/yyyy', 'dd/mm/yyyy',
                'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General', 'dd/mm/yyyy','General','General',
                'dd/mm/yyyy', 'dd/mm/yyyy', 'dd/mm/yyyy', 'dd/mm/yyyy', 'dd/mm/yyyy', 'General', 'General', '0']

    ancho_columnas = [20, 40, 40, 40, 40, 40, 30,
                      20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20,
                      20, 20, 20,
                      20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20,
                      20, 20, 20,
                      20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20,
                      20, 20, 20,
                      20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20,
                      20, 20, 20,
                      20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20,
                      20, 20, 20, 20, 20, 20, 20, 20
                      ]

    contenidos = []

    for docente in models.Docentes.objects.filter(municipio__departamento__region = region):
        contenidos.append([
            docente.municipio.departamento.nombre,
            docente.municipio.nombre,
            docente.cedula,
            docente.nombre,
            docente.estrategia.nombre,
            docente.registro,

            docente.get_estado_tablero_autoreporte(12, 5),

            docente.get_estado_tablero_listado(3, 3),
            docente.get_estado_tablero_listado_texto(3, 3),
            docente.get_estado_tablero_listado_red(3, 3),
            docente.get_estado_tablero_listado(4, 4),
            docente.get_estado_tablero_listado_texto(4, 4),
            docente.get_estado_tablero_listado_red(4, 4),
            docente.get_estado_tablero_listado(5, 0),
            docente.get_estado_tablero_listado_texto(5, 0),
            docente.get_estado_tablero_listado_red(5, 0),
            docente.get_estado_tablero_listado(6, 0),
            docente.get_estado_tablero_listado_texto(6, 0),
            docente.get_estado_tablero_listado_red(6, 0),
            docente.get_estado_tablero_listado(7, 0),
            docente.get_estado_tablero_listado_texto(7, 0),
            docente.get_estado_tablero_listado_red(7, 0),
            docente.get_estado_tablero_listado(8, 0),
            docente.get_estado_tablero_listado_texto(8, 0),
            docente.get_estado_tablero_listado_red(8, 0),
            docente.get_estado_tablero_listado(9, 0),
            docente.get_estado_tablero_listado_texto(9, 0),
            docente.get_estado_tablero_listado_red(9, 0),
            docente.get_estado_tablero_listado(10, 0),
            docente.get_estado_tablero_listado_texto(10, 0),
            docente.get_estado_tablero_listado_red(10, 0),
            docente.get_estado_tablero_listado(11, 0),
            docente.get_estado_tablero_listado_texto(11, 0),
            docente.get_estado_tablero_listado_red(11, 0),

            docente.get_estado_tablero_evaluacion(13, 0),
            '',
            '',

            docente.get_estado_tablero_listado(15, 7),
            docente.get_estado_tablero_listado_texto(15, 7),
            docente.get_estado_tablero_listado_red(15, 7),
            docente.get_estado_tablero_listado(16, 8),
            docente.get_estado_tablero_listado_texto(16, 8),
            docente.get_estado_tablero_listado_red(16, 8),
            docente.get_estado_tablero_listado(17, 9),
            docente.get_estado_tablero_listado_texto(17, 9),
            docente.get_estado_tablero_listado_red(17, 9),
            docente.get_estado_tablero_listado(18, 10),
            docente.get_estado_tablero_listado_texto(18, 10),
            docente.get_estado_tablero_listado_red(18, 10),
            docente.get_estado_tablero_listado(19, 11),
            docente.get_estado_tablero_listado_texto(19, 11),
            docente.get_estado_tablero_listado_red(19, 11),
            docente.get_estado_tablero_listado(0, 12),
            docente.get_estado_tablero_listado_texto(0, 12),
            docente.get_estado_tablero_listado_red(0, 12),
            docente.get_estado_tablero_listado(0, 13),
            docente.get_estado_tablero_listado_texto(0, 13),
            docente.get_estado_tablero_listado_red(0, 13),
            docente.get_estado_tablero_listado(0, 14),
            docente.get_estado_tablero_listado_texto(0, 14),
            docente.get_estado_tablero_listado_red(0, 14),
            docente.get_estado_tablero_listado(0, 15),
            docente.get_estado_tablero_listado_texto(0, 15),
            docente.get_estado_tablero_listado_red(0, 15),
            docente.get_estado_tablero_listado(0, 16),
            docente.get_estado_tablero_listado_texto(0, 16),
            docente.get_estado_tablero_listado_red(0, 16),

            docente.get_estado_tablero_evaluacion(21, 18),
            '',
            '',

            docente.get_estado_tablero_listado(23, 20),
            docente.get_estado_tablero_listado_texto(23, 20),
            docente.get_estado_tablero_listado_red(23, 20),
            docente.get_estado_tablero_listado(24, 21),
            docente.get_estado_tablero_listado_texto(24, 21),
            docente.get_estado_tablero_listado_red(24, 21),
            docente.get_estado_tablero_listado(25, 22),
            docente.get_estado_tablero_listado_texto(25, 22),
            docente.get_estado_tablero_listado_red(25, 22),
            docente.get_estado_tablero_listado(26, 23),
            docente.get_estado_tablero_listado_texto(26, 23),
            docente.get_estado_tablero_listado_red(26, 23),
            docente.get_estado_tablero_listado(27, 24),
            docente.get_estado_tablero_listado_texto(27, 24),
            docente.get_estado_tablero_listado_red(27, 24),
            docente.get_estado_tablero_listado(28, 25),
            docente.get_estado_tablero_listado_texto(28, 25),
            docente.get_estado_tablero_listado_red(28, 25),
            docente.get_estado_tablero_listado(0, 26),
            docente.get_estado_tablero_listado_texto(0, 26),
            docente.get_estado_tablero_listado_red(0, 26),
            docente.get_estado_tablero_listado(0, 27),
            docente.get_estado_tablero_listado_texto(0, 27),
            docente.get_estado_tablero_listado_red(0, 27),
            docente.get_estado_tablero_listado(0, 28),
            docente.get_estado_tablero_listado_texto(0, 28),
            docente.get_estado_tablero_listado_red(0, 28),
            docente.get_estado_tablero_listado(0, 29),
            docente.get_estado_tablero_listado_texto(0, 29),
            docente.get_estado_tablero_listado_red(0, 29),

            docente.get_estado_tablero_evaluacion(29, 31),
            '',
            '',

            docente.get_estado_tablero_listado(32, 33),
            docente.get_estado_tablero_listado_texto(32, 33),
            docente.get_estado_tablero_listado_red(32, 33),
            docente.get_estado_tablero_listado(33, 34),
            docente.get_estado_tablero_listado_texto(33, 34),
            docente.get_estado_tablero_listado_red(33, 34),
            docente.get_estado_tablero_listado(34, 35),
            docente.get_estado_tablero_listado_texto(34, 35),
            docente.get_estado_tablero_listado_red(34, 35),
            docente.get_estado_tablero_listado(35, 36),
            docente.get_estado_tablero_listado_texto(35, 36),
            docente.get_estado_tablero_listado_red(35, 36),
            docente.get_estado_tablero_listado(36, 37),
            docente.get_estado_tablero_listado_texto(36, 37),
            docente.get_estado_tablero_listado_red(36, 37),
            docente.get_estado_tablero_listado(37, 38),
            docente.get_estado_tablero_listado_texto(37, 38),
            docente.get_estado_tablero_listado_red(37, 38),
            docente.get_estado_tablero_listado(38, 39),
            docente.get_estado_tablero_listado_texto(38, 39),
            docente.get_estado_tablero_listado_red(38, 39),
            docente.get_estado_tablero_listado(0, 40),
            docente.get_estado_tablero_listado_texto(0, 40),
            docente.get_estado_tablero_listado_red(0, 40),
            docente.get_estado_tablero_listado(0, 41),
            docente.get_estado_tablero_listado_texto(0, 41),
            docente.get_estado_tablero_listado_red(0, 41),
            docente.get_estado_tablero_listado(0, 42),
            docente.get_estado_tablero_listado_texto(0, 42),
            docente.get_estado_tablero_listado_red(0, 42),
            docente.get_estado_tablero_listado(0, 43),
            docente.get_estado_tablero_listado_texto(0, 43),
            docente.get_estado_tablero_listado_red(0, 43),
            docente.get_estado_tablero_listado(0, 44),
            docente.get_estado_tablero_listado_texto(0, 44),
            docente.get_estado_tablero_listado_red(0, 44),
            docente.get_estado_tablero_listado(0, 45),
            docente.get_estado_tablero_listado_texto(0, 45),
            docente.get_estado_tablero_listado_red(0, 45),
            docente.get_estado_tablero_listado(0, 46),
            docente.get_estado_tablero_listado_texto(0, 46),
            docente.get_estado_tablero_listado_red(0, 46),
            docente.get_estado_tablero_listado(0, 47),
            docente.get_estado_tablero_listado_texto(0, 47),
            docente.get_estado_tablero_listado_red(0, 47),
            docente.get_estado_tablero_listado(0, 48),
            docente.get_estado_tablero_listado_texto(0, 48),
            docente.get_estado_tablero_listado_red(0, 48),

            docente.get_estado_tablero_evaluacion(39, 51),
            '',
            '',

            docente.get_estado_tablero_listado(0, 53),
            docente.get_estado_tablero_listado_texto(0, 53),
            docente.get_estado_tablero_listado_red(0, 53),
            docente.get_estado_tablero_listado(0, 54),
            docente.get_estado_tablero_listado_texto(0, 54),
            docente.get_estado_tablero_listado_red(0, 54),
            docente.get_estado_tablero_listado(0, 0),
            docente.get_estado_tablero_listado_texto(0, 0),
            docente.get_estado_tablero_listado_red(0, 0),
            docente.get_estado_tablero_listado(0, 0),
            docente.get_estado_tablero_listado_texto(0, 0),
            docente.get_estado_tablero_listado_red(0, 0),
            docente.get_estado_tablero_listado(0, 0),
            docente.get_estado_tablero_listado_texto(0, 0),
            docente.get_estado_tablero_listado_red(0, 0),
            docente.get_estado_tablero_listado(0, 0),
            docente.get_estado_tablero_listado_texto(0, 0),
            docente.get_estado_tablero_listado_red(0, 0),
            docente.get_estado_tablero_listado(0, 0),
            docente.get_estado_tablero_listado_texto(0, 0),
            docente.get_estado_tablero_listado_red(0, 0),
            docente.get_estado_tablero_listado(0, 0),
            docente.get_estado_tablero_listado_texto(0, 0),
            docente.get_estado_tablero_listado_red(0, 0),
            docente.get_estado_tablero_listado(0, 0),
            docente.get_estado_tablero_listado_texto(0, 0),
            docente.get_estado_tablero_listado_red(0, 0),

            docente.get_estado_tablero_evaluacion(0, 56),
            '',
            '',

            docente.get_estado_tablero_autoreporte(40, 57),

            docente.get_estado_tablero_producto_final(),
            docente.get_estado_tablero_producto_final_texto(),
            docente.get_estado_tablero_producto_final_red(),
            docente.get_cantidad_listados_aprobados(),
        ])

    output = construir_reporte_pagina(output, 'Hoja3', titulos, contenidos, formatos, ancho_columnas, reporte.nombre,
                                      reporte.creation, reporte.usuario, proceso)



    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))


    return "Tablero generado: " + filename
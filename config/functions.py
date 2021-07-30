from __future__ import unicode_literals
import openpyxl
from io import BytesIO
from django.conf import settings
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from pytz import timezone
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

settings_time_zone = timezone(settings.TIME_ZONE)

def construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso):
    if len(ancho_columnas) != len(formatos) != len(titulos) != len(contenidos[0]):
        raise Exception('El arreglo de filas y columnas tienen distinta longitud')
    else:
        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/formato_informe.xlsx')
        ws = wb.get_sheet_by_name('Hoja1')
        #logo_sican = Image(settings.STATICFILES_DIRS[0] + '/img/icon-192.png')
        #logo_sican.drawing.top = 0
        #logo_sican.drawing.left = 0
        #ws.add_image(logo_sican)

        ws['B1'] = "   Nombre: " + nombre
        ws['B3'] = "   Fecha: " + fecha.astimezone(settings_time_zone).strftime('%d/%m/%Y a las %I:%M:%S %p')
        ws['B5'] = "   Usuario: " + usuario.email
        ws['B7'] = "   Proceso: " + proceso


        row_num = 9
        for col_num in range(len(titulos)):
            ws.cell(row=row_num, column=col_num+1).value = titulos[col_num]
            if col_num != 0:
                ws.column_dimensions[get_column_letter(col_num+1)].width = ancho_columnas[col_num]

            ws.cell(row=row_num, column=col_num + 1).font = Font(name='Arial',size=11,bold=True,color='FFFFFFFF')
            ws.cell(row=row_num, column=col_num + 1).fill = PatternFill(fill_type='solid',start_color='FF4C666E',end_color='FF000000')
            ws.cell(row=row_num, column=col_num + 1).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            ws.cell(row=row_num, column=col_num + 1).number_format = 'General'



        for contenido in contenidos:
            row_num += 1
            for col_num in range(len(contenido)):


                if isinstance(contenido[col_num], tuple):
                    if contenido[col_num][0] == True and contenido[col_num][0] != 1:
                        ws.cell(row=row_num,column=col_num+1).value = "SI"
                        ws.cell(row=row_num, column=col_num + 1).hyperlink = contenido[col_num][1]
                    elif contenido[col_num][0] == False and contenido[col_num][0] != 0:
                        ws.cell(row=row_num,column=col_num+1).value = "NO"
                        ws.cell(row=row_num, column=col_num + 1).hyperlink = contenido[col_num][1]
                    elif contenido[col_num][0] == None:
                        ws.cell(row=row_num,column=col_num+1).value = ""
                    else:
                        ws.cell(row=row_num,column=col_num+1).value = contenido[col_num][0]
                        ws.cell(row=row_num, column=col_num + 1).hyperlink = contenido[col_num][1]

                    ws.cell(row=row_num, column=col_num + 1).font = Font(name='Arial', size=10, bold=True,color='FF4C666E')
                    ws.cell(row=row_num, column=col_num + 1).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                    ws.cell(row=row_num, column=col_num + 1).number_format = formatos[col_num][0]


                elif isinstance(contenido[col_num], list):

                    if len(contenido[col_num]) == 4:

                        if contenido[col_num][0] == True and contenido[col_num][0] != 1:
                            ws.cell(row=row_num,column=col_num+1).value = "SI"
                            ws.cell(row=row_num, column=col_num + 1).comment = Comment(contenido[col_num][-1],'')
                        elif contenido[col_num][0] == False and contenido[col_num][0] != 0:
                            ws.cell(row=row_num,column=col_num+1).value = "NO"
                            ws.cell(row=row_num, column=col_num + 1).comment = Comment(contenido[col_num][-1],'')
                        elif contenido[col_num][0] == None:
                            ws.cell(row=row_num,column=col_num+1).value = ""
                            ws.cell(row=row_num, column=col_num + 1).comment = Comment(contenido[col_num][-1],'')
                        else:
                            ws.cell(row=row_num,column=col_num+1).value = contenido[col_num][0]
                            ws.cell(row=row_num, column=col_num + 1).comment = Comment(contenido[col_num][-1],'')


                        if contenido[col_num][1] == 'FFFFFFFF':

                            ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10)
                            ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                            ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num][0]

                        else:

                            ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10, bold=True, color=contenido[col_num][2])
                            ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                            ws.cell(row=row_num, column=col_num + 1).fill=PatternFill(fill_type='solid',start_color=contenido[col_num][1],end_color=contenido[col_num][1])
                            ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num][0]


                    else:
                        if contenido[col_num][0] == True and contenido[col_num][0] != 1:
                            ws.cell(row=row_num, column=col_num + 1).value = "SI"
                        elif contenido[col_num][0] == False and contenido[col_num][0] != 0:
                            ws.cell(row=row_num, column=col_num + 1).value = "NO"
                        elif contenido[col_num][0] == None:
                            ws.cell(row=row_num, column=col_num + 1).value = ""
                        else:
                            ws.cell(row=row_num, column=col_num + 1).value = contenido[col_num][0]

                        if contenido[col_num][1] == 'FFFFFFFF':

                            ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10)
                            ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                            ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num][0]


                        else:

                            ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10, bold=True, color=contenido[col_num][2])
                            ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                            ws.cell(row=row_num, column=col_num + 1).fill=PatternFill(fill_type='solid',start_color=contenido[col_num][1],end_color=contenido[col_num][1])
                            ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num][0]

                else:

                    if contenido[col_num] == True and contenido[col_num] != 1:
                        ws.cell(row=row_num,column=col_num+1).value = "SI"
                    elif contenido[col_num] == False and contenido[col_num] != 0:
                        ws.cell(row=row_num,column=col_num+1).value = "NO"
                    elif contenido[col_num] == None:
                        ws.cell(row=row_num,column=col_num+1).value = ""
                    else:
                        ws.cell(row=row_num,column=col_num+1).value = contenido[col_num]

                    ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10)
                    ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                    ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num]


        wb.save(output)
        return output

def construir_reporte_pagina(output_in,sheet_name,titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso):
    if len(ancho_columnas) != len(formatos) != len(titulos) != len(contenidos[0]):
        raise Exception('El arreglo de filas y columnas tienen distinta longitud')
    else:
        output = BytesIO()
        wb = openpyxl.load_workbook(output_in)
        ws = wb.create_sheet(sheet_name)

        ws2 = wb.get_sheet_by_name('Hoja1')
        #logo_sican = Image(settings.STATICFILES_DIRS[0] + '/img/icon-192.png', size=(151, 151))
        #logo_sican.drawing.top = 0
        #logo_sican.drawing.left = 0
        #ws2.add_image(logo_sican)


        row_num = 1
        for col_num in range(len(titulos)):
            ws.cell(row=row_num, column=col_num+1).value = titulos[col_num]
            ws.column_dimensions[get_column_letter(col_num+1)].width = ancho_columnas[col_num]

            ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial',size=11,bold=True,color='FFFFFFFF')
            ws.cell(row=row_num, column=col_num + 1).fill=PatternFill(fill_type='solid',start_color='ef6c00',end_color='FF000000')
            ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            ws.cell(row=row_num, column=col_num + 1).number_format='General'


        for contenido in contenidos:
            row_num += 1
            for col_num in range(len(contenido)):


                if isinstance(contenido[col_num], tuple):
                    if contenido[col_num][0] == True and contenido[col_num][0] != 1:
                        ws.cell(row=row_num,column=col_num+1).value = "SI"
                        ws.cell(row=row_num, column=col_num + 1).hyperlink = contenido[col_num][1]
                    elif contenido[col_num][0] == False and contenido[col_num][0] != 0:
                        ws.cell(row=row_num,column=col_num+1).value = "NO"
                        ws.cell(row=row_num, column=col_num + 1).hyperlink = contenido[col_num][1]
                    elif contenido[col_num][0] == None:
                        ws.cell(row=row_num,column=col_num+1).value = ""
                    else:
                        ws.cell(row=row_num,column=col_num+1).value = contenido[col_num][0]
                        ws.cell(row=row_num, column=col_num + 1).hyperlink = contenido[col_num][1]

                    ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10, bold=True,color='FF4C666E')
                    ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                    ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num][0]


                elif isinstance(contenido[col_num], list):

                    if len(contenido[col_num]) == 4:

                        if contenido[col_num][0] == True and contenido[col_num][0] != 1:
                            ws.cell(row=row_num,column=col_num+1).value = "SI"
                            ws.cell(row=row_num, column=col_num + 1).comment = Comment(contenido[col_num][-1], '')
                        elif contenido[col_num][0] == False and contenido[col_num][0] != 0:
                            ws.cell(row=row_num,column=col_num+1).value = "NO"
                            ws.cell(row=row_num, column=col_num + 1).comment = Comment(contenido[col_num][-1], '')
                        elif contenido[col_num][0] == None:
                            ws.cell(row=row_num,column=col_num+1).value = ""
                            ws.cell(row=row_num, column=col_num + 1).comment = Comment(contenido[col_num][-1], '')
                        else:
                            ws.cell(row=row_num,column=col_num+1).value = contenido[col_num][0]
                            ws.cell(row=row_num, column=col_num + 1).comment = Comment(contenido[col_num][-1], '')


                        if contenido[col_num][1] == 'FFFFFFFF':

                            ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10)
                            ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                            ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num][0]

                        else:

                            ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10, bold=True, color=contenido[col_num][2])
                            ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                            ws.cell(row=row_num, column=col_num + 1).fill=PatternFill(fill_type='solid',start_color=contenido[col_num][1],end_color=contenido[col_num][1])
                            ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num][0]

                    else:
                        if contenido[col_num][0] == True and contenido[col_num][0] != 1:
                            ws.cell(row=row_num, column=col_num + 1).value = "SI"
                        elif contenido[col_num][0] == False and contenido[col_num][0] != 0:
                            ws.cell(row=row_num, column=col_num + 1).value = "NO"
                        elif contenido[col_num][0] == None:
                            ws.cell(row=row_num, column=col_num + 1).value = ""
                        else:
                            ws.cell(row=row_num, column=col_num + 1).value = contenido[col_num][0]

                        if contenido[col_num][1] == 'FFFFFFFF':

                            ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10)
                            ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                            ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num][0]

                        else:

                            ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10, bold=True, color=contenido[col_num][2])
                            ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                            ws.cell(row=row_num, column=col_num + 1).fill=PatternFill(fill_type='solid',start_color=contenido[col_num][1],end_color=contenido[col_num][1])
                            ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num][0]


                else:

                    if contenido[col_num] == True and contenido[col_num] != 1:
                        ws.cell(row=row_num,column=col_num+1).value = "SI"
                    elif contenido[col_num] == False and contenido[col_num] != 0:
                        ws.cell(row=row_num,column=col_num+1).value = "NO"
                    elif contenido[col_num] == None:
                        ws.cell(row=row_num,column=col_num+1).value = ""
                    else:
                        ws.cell(row=row_num,column=col_num+1).value = contenido[col_num]


                    ws.cell(row=row_num, column=col_num + 1).font=Font(name='Arial', size=10)
                    ws.cell(row=row_num, column=col_num + 1).alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
                    ws.cell(row=row_num, column=col_num + 1).number_format=formatos[col_num]


        wb.save(output)
        return output
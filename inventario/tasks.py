from io import BytesIO

import openpyxl
from django.conf import settings
from django.core.files import File
from django.db.models import Sum

from config.celery import app
from config.functions import construir_reporte
from inventario.models import Despachos, Sustracciones, Productos, Adiciones
from reportes import models as models_reportes
from reportes.models import Reportes


@app.task
def build_remision(id):
    despacho = Despachos.objects.get(id=id)
    sustracciones = Sustracciones.objects.filter(despacho = despacho)


    if sustracciones.count() > 0 and sustracciones.count() <= 19:

        cantidad = Sustracciones.objects.filter(despacho = despacho).count()

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/despacho_1.xlsx')
        ws = wb.get_sheet_by_name('remision')
        #logo_sican.drawing.top = 46
        #logo_sican.drawing.left = 66

        ws.merge_cells('G3:J5')
        top_left_cell = ws['G3']
        top_left_cell.value = str(despacho.consecutivo)

        ws['B6'] = str(despacho.cliente.get_nombre_completo())
        ws['D6'] = str(despacho.cliente.ciudad.nombre)
        ws['H6'] = str(despacho.pretty_fecha_envio_datetime())

        ws['B7'] = str(despacho.cliente.documento)
        ws['D7'] = str(despacho.cliente.direccion)

        if despacho.transportador != None and despacho.transportador != "":
            ws['B8'] = str(despacho.transportador)
        if despacho.conductor != None and despacho.conductor != "":
            ws['E8'] = str(despacho.conductor)
        if despacho.placa != None and despacho.placa != "":
            ws['J8'] = str(despacho.placa)

        i = 12
        contador = 1

        for sustraccion in sustracciones:


            ws['A' + str(i)] = contador
            ws['B' + str(i)] = sustraccion.producto.nombre
            ws['F' + str(i)] = sustraccion.producto.get_marca()
            ws['G' + str(i)] = sustraccion.cantidad
            ws['H' + str(i)] = str(str(sustraccion.producto.valor).replace('COL',''))
            ws['I' + str(i)] = str(str(sustraccion.producto.impuesto) + "%")
            ws['J' + str(i)] = str(sustraccion.pretty_print_valor_total())

            i += 1
            contador += 1

        total = sustracciones.aggregate(Sum('valor_total'))['valor_total__sum']
        total = "${:,.2f}".format(total)
        ws['J33'] = total

        ws['A31'] = despacho.observacion

        wb.save(output)

        filename = str(despacho.id) + '.xlsx'
        despacho.respaldo.save(filename, File(output))

    elif sustracciones.count() > 19 and sustracciones.count() <= 41:

        cantidad = Sustracciones.objects.filter(despacho = despacho).count()

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/despacho_2.xlsx')
        ws = wb.get_sheet_by_name('remision')
        #logo_sican.drawing.top = 46
        #logo_sican.drawing.left = 66

        ws.merge_cells('G3:J5')
        top_left_cell = ws['G3']
        top_left_cell.value = str(despacho.consecutivo)

        ws.merge_cells('G40:J42')
        top_left_cell = ws['G40']
        top_left_cell.value = str(despacho.consecutivo)

        ws['B6'] = str(despacho.nombre_cliente)
        ws['D6'] = str(despacho.ciudad.nombre)
        ws['H6'] = str(despacho.pretty_fecha_envio_datetime())

        ws['B7'] = str(despacho.documento)
        ws['D7'] = str(despacho.direccion)
        ws['D7'] = str(despacho.direccion)

        if despacho.transportador != None and despacho.transportador != "":
            ws['B8'] = str(despacho.transportador)
        if despacho.conductor != None and despacho.conductor != "":
            ws['E8'] = str(despacho.conductor)
        if despacho.placa != None and despacho.placa != "":
            ws['J8'] = str(despacho.placa)

        i = 12
        contador = 1

        for sustraccion in sustracciones:


            ws['A' + str(i)] = contador
            ws['B' + str(i)] = sustraccion.producto.nombre
            ws['F' + str(i)] = sustraccion.producto.get_marca()
            ws['G' + str(i)] = sustraccion.cantidad
            ws['H' + str(i)] = str(str(sustraccion.producto.valor).replace('COL',''))
            ws['I' + str(i)] = str(str(sustraccion.producto.impuesto) + "%")
            ws['J' + str(i)] = str(sustraccion.pretty_print_valor_total())

            i += 1
            contador += 1
            if contador == 20:
                i = i + 16

        total = sustracciones.aggregate(Sum('valor_total'))['valor_total__sum']
        total = "${:,.2f}".format(total)
        ws['J70'] = total

        ws['A68'] = despacho.observacion

        wb.save(output)

        filename = str(despacho.id) + '.xlsx'
        despacho.respaldo.save(filename, File(output))

    elif sustracciones.count() > 42 and sustracciones.count() <= 63:

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/despacho_3.xlsx')
        ws = wb.get_sheet_by_name('remision')
        #logo_sican.drawing.top = 46
        #logo_sican.drawing.left = 66

        ws.merge_cells('G3:J5')
        top_left_cell = ws['G3']
        top_left_cell.value = str(despacho.consecutivo)

        ws.merge_cells('G40:J42')
        top_left_cell = ws['G40']
        top_left_cell.value = str(despacho.consecutivo)

        ws.merge_cells('G77:J70')
        top_left_cell = ws['G77']
        top_left_cell.value = str(despacho.consecutivo)

        ws['B6'] = str(despacho.nombre_cliente)
        ws['D6'] = str(despacho.ciudad.nombre)
        ws['H6'] = str(despacho.pretty_fecha_envio_datetime())

        ws['B7'] = str(despacho.documento)
        ws['D7'] = str(despacho.direccion)
        ws['D7'] = str(despacho.direccion)

        if despacho.transportador != None and despacho.transportador != "":
            ws['B8'] = str(despacho.transportador)
        if despacho.conductor != None and despacho.conductor != "":
            ws['E8'] = str(despacho.conductor)
        if despacho.placa != None and despacho.placa != "":
            ws['J8'] = str(despacho.placa)

        i = 12
        contador = 1

        for sustraccion in sustracciones:


            ws['A' + str(i)] = contador
            ws['B' + str(i)] = sustraccion.producto.nombre
            ws['F' + str(i)] = sustraccion.producto.get_marca()
            ws['G' + str(i)] = sustraccion.cantidad
            ws['H' + str(i)] = str(str(sustraccion.producto.valor).replace('COL',''))
            ws['I' + str(i)] = str(str(sustraccion.producto.impuesto) + "%")
            ws['J' + str(i)] = str(sustraccion.pretty_print_valor_total())

            i += 1
            contador += 1
            if contador == 20:
                i = i + 16
            if contador == 22:
                i = i + 16

        total = sustracciones.aggregate(Sum('valor_total'))['valor_total__sum']
        total = "${:,.2f}".format(total)
        ws['J107'] = total

        ws['A105'] = despacho.observacion

        wb.save(output)

        filename = str(despacho.id) + '.xlsx'
        despacho.respaldo.save(filename, File(output))

    return "Reporte generado"

@app.task
def build_remision_proyect(id):
    despacho = Despachos.objects.get(id=id)
    sustracciones = Sustracciones.objects.filter(despacho = despacho)


    if sustracciones.count() > 0 and sustracciones.count() <= 19:

        cantidad = Sustracciones.objects.filter(despacho = despacho).count()

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/despacho_sin_valor_1.xlsx')
        ws = wb.get_sheet_by_name('remision')
        #logo_sican.drawing.top = 46
        #logo_sican.drawing.left = 66

        ws.merge_cells('G3:J5')
        top_left_cell = ws['G3']
        top_left_cell.value = str(despacho.consecutivo)

        ws['B6'] = str(despacho.cliente.get_nombre_completo())
        ws['D6'] = str(despacho.cliente.ciudad.nombre)
        ws['H6'] = str(despacho.pretty_fecha_envio_datetime())

        ws['B7'] = str(despacho.cliente.documento)
        ws['D7'] = str(despacho.cliente.direccion)

        if despacho.transportador != None and despacho.transportador != "":
            ws['B8'] = str(despacho.transportador)
        if despacho.conductor != None and despacho.conductor != "":
            ws['E8'] = str(despacho.conductor)
        if despacho.placa != None and despacho.placa != "":
            ws['J8'] = str(despacho.placa)

        i = 12
        contador = 1

        for sustraccion in sustracciones:


            ws['A' + str(i)] = contador
            ws['B' + str(i)] = sustraccion.producto.nombre
            ws['H' + str(i)] = sustraccion.producto.nombre
            ws['J' + str(i)] = sustraccion.cantidad

            i += 1
            contador += 1

        total = sustracciones.aggregate(Sum('valor_total'))['valor_total__sum']
        total = "${:,.2f}".format(total)
        ws['J33'] = total

        ws['A31'] = despacho.observacion

        wb.save(output)

        filename = str(despacho.id) + '.xlsx'
        despacho.respaldo.save(filename, File(output))

    elif sustracciones.count() > 19 and sustracciones.count() <= 41:

        cantidad = Sustracciones.objects.filter(despacho = despacho).count()

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/despacho_sin_valor_2.xlsx')
        ws = wb.get_sheet_by_name('remision')
        #logo_sican.drawing.top = 46
        #logo_sican.drawing.left = 66

        ws.merge_cells('G3:J5')
        top_left_cell = ws['G3']
        top_left_cell.value = str(despacho.consecutivo)

        ws.merge_cells('G40:J42')
        top_left_cell = ws['G40']
        top_left_cell.value = str(despacho.consecutivo)

        ws['B6'] = str(despacho.nombre_cliente)
        ws['D6'] = str(despacho.ciudad.nombre)
        ws['H6'] = str(despacho.pretty_fecha_envio_datetime())

        ws['B7'] = str(despacho.documento)
        ws['D7'] = str(despacho.direccion)
        ws['D7'] = str(despacho.direccion)

        if despacho.transportador != None and despacho.transportador != "":
            ws['B8'] = str(despacho.transportador)
        if despacho.conductor != None and despacho.conductor != "":
            ws['E8'] = str(despacho.conductor)
        if despacho.placa != None and despacho.placa != "":
            ws['J8'] = str(despacho.placa)

        i = 12
        contador = 1

        for sustraccion in sustracciones:


            ws['A' + str(i)] = contador
            ws['B' + str(i)] = sustraccion.producto.nombre
            ws['H' + str(i)] = sustraccion.producto.nombre
            ws['J' + str(i)] = sustraccion.cantidad
            i += 1
            contador += 1
            if contador == 20:
                i = i + 16

        total = sustracciones.aggregate(Sum('valor_total'))['valor_total__sum']
        total = "${:,.2f}".format(total)
        ws['J70'] = total

        ws['A68'] = despacho.observacion

        wb.save(output)

        filename = str(despacho.id) + '.xlsx'
        despacho.respaldo.save(filename, File(output))

    elif sustracciones.count() > 42 and sustracciones.count() <= 63:

        output = BytesIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/despacho_sin_valor_3.xlsx')
        ws = wb.get_sheet_by_name('remision')
        #logo_sican.drawing.top = 46
        #logo_sican.drawing.left = 66

        ws.merge_cells('G3:J5')
        top_left_cell = ws['G3']
        top_left_cell.value = str(despacho.consecutivo)

        ws.merge_cells('G40:J42')
        top_left_cell = ws['G40']
        top_left_cell.value = str(despacho.consecutivo)

        ws.merge_cells('G77:J70')
        top_left_cell = ws['G77']
        top_left_cell.value = str(despacho.consecutivo)

        ws['B6'] = str(despacho.nombre_cliente)
        ws['D6'] = str(despacho.ciudad.nombre)
        ws['H6'] = str(despacho.pretty_fecha_envio_datetime())

        ws['B7'] = str(despacho.documento)
        ws['D7'] = str(despacho.direccion)
        ws['D7'] = str(despacho.direccion)

        if despacho.transportador != None and despacho.transportador != "":
            ws['B8'] = str(despacho.transportador)
        if despacho.conductor != None and despacho.conductor != "":
            ws['E8'] = str(despacho.conductor)
        if despacho.placa != None and despacho.placa != "":
            ws['J8'] = str(despacho.placa)

        i = 12
        contador = 1

        for sustraccion in sustracciones:


            ws['A' + str(i)] = contador
            ws['B' + str(i)] = sustraccion.producto.nombre
            ws['H' + str(i)] = sustraccion.producto.nombre
            ws['J' + str(i)] = sustraccion.cantidad

            i += 1
            contador += 1
            if contador == 20:
                i = i + 16
            if contador == 22:
                i = i + 16

        total = sustracciones.aggregate(Sum('valor_total'))['valor_total__sum']
        total = "${:,.2f}".format(total)
        ws['J107'] = total

        ws['A105'] = despacho.observacion

        wb.save(output)

        filename = str(despacho.id) + '.xlsx'
        despacho.respaldo.save(filename, File(output))

    return "Reporte generado"

@app.task
def build_list_reports(id):
    reporte = Reportes.objects.get(id=id)
    productos = Productos.objects.all()
    proceso = "SICAN-LIST-PRODUCTOS"


    titulos = ['Consecutivo', 'Código','Nombre', 'Valor' ,'Stock']

    formatos = ['0', 'General','General', '"$"#,##0_);("$"#,##0)','0']

    ancho_columnas = [20, 30, 30, 40, 30]

    contenidos = []
    order = []


    i = 0
    for producto in productos:
        i += 1


        lista = [
            int(i),
            producto.codigo,
            producto.nombre,
            producto.pretty_print_precio(),
            producto.stock,
        ]

        contenidos.append(lista)


    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))


    return "Archivo paquete ID: " + filename

@app.task
def build_reporte_cargue(reporte_id):
    reporte = models_reportes.Reportes.objects.get(id = reporte_id)

    proceso = "UNI2DATA-REPORTE-CARGUES"


    titulos = ['Consecutivo', 'Fecha creación', 'Cargue','Observacion general', 'Producto', 'Codigo', 'Valor Compra', 'Cantidad',
               'Estado','Observacion']

    formatos = ['0', 'dd/mm/yy', '0','General', 'General', 'General', '"$"#,##0.00_);[Red]("$"#,##0.00)', '0',
                 'General','General']

    ancho_columnas = [20, 30, 30, 40, 30, 30, 30, 30, 30, 30]

    contenidos = []

    i = 0
    for adicion in Adiciones.objects.all().order_by('creacion'):
        i += 1
        contenidos.append([
            int(i),
            adicion.creacion,
            adicion.cargue.consecutivo,
            adicion.cargue.observacion,
            adicion.producto.nombre,
            adicion.producto.codigo,
            adicion.producto.valor_compra.amount,
            adicion.cantidad,
            adicion.cargue.estado,
            adicion.observacion,
        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))


    return "Archivo paquete ID: " + filename

@app.task
def build_reporte_despacho(reporte_id):
    reporte = models_reportes.Reportes.objects.get(id = reporte_id)

    proceso = "UNI2DATA-REPORTE-DESPACHO"


    titulos = ['Consecutivo', 'Fecha creación', 'Cargue','Observacion general', 'Producto', 'Codigo', 'Valor Compra','Marca', 'Cantidad',
               'Estado','Observacion','Fecha de envio','Proyecto']

    formatos = ['0', 'dd/mm/yy', '0','General', 'General', 'General', '"$"#,##0.00_);[Red]("$"#,##0.00)','General','0',
                 'General','General','dd/mm/yy','General']

    ancho_columnas = [20, 30, 30, 40, 30, 30, 30, 30, 30, 30, 30, 30, 30]

    contenidos = []

    i = 0
    for sustraccion in Sustracciones.objects.all().order_by('creacion'):
        i += 1
        contenidos.append([
            int(i),
            sustraccion.creacion,
            sustraccion.despacho.consecutivo,
            sustraccion.despacho.observacion,
            sustraccion.producto.nombre,
            sustraccion.producto.codigo,
            sustraccion.producto.valor_compra.amount,
            sustraccion.producto.marca,
            sustraccion.cantidad,
            sustraccion.despacho.estado,
            sustraccion.observacion,
            sustraccion.despacho.fecha_envio,
            sustraccion.despacho.get_proyect(),

        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))


    return "Archivo paquete ID: " + filename
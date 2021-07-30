from django.db import models
import uuid
from usuarios.models import Municipios, User, Departamentos, Corregimientos, Veredas, PueblosIndigenas, \
    ResguardosIndigenas, ComunidadesIndigenas, LenguasNativas, ConsejosAfro, ComunidadesAfro, CategoriaDiscapacidad, \
    DificultadesPermanentesDiscapacidad, ElementosDiscapacidad, TiposRehabilitacionDiscapacidad, ConsejosResguardosProyectosIraca, \
    ComunidadesProyectosIraca

from recursos_humanos import models as models_rh
from djmoney.models.fields import MoneyField
from django.db.models import Sum
from direccion_financiera.models import Bancos
from config.extrafields import ContentTypeRestrictedFileField
from django.db.models.signals import post_save
from django.dispatch import receiver
from pytz import timezone
from django.conf import settings
import json
from delta import html
from django.contrib.postgres.fields import JSONField
import datetime
import openpyxl
from io import BytesIO
from django.core.files import File
from openpyxl.drawing.image import Image
import dateutil.parser
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy


settings_time_zone = timezone(settings.TIME_ZONE)

# Create your models here.

class Componentes(models.Model):

    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    nombre = models.CharField(max_length=100)
    consecutivo = models.IntegerField()
    momentos = models.IntegerField()
    valor = models.IntegerField(default=0)
    ruta = models.TextField(default='')
    valor_pagado = models.IntegerField(default=0)

    def __str__(self):
        return self.nombre

    def get_numero_momentos(self):
        return Momentos.objects.filter(componente=self).count()


    def get_cantidad_instrumentos(self,hogar):
        return InstrumentosRutaObject.objects.filter(hogar = hogar,momento__componente = self).count()


    def get_valor_hogar(self, hogar):
        valor = CuposRutaObject.objects.filter(hogar=hogar, momento__componente = self,estado__in=['Reportado','Pagado']).aggregate(Sum('valor'))['valor__sum']
        return valor if valor != None else 0


    def get_ruta_hogar_componente(self, hogar):

        if self.consecutivo == 1:
            return hogar.ruta_1
        elif self.consecutivo == 2:
            return hogar.ruta_2
        elif self.consecutivo == 3:
            return hogar.ruta_3
        elif self.consecutivo == 4:
            return hogar.ruta_4
        else:
            return None

class Momentos(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    componente = models.ForeignKey(Componentes,on_delete=models.DO_NOTHING,related_name='momentos_componente')
    nombre = models.CharField(max_length=100)
    consecutivo = models.IntegerField()
    instrumentos = models.IntegerField()
    tipo = models.CharField(max_length=100)
    valor_maximo = models.BooleanField(default=True)
    novedades = models.BooleanField(default=True)
    progreso = models.BooleanField(default=True)

    def __str__(self):
        return '{0} - {1}'.format(self.componente.nombre,self.nombre)


    def get_objetos_ruta(self,ruta):
        return CuposRutaObject.objects.filter(momento = self, ruta = ruta).count()


    def get_cantidad_instrumentos(self,hogar,componente):
        return InstrumentosRutaObject.objects.filter(hogar = hogar,momento__componente = componente, momento = self).count()

    def get_novedades_mis_rutas_actividades(self,ruta):
        return InstrumentosRutaObject.objects.filter(ruta = ruta, momento = self, estado = 'cargado').distinct().count()

    def get_numero_instrumentos(self):
        return Instrumentos.objects.filter(momento=self).count()

    def get_consecutivo(self):
        return '{0}.{1}'.format(self.componente.consecutivo,self.consecutivo)

    def get_valor_maximo_momento(self,ruta):



        try:
            data = json.loads(ruta.valores_actividades)
        except:
            data = []


        if 'valor_' + str(self.id) in data:
            valor = data['valor_' + str(self.id)].replace('$ ','').replace(',','')

        else:
            valor = 0

        return float(valor)


    def get_cantidad_momento(self,ruta):



        try:
            data = json.loads(ruta.valores_actividades)
        except:
            data = []


        if 'cantidad_' + str(self.id) in data:
            cantidad = data['cantidad_' + str(self.id)]

        else:
            cantidad = 0

        return cantidad


    def get_valor_maximo_momento_corte(self,ruta,corte):
        valor = CuposRutaObject.objects.filter(ruta = ruta, momento = self, corte = corte).aggregate(Sum('valor'))['valor__sum']
        if valor == None:
            valor = 0
        return float(valor)


    def get_valor_reportado_momento(self,ruta):
        valor = CuposRutaObject.objects.filter(ruta = ruta, momento = self, estado = 'Reportado').aggregate(Sum('valor'))['valor__sum']
        if valor == None:
            valor = 0
        return float(valor)

    def get_valor_pagado_momento(self,ruta):
        valor = CuposRutaObject.objects.filter(ruta = ruta, momento = self, estado = 'Pagado').aggregate(Sum('valor'))['valor__sum']
        if valor == None:
            valor = 0
        return float(valor)

    def get_progreso_momento(self,ruta):
        valor_maximo = self.get_valor_maximo_momento(ruta)
        valor_reportado = self.get_valor_reportado_momento(ruta)
        valor_pagado = self.get_valor_pagado_momento(ruta)

        if valor_maximo != 0:
            progreso = ((valor_reportado + valor_pagado)/valor_maximo)*100.0
        else:
            progreso = 0.0

        return (progreso,valor_maximo,valor_reportado,valor_pagado)


    def get_valor_pagado(self,hogar):
        query = CuposRutaObject.objects.filter(momento = self, estado__in = ['Reportado','Pagado'], hogar = hogar)
        valor = query.aggregate(Sum('valor'))['valor__sum']
        if valor == None:
            valor = 0

        return valor

class Instrumentos(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    momento = models.ForeignKey(Momentos,on_delete=models.DO_NOTHING,related_name='instrumento_omento')
    nombre = models.CharField(max_length=100)
    short_name = models.CharField(max_length=100)
    consecutivo = models.IntegerField()
    modelo = models.CharField(max_length=100)
    color = models.CharField(max_length=100)
    icon = models.CharField(max_length=100)
    nivel = models.CharField(max_length=100)

    def __str__(self):
        return self.nombre

    def get_consecutivo(self):
        return '{0}.{1}.{2}'.format(self.momento.componente.consecutivo,self.momento.consecutivo,self.consecutivo)

class Rutas(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    creation = models.DateTimeField(auto_now_add=True)
    nombre = models.CharField(unique=True, max_length=100)
    contrato = models.ForeignKey(models_rh.Contratos, on_delete=models.DO_NOTHING,
                                 related_name='contrato_ruta_fest_2019')
    valor = MoneyField(max_digits=10, decimal_places=2, default_currency='COP',default=0)
    valor_transporte = MoneyField(max_digits=10, decimal_places=2, default_currency='COP', default=0)
    valor_otros = MoneyField(max_digits=10, decimal_places=2, default_currency='COP', default=0)

    novedades = models.IntegerField(default=0)
    progreso = models.DecimalField(max_digits=10, decimal_places=2,default=0.0)
    usuario_creacion = models.ForeignKey(User, related_name="usuario_creacion_ruta_fest_2019", on_delete=models.DO_NOTHING)
    update_datetime = models.DateTimeField(auto_now=True)
    usuario_actualizacion = models.ForeignKey(User, related_name="usuario_actualizacion_ruta_fest_2019",
                                              on_delete=models.DO_NOTHING,
                                              blank=True, null=True)

    estado = models.CharField(max_length=100,blank=True)
    componente = models.ForeignKey(Componentes, on_delete=models.DO_NOTHING,blank=True,null=True)

    hogares_inscritos = models.IntegerField(default=0)

    tipo_pago = models.CharField(max_length=100)
    valores_actividades = models.TextField()


    def __str__(self):
        return self.nombre




    def get_aprobable_valor(self, momento):

        dict = {'aprobable': 'no','valor': 0}

        if self.tipo_pago == 'actividad':

            valores_actividades = json.loads(self.valores_actividades)

            try:
                cantidad = int(valores_actividades['cantidad_' + str(momento.id)])
            except:
                cantidad = 0

            try:
                valor = float(valores_actividades['valor_' + str(momento.id)].replace("$ ",'').replace(',',''))
            except:
                valor = float(0)


            objetos_aprobados = CuposRutaObject.objects.filter(ruta = self, momento = momento, estado__in = ["aprobado","Reportado","Pagado"])
            valor_aprobados = objetos_aprobados.aggregate(Sum('valor'))['valor__sum']

            if valor_aprobados == None:
                valor_aprobados = float(0)
            else:
                valor_aprobados = float(valor_aprobados)

            if objetos_aprobados.count() < cantidad:

                try:
                    valor_actividad = (valor - valor_aprobados) / (cantidad - objetos_aprobados.count())
                except:
                    valor_actividad = 0


                dict['aprobable']= 'si'
                dict['valor']= valor_actividad



        else:
            pass

        return dict





    def translado(self, hogar, componente):

        update = False

        momentos_ids = []

        for cupo in CuposRutaObject.objects.filter(hogar = hogar, estado__in = ['Pagado','Reportado'], translado = False,momento__componente = componente).exclude(ruta=self):

            if cupo.momento.id not in momentos_ids:

                update = True

                CuposRutaObject.objects.get_or_create(
                    ruta = self,
                    momento = cupo.momento,
                    tipo = cupo.tipo,
                    estado = 'Pagado',
                    valor = 0,
                    hogar = cupo.hogar,
                    translado = False
                )

            cupo.translado = True
            cupo.save()

        ruta = None

        if componente.consecutivo == 1:

            if hogar.ruta_1 != None:
                ruta = hogar.ruta_1

            hogar.ruta_1 = self
            hogar.save()

        elif componente.consecutivo == 2:

            if hogar.ruta_2 != None:
                ruta = hogar.ruta_2

            hogar.ruta_2 = self
            hogar.save()

        elif componente.consecutivo == 3:

            if hogar.ruta_3 != None:
                ruta = hogar.ruta_3

            hogar.ruta_3 = self
            hogar.save()

        elif componente.consecutivo == 4:

            if hogar.ruta_4 != None:
                ruta = hogar.ruta_4

            hogar.ruta_4 = self
            hogar.save()

        if ruta != None:
            ruta.update_hogares_inscritos()
            if update:
                ruta.actualizar_objetos()

        return update



    def translado_vinculacion(self, hogar):

        update = False

        momentos_ids = []

        for cupo in CuposRutaObject.objects.filter(hogar = hogar, estado__in = ['Pagado','Reportado'], translado = False, momento__tipo = 'vinculacion').exclude(ruta=self):

            if cupo.momento.id not in momentos_ids:

                update = True

                CuposRutaObject.objects.get_or_create(
                    ruta = self,
                    momento = cupo.momento,
                    tipo = cupo.tipo,
                    estado = 'Pagado',
                    valor = 0,
                    hogar = cupo.hogar,
                    translado = False
                )

            cupo.translado = True
            cupo.save()

        ruta = None


        if hogar.ruta_vinculacion != None:
            ruta = hogar.ruta_vinculacion

        hogar.ruta_vinculacion = self
        hogar.save()



        if ruta != None:
            ruta.update_hogares_inscritos()
            if update:
                ruta.actualizar_objetos()

        return update


    def update_visita_1(self):

        momento = Momentos.objects.filter(componente = self.componente).get(nombre = "Visita 1")
        valor_total = CuposRutaObject.objects.filter(ruta = self,momento = momento).aggregate(Sum('valor'))['valor__sum']
        CuposRutaObject.objects.filter(ruta = self, momento= momento, estado = "asignado").delete()
        asignados = CuposRutaObject.objects.filter(ruta=self,momento=momento).count()

        for i in range(0,self.meta_vinculacion-asignados):

            CuposRutaObject.objects.create(
                ruta=self,
                momento=momento,
                tipo=momento.tipo,
                estado='asignado',
                valor=0
            )

        objetos = CuposRutaObject.objects.filter(ruta=self,momento=momento)
        valor = valor_total/objetos.count()
        objetos.update(valor=valor)

        return '{0} - {1}'.format(self.meta_vinculacion,objetos.count())

    def update_progreso(self):

        dict = json.loads(self.valores_actividades)
        cantidad = 0
        progreso = 0

        for momento in Momentos.objects.filter(componente = self.componente):
            cantidad += dict['cantidad_' + str(momento.id)]


        cupos = CuposRutaObject.objects.filter(ruta = self)
        revisados = cupos.filter(estado__in = ['Reportado','Pagado'])

        try:
            progreso = (revisados.count()/cantidad)*100.0
        except:
            pass

        self.progreso = progreso
        self.save()

        return self.progreso

    def set_novedades_ruta(self):
        novedades = 0
        for momento in Momentos.objects.filter(componente = self.componente):
            novedades += momento.get_novedades_mis_rutas_actividades(self)
        Rutas.objects.filter(id = self.id).update(novedades = novedades)
        self.update_progreso()
        return novedades

    def update_novedades(self):
        self.novedades = InstrumentosRutaObject.objects.filter(ruta=self, estado='cargado').count()
        self.save()
        self.update_progreso()
        return 'Ok'

    def update_hogares_inscritos(self):

        objetos_hogares = Hogares.objects.filter(rutas = self)

        self.hogares_inscritos = objetos_hogares.count()
        self.save()

        return None

    def crear_objetos_momento(self,momento,procesados,meta,valor,cantidad_momentos,meta_vinculacion):

        valor = valor/cantidad_momentos

        valor_procesados = procesados.filter(momento=momento).aggregate(Sum('valor'))['valor__sum']
        if valor_procesados == None:
            valor_procesados = 0

        nueva_meta = (meta - procesados.filter(momento=momento).exclude(translado = True).count())



        try:
            nuevo_valor = (valor - float(valor_procesados)) / (nueva_meta)
        except:
            return (0,0)

        else:
            if momento.nombre == 'Visita 1':

                nueva_meta_visita_1 = (meta_vinculacion - procesados.filter(momento=momento).count())

                try:
                    nuevo_valor_visita_1 = (valor - float(valor_procesados)) / (nueva_meta_visita_1)
                except:
                    return (0, 0)

                for i in range(0, nueva_meta_visita_1):
                    CuposRutaObject.objects.create(
                        ruta=self,
                        momento=momento,
                        tipo=momento.tipo,
                        estado='asignado',
                        valor=nuevo_valor_visita_1
                    )
            else:
                for i in range(0, nueva_meta):
                    CuposRutaObject.objects.create(
                        ruta=self,
                        momento=momento,
                        tipo=momento.tipo,
                        estado='asignado',
                        valor=nuevo_valor
                    )

            return (nueva_meta,nuevo_valor)

    def clean_objetos(self,tipo):

        CuposRutaObject.objects.filter(ruta=self, estado = "asignado",momento__tipo = tipo).delete()
        procesados = CuposRutaObject.objects.filter(ruta = self, estado__in = ["Reportado","Pagado"], momento__tipo = tipo)

        return procesados

    def actualizar_objetos(self):

        tipos = {
            'vinculacion':{'meta':self.meta_vinculacion,'valor':float(self.valor_vinculacion.amount)},
            'visita': {'meta': self.meta_hogares, 'valor': (float(self.valor_actividades.amount)*self.peso_visitas)/100},
            'encuentro': {'meta': self.meta_hogares, 'valor': (float(self.valor_actividades.amount)*self.peso_encuentros)/100},
            'otros': {'meta': self.meta_hogares, 'valor': (float(self.valor_actividades.amount) * self.peso_otros)/100},
        }
        momentos = Momentos.objects.filter(componente=self.componente)

        #for key in tipos.keys():

        #    procesados = self.clean_objetos(key)

        #    cantidad_momentos = momentos.filter(tipo=key).count()

        #    for momento in momentos.filter(tipo=key):

        #        meta,valor = self.crear_objetos_momento(momento,procesados,tipos[key]['meta'],tipos[key]['valor'],cantidad_momentos,self.meta_vinculacion)


        if tipos['vinculacion']['meta'] > 0:

            for key in tipos.keys():

                valor_total = tipos[key]['valor'] # valor total del tipo de momento
                cantidad_momentos = momentos.filter(tipo=key).count() # cantidad total de momentos del componente
                procesados = self.clean_objetos(key) #query de objetos pagos o reportados
                valor_procesados = procesados.aggregate(Sum('valor'))['valor__sum']
                if valor_procesados == None:
                    valor_procesados = 0


                nueva_meta = (cantidad_momentos * tipos[key]['meta']) - procesados.filter(ruta = self).exclude(translado=True).count()
                nuevo_valor_total = valor_total - float(valor_procesados)




                try:
                    valor_momento = nuevo_valor_total / nueva_meta
                except:
                    pass

                else:

                    for momento in momentos.filter(tipo=key):

                        if momento.nombre == 'Visita 1':
                            nueva_meta_visita_1 = (tipos['vinculacion']['meta'] - procesados.filter(ruta = self,momento=momento).exclude(translado = True).count())

                            try:
                                nuevo_valor_visita_1 = (nuevo_valor_total/cantidad_momentos)/ nueva_meta_visita_1
                            except:
                                pass
                            else:

                                for i in range(0, nueva_meta_visita_1):
                                    CuposRutaObject.objects.create(
                                        ruta=self,
                                        momento=momento,
                                        tipo=momento.tipo,
                                        estado='asignado',
                                        valor=0
                                    )
                        else:
                            for i in range(0, tipos[key]['meta'] - procesados.filter(ruta = self,momento=momento).exclude(translado = True).count()):
                                CuposRutaObject.objects.create(
                                    ruta=self,
                                    momento=momento,
                                    tipo=momento.tipo,
                                    estado='asignado',
                                    valor=0
                                )


        else:

            for key in tipos.keys():

                valor_total = tipos[key]['valor']  # valor total del tipo de momento
                cantidad_momentos = momentos.filter(tipo=key).exclude(nombre = 'Visita 1').count()  # cantidad total de momentos del componente
                procesados = self.clean_objetos(key)  # query de objetos pagos o reportados
                valor_procesados = procesados.aggregate(Sum('valor'))['valor__sum']
                if valor_procesados == None:
                    valor_procesados = 0

                nueva_meta = (cantidad_momentos * tipos[key]['meta']) - procesados.filter(ruta=self).exclude(translado=True).exclude(momento__nombre="Visita 1").count()
                nuevo_valor_total = valor_total - float(valor_procesados)

                try:
                    valor_momento = nuevo_valor_total / nueva_meta
                except:
                    pass

                else:

                    print("Valor total: {0}".format(valor_total))

                    for momento in momentos.filter(tipo=key).exclude(nombre = 'Visita 1'):

                        for i in range(0,tipos[key]['meta'] - procesados.filter(ruta=self, momento=momento).exclude(translado=True).count()):
                            CuposRutaObject.objects.create(
                                ruta=self,
                                momento=momento,
                                tipo=momento.tipo,
                                estado='asignado',
                                valor=0
                            )


        for key in tipos.keys():

            valor_total = tipos[key]['valor']

            valor_pagados = CuposRutaObject.objects.filter(ruta = self, estado__in = ["Reportado","Pagado"], momento__tipo = key).aggregate(Sum('valor'))['valor__sum']


            if valor_pagados == None:
                valor_pagados = 0


            if valor_total > valor_pagados:

                valor_dividir = float(valor_total) - float(valor_pagados)

                objetos_creados = CuposRutaObject.objects.filter(ruta=self, estado = 'asignado', tipo = key)

                if objetos_creados.count() > 0:

                    valor_objeto = valor_dividir/objetos_creados.count()

                    objetos_creados.update(valor = valor_objeto)


        #valor_total = 0

        #for key in tipos.keys():
        #    valor_total += float(tipos[key]['valor'])


        #valor_objetos = CuposRutaObject.objects.filter(ruta = self).aggregate(Sum('valor'))['valor__sum']

        #if valor_objetos == None:
        #    valor_objetos = float(0)

        #else:
        #    valor_objetos = float(valor_objetos)


        #if valor_objetos < valor_total:

        #    adicional = valor_total - valor_objetos


        #    cantidad = CuposRutaObject.objects.filter(ruta = self, estado = 'asignado').count()


        #    if cantidad > 0:

        #        sumando = adicional / cantidad

        #        for cupo in CuposRutaObject.objects.filter(ruta = self, estado = 'asignado'):

        #            cupo.valor = float(cupo.valor.amount) + float(sumando)
        #            cupo.save()



        return 'Ok'

    def get_instrumentos_list(self,momento):

        instrumentos_return = []

        for instrumento in Instrumentos.objects.filter(momento = momento).order_by('consecutivo'):
            instrumentos_return.append({
                'id': instrumento.id,
                'short_name': instrumento.short_name,
                'icon': instrumento.icon,
                'color': instrumento.color
            })

        return instrumentos_return

    def get_valor_corte(self):
        objetos = CuposRutaObject.objects.filter(estado = "Reportado", ruta = self).exclude(momento__tipo = 'vinculacion')
        valor = objetos.aggregate(Sum('valor'))['valor__sum']
        return valor if valor != None else 0

    def get_cupo_componente(self,componente):

        if componente.consecutivo == 1:
            cupos = self.meta_hogares - Hogares.objects.filter(ruta_1=self).count()

        elif componente.consecutivo == 2:
            cupos = self.meta_hogares - Hogares.objects.filter(ruta_2=self).count()

        elif componente.consecutivo == 3:
            cupos = self.meta_hogares - Hogares.objects.filter(ruta_3=self).count()

        elif componente.consecutivo == 4:
            cupos = self.meta_hogares - Hogares.objects.filter(ruta_4=self).count()

        else:
            cupos = 0

        return cupos


    def get_cupo_vinculacion(self):

        cupos = self.meta_vinculacion - Hogares.objects.filter(ruta_vinculacion=self).count()

        return cupos


class Hogares(models.Model):

    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)

    documento = models.BigIntegerField()
    municipio = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING, related_name='hogares_municipio_inscripcion')

    primer_apellido = models.CharField(max_length=100)
    segundo_apellido = models.CharField(max_length=100,blank=True,null=True)
    primer_nombre = models.CharField(max_length=100)
    segundo_nombre = models.CharField(max_length=100,blank=True,null=True)
    fecha_nacimiento = models.DateField()
    telefono = models.CharField(max_length=100,blank=True,null=True)
    celular1 = models.CharField(max_length=100,blank=True,null=True)
    celular2 = models.CharField(max_length=100,blank=True,null=True)
    municipio_residencia = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING, related_name='hogares_municipio_residencia')

    rutas = models.ManyToManyField(Rutas, related_name='rutas',blank=True)


    def __str__(self):
        return self.get_nombres() + ' ' + self.get_apellidos() + ' - ' + str(self.documento)



    def get_rutas(self):

        rutas = ''

        for ruta in self.rutas.all():
            rutas += ruta.nombre + ', '

        if rutas != '':
            rutas = rutas[:-2]

        return rutas



    def get_valor_ruta_vinculacion(self):
        valor = CuposRutaObject.objects.filter(hogar=self,momento__tipo='vinculacion',estado__in=['Reportado','Pagado']).aggregate(Sum('valor'))['valor__sum']
        return valor if valor != None else 0



    def get_nombre_ruta_componente(self,componente):


        ruta = ''


        if componente.consecutivo == 1:
            if self.ruta_1 != None:
                ruta = self.ruta_1.nombre

        elif componente.consecutivo == 2:
            if self.ruta_2 != None:
                ruta = self.ruta_2.nombre

        elif componente.consecutivo == 3:
            if self.ruta_3 != None:
                ruta = self.ruta_3.nombre

        elif componente.consecutivo == 4:
            if self.ruta_4 != None:
                ruta = self.ruta_4.nombre

        return ruta



    def get_ruta_componente(self,componente):


        ruta = None


        if componente.consecutivo == 1:
            if self.ruta_1 != None:
                ruta = self.ruta_1

        elif componente.consecutivo == 2:
            if self.ruta_2 != None:
                ruta = self.ruta_2

        elif componente.consecutivo == 3:
            if self.ruta_3 != None:
                ruta = self.ruta_3

        elif componente.consecutivo == 4:
            if self.ruta_4 != None:
                ruta = self.ruta_4

        return ruta



    def get_nombre_ruta_vinculacion(self):

        ruta = ''

        if self.ruta_vinculacion != None:
            ruta = self.ruta_vinculacion.nombre

        return ruta



    def get_ruta_vinculacion(self):

        ruta = None

        if self.ruta_vinculacion != None:
            ruta = self.ruta_vinculacion

        return ruta



    def get_vinculacion_ruta(self, ruta):
        respuesta = 'No'

        if self.ruta_vinculacion == ruta:
            respuesta = 'Si'

        return respuesta


    def get_estado_valor(self,ruta,momento):
        estado = ''
        try:
            obj = CuposRutaObject.objects.get(ruta=ruta,momento=momento,hogar=self)
        except:
            pass
        else:
            estado = str(obj.valor).replace('COL','')
        return estado


    def get_estado(self,ruta,momento):
        estado = ''
        try:
            obj = CuposRutaObject.objects.get(ruta=ruta,momento=momento,hogar=self)
        except:
            pass
        else:
            estado = obj.estado
        return estado


    def get_cantidad_instrumentos(self):
        return InstrumentosRutaObject.objects.filter(hogar = self).count()


    def get_novedades_mis_rutas_momento(self,ruta,momento):
        return InstrumentosRutaObject.objects.filter(ruta = ruta, momento = momento, hogar = self, estado__in = ['cargado','preaprobado']).count()


    def get_nombres(self):
        if self.segundo_nombre != None:
            nombres = '{0} {1}'.format(self.primer_nombre,self.segundo_nombre)
        else:
            nombres = self.primer_nombre
        return nombres

    def get_apellidos(self):
        if self.segundo_apellido != None:
            apellidos = '{0} {1}'.format(self.primer_apellido,self.segundo_apellido)
        else:
            apellidos = self.primer_apellido
        return apellidos

    def get_full_name(self):
        return '{0} {1}'.format(self.get_nombres(),self.get_apellidos())


    def get_gull_name(self):
        return '{0} {1}'.format(self.get_nombres(),self.get_apellidos())


    def get_cantidad_miembros(self):
        return MiembroNucleoHogar.objects.filter(hogar=self).count()


    def get_valor_total(self):
        valor = CuposRutaObject.objects.filter(hogar=self).aggregate(Sum('valor'))['valor__sum']
        return valor if valor != None else 0

    def get_cantidad_componentes(self):
        return CuposRutaObject.objects.filter(hogar=self).values_list('momento__componente__id',flat=True).distinct().count()

    def get_valor_vinculacion(self):
        valor = CuposRutaObject.objects.filter(hogar=self,momento__tipo='vinculacion').aggregate(Sum('valor'))['valor__sum']
        return valor if valor != None else 0





class MiembroNucleoHogar(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogar = models.ForeignKey(Hogares,on_delete=models.DO_NOTHING,related_name='hogar_miembro_nucleo')


    #Datos personales

    tipo_documento = models.CharField(max_length=100,blank=True,null=True)
    numero_documento = models.IntegerField(blank=True,null=True)
    primer_apellido = models.CharField(max_length=100,blank=True,null=True)
    segundo_apellido = models.CharField(max_length=100,blank=True, null=True)
    primer_nombre = models.CharField(max_length=100,blank=True,null=True)
    segundo_nombre = models.CharField(max_length=100,blank=True, null=True)
    celular_1 = models.CharField(max_length=100,blank=True,null=True)
    celular_2 = models.CharField(max_length=100,blank=True, null=True)
    correo_electronico = models.EmailField(max_length=100,blank=True, null=True)

    # Lugar y fecha de nacimiento

    departamento_nacimiento = models.ForeignKey(Departamentos, on_delete=models.DO_NOTHING,related_name='departamento_nacimiento_miembro_nucleo',blank=True,null=True)
    municipio_nacimiento = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_nacimiento_miembro_nucleo',blank=True,null=True)
    fecha_nacimiento = models.DateField(blank=True,null=True)

    # Lugar y fecha de expedición del documento

    departamento_expedicion = models.ForeignKey(Departamentos, on_delete=models.DO_NOTHING,related_name='departamento_expedicion_miembro_nucleo',blank=True,null=True)
    municipio_expedicion = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_expedicion_miembro_nucleo',blank=True,null=True)
    fecha_expedicion = models.DateField(blank=True,null=True)

    longitud = models.DecimalField(max_digits=15, decimal_places=10, blank=True, null=True)
    latitud = models.DecimalField(max_digits=15, decimal_places=10, blank=True, null=True)
    precision = models.DecimalField(max_digits=15, decimal_places=10, blank=True, null=True)
    altitud = models.DecimalField(max_digits=15, decimal_places=10, blank=True, null=True)

    # -------------------------------------------------------------

    # Caracteristicas generales

    sexo = models.CharField(max_length=100,blank=True,null=True)
    tiene_libreta = models.BooleanField(default=False,blank=True,null=True)
    numero_libreta = models.CharField(max_length=100, blank=True, null=True)

    identidad_genero = models.CharField(max_length=100, blank=True, null=True)
    condicion_sexual = models.CharField(max_length=100, blank=True, null=True)
    estado_civil = models.CharField(max_length=100,blank=True,null=True)
    etnia = models.CharField(max_length=100,blank=True,null=True)

    pueblo_indigena = models.ForeignKey(PueblosIndigenas, on_delete=models.DO_NOTHING, blank=True, null=True)#aparece si se selecciona "INDIGENA" en la etnia
    resguardo_indigena = models.ForeignKey(ResguardosIndigenas, on_delete=models.DO_NOTHING, blank=True, null=True)  # aparece si se selecciona "INDIGENA" en la etnia
    comunidad_indigena = models.ForeignKey(ComunidadesIndigenas, on_delete=models.DO_NOTHING, blank=True, null=True)  # aparece si se selecciona "INDIGENA" en la etnia
    lengua_nativa_indigena = models.BooleanField(blank=True,null=True)  # aparece si se selecciona "INDIGENA" en la etnia
    cual_lengua_indigena = models.ForeignKey(LenguasNativas, related_name='lengua_indigena_miembro_nucleo', on_delete=models.DO_NOTHING,blank=True,null=True)# aparece si se selecciona "INDIGENA" en la etnia y si se activa lengua nativa

    consejo_afro = models.ForeignKey(ConsejosAfro, on_delete=models.DO_NOTHING,blank=True,null=True)#aparece si se selecciona "AFROCOLOMBIANO" en la etnia
    comunidad_afro = models.ForeignKey(ComunidadesAfro, on_delete=models.DO_NOTHING, blank=True,null=True)  # aparece si se selecciona "AFROCOLOMBIANO" en la etnia
    lengua_nativa_afro = models.BooleanField(blank=True,null=True)  # aparece si se selecciona "AFROCOLOMBIANO" en la etnia
    cual_lengua_afro = models.ForeignKey(LenguasNativas, related_name='lengua_afro_miembro_nucleo', on_delete=models.DO_NOTHING, blank=True,null=True)  # aparece si se selecciona "AFROCOLOMBIANO" en la etnia y si se activa lengua nativa

    discapacidad = models.BooleanField(blank=True,null=True)

    registro_discapacidad = models.CharField(max_length=100,blank=True,null=True) #aparece si hay discapacidad
    categoria_discapacidad = models.ManyToManyField(CategoriaDiscapacidad, blank=True) #aparece si hay discapacidad
    dificultades_permanentes = models.ManyToManyField(DificultadesPermanentesDiscapacidad, blank=True) #aparece si hay discapacidad
    utiliza_actualmente = models.ManyToManyField(ElementosDiscapacidad, blank=True) #aparece si hay discapacidad
    rehabilitacion = models.ManyToManyField(TiposRehabilitacionDiscapacidad, blank=True) #aparece si hay discapacidad
    tiene_cuidador = models.BooleanField(blank=True,null=True) #aparece si hay discapacidad
    cuidador = models.CharField(max_length=100,blank=True,null=True)

    parentezco = models.CharField(max_length=100,blank=True,null=True)
    es_jefe = models.BooleanField(blank=True,null=True)

    nivel_escolaridad = models.CharField(max_length=100,blank=True,null=True)
    grado_titulo = models.CharField(max_length=100,blank=True,null=True)
    sabe_leer = models.BooleanField(blank=True,null=True)
    sabe_sumar_restar = models.BooleanField(blank=True,null=True)
    actualmente_estudia = models.BooleanField(blank=True,null=True)
    recibe_alimentos = models.BooleanField(blank=True,null=True)

    razon_no_estudia = models.CharField(max_length=100, blank=True, null=True)  # se activa si no estudia
    razon_no_estudia_otra = models.CharField(max_length=100, blank=True, null=True)  # se activa si no estudia y hay otra razon
    regimen_seguridad_social = models.CharField(max_length=100,blank=True,null=True)


    def get_str_list_categoria_discapacidad(self):
        string = ', '

        for categoria in self.categoria_discapacidad.all():
            string += categoria.nombre + ', '

        return string[:-2]

    def get_str_list_dificultades_permanentes(self):
        string = ', '

        for categoria in self.dificultades_permanentes.all():
            string += categoria.nombre + ', '

        return string[:-2]

    def get_str_list_utiliza_actualmente(self):
        string = ', '

        for categoria in self.utiliza_actualmente.all():
            string += categoria.nombre + ', '

        return string[:-2]

    def get_str_list_rehabilitacion(self):
        string = ', '

        for categoria in self.rehabilitacion.all():
            string += categoria.nombre + ', '

        return string[:-2]


    def get_fullname(self):
        return '{0} {1} {2} {3}'.format(self.primer_nombre,self.segundo_nombre,self.primer_apellido,self.segundo_apellido)



class PermisosCuentasRutas(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    user = models.OneToOneField(User,on_delete=models.DO_NOTHING)
    rutas_ver = models.ManyToManyField(Rutas,related_name="permisos_cuentas_ver",blank=True)
    #rutas_preaprobar = models.ManyToManyField(Rutas,related_name="permisos_cuentas_preaprobar",blank=True)
    rutas_aprobar = models.ManyToManyField(Rutas,related_name="permisos_cuentas_aprobar",blank=True)

    def __str__(self):
        return self.user.email



class PermisosCuentasDepartamentos(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    users = models.ManyToManyField(User, blank=True)
    departamento = models.OneToOneField(Departamentos, on_delete=models.DO_NOTHING)

    def __str__(self):
        return self.departamento.nombre





class Cortes(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    consecutivo = models.IntegerField()
    creation = models.DateTimeField(auto_now_add=True)
    usuario_creacion = models.ForeignKey(User, on_delete=models.DO_NOTHING,related_name='cortes_usuario_creacion_fest')
    valor = MoneyField(max_digits=10, decimal_places=2, default_currency='COP', default=0, blank=True, null=True)
    descripcion = models.CharField(max_length=200)

    def __str__(self):
        return self.descripcion


    def pretty_creation_datetime(self):
        return self.creation.astimezone(settings_time_zone).strftime('%d/%m/%Y - %I:%M:%S %p')


    def get_valor(self):
        valor = CuposRutaObject.objects.filter(corte = self).aggregate(Sum('valor'))['valor__sum']
        return valor if valor != None else 0


    def get_novedades(self):
        cuentas_cobro = CuentasCobro.objects.filter(corte = self, estado__in = ['Creado', 'Cargado'])
        return cuentas_cobro.count()

    def get_cantidad_cuentas_cobro(self):
        return CuentasCobro.objects.filter(corte = self).count()

    def create_cuentas_cobro(self, user):
        objetos = CuposRutaObject.objects.filter(corte = self, estado = "Pagado")
        rutas_ids = objetos.values_list('ruta__id', flat=True).distinct()
        for ruta_id in rutas_ids:
            ruta = Rutas.objects.get(id = ruta_id)

            try:
                cuenta_cobro = CuentasCobro.objects.get(
                    ruta = ruta,
                    corte = self
                )
            except:

                valor = objetos.filter(ruta = ruta).aggregate(Sum('valor'))['valor__sum']

                if valor == None:
                    valor = 0

                cuenta_cobro = CuentasCobro.objects.create(
                    ruta = ruta,
                    corte = self,
                    usuario_creacion = user,
                    estado = 'Creado',
                    valor = valor
                )

            else:
                pass


        return None


class CuposRutaObject(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)

    ruta = models.ForeignKey(Rutas,on_delete=models.DO_NOTHING,related_name='cupo_ruta')
    momento = models.ForeignKey(Momentos,on_delete=models.DO_NOTHING,related_name='cupo_momento')
    tipo = models.CharField(max_length=100)
    estado = models.CharField(max_length=100)
    valor = MoneyField(max_digits=10, decimal_places=2, default_currency='COP', default=0)
    hogares = models.ManyToManyField(Hogares,related_name='cupo_hogares',blank=True)
    corte = models.ForeignKey(Cortes, on_delete=models.DO_NOTHING, blank=True, null=True)
    translado = models.BooleanField(default=False)




class InstrumentosRutaObject(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    creacion = models.DateTimeField(auto_now_add=True)
    usuario_creacion = models.ForeignKey(User, on_delete=models.DO_NOTHING,related_name='instrumento_usuario_creacion',blank=True,null=True)

    ruta = models.ForeignKey(Rutas, on_delete=models.DO_NOTHING, related_name='instrumento_ruta')
    momento = models.ForeignKey(Momentos, on_delete=models.DO_NOTHING, related_name='instrumento_momento')
    hogares = models.ManyToManyField(Hogares, related_name='instrumento_hogar', blank=True)
    instrumento = models.ForeignKey(Instrumentos, on_delete=models.DO_NOTHING, related_name='instrumento_instrumento',blank=True,null=True)

    modelo = models.CharField(max_length=100)
    soporte = models.UUIDField(blank=True,null=True)
    observacion = models.TextField(blank=True,null=True)
    fecha_actualizacion = models.DateTimeField(blank=True,null=True)
    usuario_actualizacion = models.ForeignKey(User, on_delete=models.DO_NOTHING,related_name='instrumento_usuario_actualizacion',blank=True,null=True)
    estado = models.CharField(max_length=100,blank=True,null=True)
    nombre = models.CharField(max_length=100,blank=True,null=True)
    consecutivo = models.IntegerField(blank=True,null=True)
    cupo_object = models.ForeignKey(CuposRutaObject,on_delete=models.DO_NOTHING,blank=True,null=True)



    def clean_similares(self):

        from fest_2019 import modelos_instrumentos

        self.modelos = modelos_instrumentos.get_modelo(self.instrumento.modelo)

        if self.instrumento.nivel != 'ruta':

            for instrumento_object in InstrumentosRutaObject.objects.filter(ruta = self.ruta, momento = self.momento, instrumento = self.instrumento).exclude(id = self.id):

                for hogar in self.hogares.all():

                    if hogar in instrumento_object.hogares.all():
                        instrumento_object.hogares.remove(hogar)

                if instrumento_object.hogares.all().count() == 0:
                    self.modelos.get('model').objects.get(id = instrumento_object.soporte).delete()
                    InstrumentosTrazabilidadRutaObject.objects.filter(instrumento = instrumento_object).delete()
                    ObservacionesInstrumentoRutaObject.objects.filter(instrumento = instrumento_object).delete()
                    instrumento_object.delete()
                    self.ruta.update_novedades()



    def get_hogares_list(self):
        hogares = ''

        for hogar in self.hogares.all():
            hogares += '<p>{0} - {1} {2}</p>'.format(hogar.documento,hogar.get_nombres(),hogar.get_apellidos())

        return hogares

    def get_hogares_reporte(self):
        hogares = ''

        for hogar in self.hogares.all():
            hogares += '{0} - {1} {2}\n'.format(hogar.documento,hogar.get_nombres(),hogar.get_apellidos())

        return hogares

    def get_departamentos_reporte(self):
        departamentos = ''

        for hogar in self.hogares.all():
            departamentos += '{0}, {1}\n'.format(hogar.municipio_residencia.nombre,hogar.municipio_residencia.departamento.nombre)

        return departamentos


class ObservacionesInstrumentoRutaObject(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    instrumento = models.ForeignKey(InstrumentosRutaObject, on_delete=models.DO_NOTHING)
    creacion = models.DateTimeField(auto_now_add=True)
    usuario_creacion = models.ForeignKey(User, on_delete=models.DO_NOTHING, related_name='instrumento_observacion_usuario_creacion',blank=True, null=True)
    observacion = models.TextField(blank=True,null=True)


    def pretty_creation_datetime(self):
        return self.creacion.astimezone(settings_time_zone).strftime('%d/%m/%Y - %I:%M:%S %p')



class InstrumentosTrazabilidadRutaObject(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    instrumento = models.ForeignKey(InstrumentosRutaObject,on_delete=models.DO_NOTHING,related_name="trazabilidad_instrumento")
    creacion = models.DateTimeField(auto_now_add=True)
    user = models.ForeignKey(User, on_delete=models.DO_NOTHING, related_name='trazabilidad_instrumento_usuario')
    observacion = models.TextField()









def upload_dinamic_cuentas_cobro(instance, filename):
    return '/'.join(['FEST 2019', 'Cuentas de Cobro', str(instance.id), filename])

class CuentasCobro(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    ruta = models.ForeignKey(Rutas, on_delete=models.DO_NOTHING)
    creation = models.DateTimeField(auto_now_add=True)
    usuario_creacion = models.ForeignKey(User, on_delete=models.DO_NOTHING, blank=True, null=True,related_name='cuentas_cobro_usuario_creacion_fest')

    fecha_actualizacion = models.DateTimeField(blank=True,null=True)
    usuario_actualizacion = models.ForeignKey(User, on_delete=models.DO_NOTHING, blank=True, null=True, related_name='usuario_actualizacion_cuentas_cobro_fest')

    corte = models.ForeignKey(Cortes, on_delete = models.DO_NOTHING, blank=True, null=True)
    estado = models.CharField(max_length=100, blank=True, null=True)
    valor = MoneyField(max_digits=10, decimal_places=2, default_currency='COP', default=0, blank=True, null=True)
    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_cuentas_cobro,
        content_types=['application/pdf'],
        max_upload_size=5242880,
        max_length=255,
        blank=True,
        null=True
    )
    file2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_cuentas_cobro,
        content_types=['application/pdf'],
        max_upload_size=5242880,
        max_length=255,
        blank=True,
        null=True
    )
    html = models.FileField(upload_to=upload_dinamic_cuentas_cobro, blank=True, null=True)
    delta = models.TextField(blank=True, null=True)
    data_json = models.TextField(blank=True,null=True)
    valores_json = models.TextField(default='[]',blank=True,null=True)
    observaciones = models.TextField(default='',blank=True,null=True)



    def pretty_creation_datetime(self):
        return self.creation.astimezone(settings_time_zone).strftime('%d/%m/%Y - %I:%M:%S %p')

    def create_delta(self):
        from fest_2019.functions import delta_cuenta_cobro
        self.delta = json.dumps(delta_cuenta_cobro(self))
        self.save()
        return None

    def get_html_delta(self):
        delta = json.loads(self.delta)
        return html.render(delta['ops'])

    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url

    def url_file2(self):
        url = None
        try:
            url = self.file2.url
        except:
            pass
        return url

    def pretty_print_url_file2(self):
        try:
            url = self.file2.url
        except:
            return '<p style="display:inline;margin-left:5px;">No hay archivos cargados.</p>'
        else:
            return '<a href="'+ url +'"> '+ str(self.file2.name) +'</a>'








def upload_dinamic_fest(instance, filename):
    return '/'.join(['FEST 2019', str(instance.ruta.id), instance.nombre, filename])

class Documento(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogares = models.ManyToManyField(Hogares,related_name='hogares_documento',blank=True)
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_documento',blank=True,null=True)
    ruta = models.ForeignKey(Rutas,on_delete=models.DO_NOTHING,related_name='ruta_documento',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )



    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url


    def get_extension(self):
        return self.file.name.split('.')[-1]








def upload_dinamic_formulario_caracterizacion(instance, filename):
    return '/'.join(['FEST 2019', str(instance.ruta.id), instance.nombre, filename])

class FormularioCaracterizacion(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogares = models.ManyToManyField(Hogares,related_name='hogares_formulario_caracterizacion',blank=True)
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_formulario_caracterizacion',blank=True,null=True)
    ruta = models.ForeignKey(Rutas,on_delete=models.DO_NOTHING,related_name='ruta_formulario_caracterizacion',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_formulario_caracterizacion,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    file2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_formulario_caracterizacion,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )



    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url


    def url_file2(self):
        url = None
        try:
            url = self.file2.url
        except:
            pass
        return url


    def get_extension(self):
        return self.file.name.split('.')[-1]







def upload_dinamic_ficha_icoe(instance, filename):
    return '/'.join(['FEST 2019', str(instance.ruta.id), instance.nombre, filename])

class FichaIcoe(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogares = models.ManyToManyField(Hogares,related_name='hogares_ficha_icoe',blank=True)
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_ficha_icoe',blank=True,null=True)
    ruta = models.ForeignKey(Rutas,on_delete=models.DO_NOTHING,related_name='ruta_ficha_icoe',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    municipio = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_ficha_icoe')
    nombre_comunidad = models.CharField(max_length=100)
    resguado_indigena_consejo_comunitario = models.CharField(max_length=100)

    fecha_entrada = models.DateField()
    fecha_salida = models.DateField()


    aspecto_1_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_1_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_1_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    aspecto_2_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_2_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_2_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    aspecto_3_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_3_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_3_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    aspecto_4_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_4_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_4_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    subindice_1_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    subindice_1_salida = models.DecimalField(max_digits=10, decimal_places=2)
    subindice_1_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    aspecto_5_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_5_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_5_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    aspecto_6_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_6_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_6_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    aspecto_7_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_7_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_7_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    aspecto_8_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_8_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_8_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    subindice_2_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    subindice_2_salida = models.DecimalField(max_digits=10, decimal_places=2)
    subindice_2_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    aspecto_9_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_9_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_9_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    aspecto_10_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_10_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_10_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    aspecto_11_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_11_salida = models.DecimalField(max_digits=10, decimal_places=2)
    aspecto_11_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    subindice_3_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    subindice_3_salida = models.DecimalField(max_digits=10, decimal_places=2)
    subindice_3_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    total_entrada = models.DecimalField(max_digits=10, decimal_places=2)
    total_salida = models.DecimalField(max_digits=10, decimal_places=2)
    total_variacion = models.DecimalField(max_digits=10, decimal_places=2)

    #Ficha ICOE F-GI-IP 101
    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_icoe,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    #Listado de Asistencia
    file2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_icoe,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    #FICHA ICOE en excel
    file3 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_icoe,
        content_types=[
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    #registro fotografico
    foto1 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_icoe,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png',
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    foto2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_icoe,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png',
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    foto3 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_icoe,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png',
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )





    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url


    def url_file2(self):
        url = None
        try:
            url = self.file2.url
        except:
            pass
        return url

    def url_file3(self):
        url = None
        try:
            url = self.file3.url
        except:
            pass
        return url

    def url_foto1(self):
        url = None
        try:
            url = self.foto1.url
        except:
            pass
        return url

    def url_foto2(self):
        url = None
        try:
            url = self.foto2.url
        except:
            pass
        return url

    def url_foto3(self):
        url = None
        try:
            url = self.foto3.url
        except:
            pass
        return url

    def get_extension_file(self):
        return self.file.name.split('.')[-1]

    def get_extension_file2(self):
        return self.file2.name.split('.')[-1]

    def get_extension_file3(self):
        return self.file3.name.split('.')[-1]








def upload_dinamic_acta_socializacion_comunidades(instance, filename):
    return '/'.join(['FEST 2019', str(instance.ruta.id), instance.nombre, filename])

class ActaSocializacionComunidades(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogares = models.ManyToManyField(Hogares,related_name='hogares_acta_socializacion_comunidades',blank=True)
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_acta_socializacion_comunidades',blank=True,null=True)
    ruta = models.ForeignKey(Rutas,on_delete=models.DO_NOTHING,related_name='ruta_acta_socializacion_comunidades',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    nombre_comunidad = models.CharField(max_length=100)
    resguado_indigena_consejo_comunitario = models.CharField(max_length=100)
    municipio = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_acta_socializacion_comunidades')
    nombre_representante = models.CharField(max_length=200)
    documento_representante = models.IntegerField()
    cargo_representante = models.CharField(max_length=200)
    fecha_firma = models.DateField()


    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_socializacion_comunidades,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    file2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_socializacion_comunidades,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    file3 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_socializacion_comunidades,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    foto1 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_socializacion_comunidades,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )
    foto2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_socializacion_comunidades,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )



    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url

    def url_file2(self):
        url = None
        try:
            url = self.file2.url
        except:
            pass
        return url

    def url_file3(self):
        url = None
        try:
            url = self.file3.url
        except:
            pass
        return url


    def url_foto1(self):
        url = None
        try:
            url = self.foto1.url
        except:
            pass
        return url

    def url_foto2(self):
        url = None
        try:
            url = self.foto2.url
        except:
            pass
        return url


    def get_extension(self):
        return self.file.name.split('.')[-1]








def upload_dinamic_ficha_vision_desarrollo(instance, filename):
    return '/'.join(['FEST 2019', str(instance.ruta.id), instance.nombre, filename])

class FichaVisionDesarrollo(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogares = models.ManyToManyField(Hogares,related_name='hogares_ficha_vision_desarrollo',blank=True)
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_ficha_vision_desarrollo',blank=True,null=True)
    ruta = models.ForeignKey(Rutas,on_delete=models.DO_NOTHING,related_name='ruta_ficha_vision_desarrollo',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    municipio = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_ficha_vision_desarrollo')
    fecha = models.DateField()
    lugar = models.CharField(max_length=100)
    dependencia = models.CharField(max_length=100)

    asistentes = models.IntegerField()

    #ACTA DE REUNION
    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_vision_desarrollo,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    #LISTADO DE ASISTENCIA
    file2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_vision_desarrollo,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    #REGISTRO FOTOGRAFICO

    foto1 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_vision_desarrollo,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )
    foto2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_vision_desarrollo,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    #Fotografia Visión en texto
    foto3 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_vision_desarrollo,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    #Fotografia Cartografia de la Visión
    foto4 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_vision_desarrollo,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    # Formulario Gforms
    foto5 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_vision_desarrollo,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )



    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url

    def url_file2(self):
        url = None
        try:
            url = self.file2.url
        except:
            pass
        return url


    def url_foto1(self):
        url = None
        try:
            url = self.foto1.url
        except:
            pass
        return url

    def url_foto2(self):
        url = None
        try:
            url = self.foto2.url
        except:
            pass
        return url

    def url_foto3(self):
        url = None
        try:
            url = self.foto3.url
        except:
            pass
        return url

    def url_foto4(self):
        url = None
        try:
            url = self.foto4.url
        except:
            pass
        return url

    def url_foto5(self):
        url = None
        try:
            url = self.foto5.url
        except:
            pass
        return url


    def get_extension_file(self):
        return self.file.name.split('.')[-1]

    def get_extension_file2(self):
        return self.file2.name.split('.')[-1]

    def get_extension_foto1(self):
        return self.foto1.name.split('.')[-1]

    def get_extension_foto2(self):
        return self.foto2.name.split('.')[-1]

    def get_extension_foto3(self):
        return self.foto3.name.split('.')[-1]

    def get_extension_foto4(self):
        return self.foto4.name.split('.')[-1]

    def get_extension_foto5(self):
        return self.foto5.name.split('.')[-1]







def upload_dinamic_diagnostico_comunitario(instance, filename):
    return '/'.join(['FEST 2019', str(instance.ruta.id), instance.nombre, filename])

class DiagnosticoComunitario(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogares = models.ManyToManyField(Hogares,related_name='hogares_diagnostico_comunitario',blank=True)
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_diagnostico_comunitario',blank=True,null=True)
    ruta = models.ForeignKey(Rutas,on_delete=models.DO_NOTHING,related_name='ruta_diagnostico_comunitario',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    municipio = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_diagnostico_comunitario')
    fecha = models.DateField()
    lugar = models.CharField(max_length=100)
    dependencia = models.CharField(max_length=100)

    asistentes = models.IntegerField()


    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_diagnostico_comunitario,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    file2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_diagnostico_comunitario,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    file3 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_diagnostico_comunitario,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    file4 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_diagnostico_comunitario,
        content_types=[
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    file5 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_diagnostico_comunitario,
        content_types=[
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    foto1 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_diagnostico_comunitario,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )
    foto2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_diagnostico_comunitario,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    foto3 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_diagnostico_comunitario,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png',
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    foto4 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_diagnostico_comunitario,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png',
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )

    foto5 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_diagnostico_comunitario,
        content_types=[
            'application/pdf',
            'image/jpg',
            'image/jpeg',
            'image/png',
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )



    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url

    def url_file2(self):
        url = None
        try:
            url = self.file2.url
        except:
            pass
        return url

    def url_file3(self):
        url = None
        try:
            url = self.file3.url
        except:
            pass
        return url

    def url_file4(self):
        url = None
        try:
            url = self.file4.url
        except:
            pass
        return url

    def url_file5(self):
        url = None
        try:
            url = self.file5.url
        except:
            pass
        return url


    def url_foto1(self):
        url = None
        try:
            url = self.foto1.url
        except:
            pass
        return url

    def url_foto2(self):
        url = None
        try:
            url = self.foto2.url
        except:
            pass
        return url

    def url_foto3(self):
        url = None
        try:
            url = self.foto3.url
        except:
            pass
        return url

    def url_foto4(self):
        url = None
        try:
            url = self.foto4.url
        except:
            pass
        return url

    def url_foto5(self):
        url = None
        try:
            url = self.foto5.url
        except:
            pass
        return url


    def get_extension_foto1(self):
        return self.foto1.name.split('.')[-1]

    def get_extension_foto2(self):
        return self.foto2.name.split('.')[-1]

    def get_extension_foto3(self):
        return self.foto3.name.split('.')[-1]

    def get_extension_foto4(self):
        return self.foto4.name.split('.')[-1]

    def get_extension_foto5(self):
        return self.foto5.name.split('.')[-1]

    def get_extension_file2(self):
        return self.file2.name.split('.')[-1]








def upload_dinamic_acta_socializacion_concertacion(instance, filename):
    return '/'.join(['FEST 2019', str(instance.ruta.id), instance.nombre, filename])

class ActaSocializacionConcertacion(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogares = models.ManyToManyField(Hogares,related_name='hogares_acta_socializacion_concertacion',blank=True)
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_acta_socializacion_concertacion',blank=True,null=True)
    ruta = models.ForeignKey(Rutas,on_delete=models.DO_NOTHING,related_name='ruta_acta_socializacion_concertacion',blank=True,null=True)
    nombre = models.CharField(max_length=100)


    fecha_diligenciamiento = models.DateField()
    lugar = models.CharField(max_length=100)
    hora = models.CharField(max_length=100)
    municipio = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_acta_socializacion_concertacion')
    resguado_indigena_consejo_comunitario = models.CharField(max_length=100)
    nombre_comunidad = models.CharField(max_length=100)


    nombre_representante = models.CharField(max_length=200)
    datos_contacto_representante = models.CharField(max_length=200)



    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_socializacion_concertacion,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    file2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_socializacion_concertacion,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    foto1 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_socializacion_concertacion,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )
    foto2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_socializacion_concertacion,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
        blank=True,
        null=True
    )



    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url


    def url_file2(self):
        url = None
        try:
            url = self.file2.url
        except:
            pass
        return url


    def url_foto1(self):
        url = None
        try:
            url = self.foto1.url
        except:
            pass
        return url

    def url_foto2(self):
        url = None
        try:
            url = self.foto2.url
        except:
            pass
        return url


    def get_extension(self):
        return self.file.name.split('.')[-1]







def upload_dinamic_acta_vinculacion_hogar(instance, filename):
    return '/'.join(['FEST 2019', str(instance.ruta.id), instance.nombre, filename])

class ActaVinculacionHogar(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogares = models.ManyToManyField(Hogares,related_name='hogares_acta_vinculacion_hogar',blank=True)
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_acta_vinculacion_hogar',blank=True,null=True)
    ruta = models.ForeignKey(Rutas,on_delete=models.DO_NOTHING,related_name='ruta_acta_vinculacion_hogar',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    fecha_diligenciamiento = models.DateField()
    municipio = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_acta_vinculacion_hogar')
    resguado_indigena_consejo_comunitario = models.CharField(max_length=100)
    nombre_comunidad = models.CharField(max_length=100)

    tipo_identificacion = models.CharField(max_length=100)
    documento_representante = models.IntegerField()
    nombre_representante = models.CharField(max_length=200)
    telefono_celular = models.CharField(max_length=300)


    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_vinculacion_hogar,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )

    file2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_acta_vinculacion_hogar,
        content_types=[
            'application/pdf',
        ],
        max_upload_size=10485760,
        max_length=255
    )



    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url


    def url_file2(self):
        url = None
        try:
            url = self.file2.url
        except:
            pass
        return url


    def get_extension(self):
        return self.file.name.split('.')[-1]








class DocumentoExcel(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogar = models.ForeignKey(Hogares,on_delete=models.DO_NOTHING,related_name='hogar_documento_excel')
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_documento_excel',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ],
        max_upload_size=10485760,
        max_length=255
    )



    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url


    def get_extension(self):
        return self.file.name.split('.')[-1]


class Fotos4(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogar = models.ForeignKey(Hogares,on_delete=models.DO_NOTHING,related_name='hogar_fotos4')
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_fotos4',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    foto1 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
    )
    foto2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )
    foto3 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )
    foto4 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )



    def url_foto1(self):
        url = None
        try:
            url = self.foto1.url
        except:
            pass
        return url

    def url_foto2(self):
        url = None
        try:
            url = self.foto2.url
        except:
            pass
        return url

    def url_foto3(self):
        url = None
        try:
            url = self.foto3.url
        except:
            pass
        return url

    def url_foto4(self):
        url = None
        try:
            url = self.foto4.url
        except:
            pass
        return url

class Fotos5(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogar = models.ForeignKey(Hogares,on_delete=models.DO_NOTHING,related_name='hogar_fotos5')
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_fotos5',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    foto1 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
    )
    foto2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )
    foto3 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )
    foto4 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )
    foto5 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )


    def url_foto1(self):
        url = None
        try:
            url = self.foto1.url
        except:
            pass
        return url

    def url_foto2(self):
        url = None
        try:
            url = self.foto2.url
        except:
            pass
        return url

    def url_foto3(self):
        url = None
        try:
            url = self.foto3.url
        except:
            pass
        return url

    def url_foto4(self):
        url = None
        try:
            url = self.foto4.url
        except:
            pass
        return url

    def url_foto5(self):
        url = None
        try:
            url = self.foto5.url
        except:
            pass
        return url


class Fotos6(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogar = models.ForeignKey(Hogares,on_delete=models.DO_NOTHING,related_name='hogar_fotos6')
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_fotos6',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    foto1 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
    )
    foto2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )
    foto3 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )
    foto4 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )
    foto5 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255
    )
    foto6 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
    )


    def url_foto1(self):
        url = None
        try:
            url = self.foto1.url
        except:
            pass
        return url

    def url_foto2(self):
        url = None
        try:
            url = self.foto2.url
        except:
            pass
        return url

    def url_foto3(self):
        url = None
        try:
            url = self.foto3.url
        except:
            pass
        return url

    def url_foto4(self):
        url = None
        try:
            url = self.foto4.url
        except:
            pass
        return url

    def url_foto5(self):
        url = None
        try:
            url = self.foto5.url
        except:
            pass
        return url

    def url_foto6(self):
        url = None
        try:
            url = self.foto6.url
        except:
            pass
        return url


class Fotos1(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogar = models.ForeignKey(Hogares,on_delete=models.DO_NOTHING,related_name='hogar_fotos1')
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_fotos1',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    foto1 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
    )


    def url_foto1(self):
        url = None
        try:
            url = self.foto1.url
        except:
            pass
        return url




class ArchivoRarZip(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogar = models.ForeignKey(Hogares,on_delete=models.DO_NOTHING,related_name='hogar_rar_zip')
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_rar_zip',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'application/x-rar-compressed',
            'application/zip',
            'application/x-7z-compressed'
        ],
        max_upload_size=10485760,
        max_length=255,
    )


    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url



class Fotos2(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogar = models.ForeignKey(Hogares,on_delete=models.DO_NOTHING,related_name='hogar_fotos2')
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_fotos2',blank=True,null=True)
    nombre = models.CharField(max_length=100)

    foto1 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
    )

    foto2 = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_fest,
        content_types=[
            'image/jpg',
            'image/jpeg',
            'image/png'
        ],
        max_upload_size=10485760,
        max_length=255,
    )


    def url_foto1(self):
        url = None
        try:
            url = self.foto1.url
        except:
            pass
        return url

    def url_foto2(self):
        url = None
        try:
            url = self.foto2.url
        except:
            pass
        return url

class CaracterizacionInicial(models.Model):
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    hogar = models.ForeignKey(Hogares,on_delete=models.DO_NOTHING,related_name='hogar_caracterizacion_inicial')
    instrumento = models.ForeignKey(Instrumentos,on_delete=models.DO_NOTHING,related_name='instrumento_caracterizacion_inicial',blank=True,null=True)
    nombre = models.CharField(max_length=100,blank=True,null=True)

    # -------------------------------------------------------------

    #lugar de atencion
    departamento_atencion = models.ForeignKey(Departamentos,on_delete=models.DO_NOTHING,related_name='departamento_atencion_caracterizacion_inicial',blank=True,null=True)
    municipio_atencion = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_atencion_caracterizacion_inicial',blank=True,null=True)

    #lugar de residencia
    departamento_residencia = models.ForeignKey(Departamentos, on_delete=models.DO_NOTHING,related_name='departamento_residencia_caracterizacion_inicial',blank=True,null=True)
    municipio_residencia = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_residencia_caracterizacion_inicial',blank=True,null=True)
    zona_residencia = models.CharField(max_length=100,blank=True,null=True)

    localidad = models.CharField(max_length=1000,blank=True,null=True) #aparece si se selecciona "cabecera municipal" en la zona de residencia
    barrio = models.CharField(max_length=1000,blank=True,null=True) #aparece si se selecciona "cabecera municipal" en la zona de residencia
    direccion_predio = models.CharField(max_length=1000,blank=True,null=True) #aparece si se selecciona "cabecera municipal" en la zona de residencia

    corregimiento = models.ForeignKey(Corregimientos,on_delete=models.DO_NOTHING,blank=True,null=True) #aparece si se selecciona las opciones "centro poblado" o "rural disperso"
    vereda = models.ForeignKey(Veredas,on_delete=models.DO_NOTHING, blank=True,null=True)  # aparece si se selecciona las opciones "centro poblado" o "rural disperso"
    ubicacion_predio = models.CharField(max_length=1000, blank=True,null=True)  # aparece si se selecciona las opciones "centro poblado" o "rural disperso"

    telefono_fijo = models.CharField(max_length=100,blank=True,null=True)


    tipo_vivienda = models.CharField(max_length=100,blank=True,null=True)
    otro_tipo_vivienda = models.CharField(max_length=100,blank=True,null=True)#aparece si se selecciona "Otro tipo de vivienda"

    propiedad_vivienda = models.CharField(max_length=100,blank=True,null=True)
    estrato_vivienda = models.CharField(max_length=100,blank=True,null=True)

    #-------------------------------------------------------------

    #información sobre la familia

    otro_telefono = models.CharField(max_length=100, blank=True, null=True)
    descripcion_direccion = models.CharField(max_length=100, blank=True, null=True)
    numero_personas_familia = models.IntegerField(blank=True,null=True)
    menores_5_anios = models.IntegerField(blank=True,null=True)
    mayores_60_anios = models.IntegerField(blank=True,null=True)
    mujeres_gestantes_lactantes = models.IntegerField(blank=True,null=True)
    discapacitados_familia = models.IntegerField(blank=True,null=True)

    # -------------------------------------------------------------

    #Datos personales

    tipo_documento = models.CharField(max_length=100,blank=True,null=True)
    numero_documento = models.IntegerField(blank=True,null=True)
    primer_apellido = models.CharField(max_length=100,blank=True,null=True)
    segundo_apellido = models.CharField(max_length=100,blank=True, null=True)
    primer_nombre = models.CharField(max_length=100,blank=True,null=True)
    segundo_nombre = models.CharField(max_length=100,blank=True, null=True)
    celular_1 = models.CharField(max_length=100,blank=True,null=True)
    celular_2 = models.CharField(max_length=100,blank=True, null=True)
    correo_electronico = models.EmailField(max_length=100,blank=True, null=True)

    # Lugar y fecha de nacimiento

    departamento_nacimiento = models.ForeignKey(Departamentos, on_delete=models.DO_NOTHING,related_name='departamento_nacimiento_caracterizacion_inicial',blank=True,null=True)
    municipio_nacimiento = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_nacimiento_caracterizacion_inicial',blank=True,null=True)
    fecha_nacimiento = models.DateField(blank=True,null=True)

    # Lugar y fecha de expedición del documento

    departamento_expedicion = models.ForeignKey(Departamentos, on_delete=models.DO_NOTHING,related_name='departamento_expedicion_caracterizacion_inicial',blank=True,null=True)
    municipio_expedicion = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING,related_name='municipio_expedicion_caracterizacion_inicial',blank=True,null=True)
    fecha_expedicion = models.DateField(blank=True,null=True)

    longitud = models.DecimalField(max_digits=15, decimal_places=10, blank=True, null=True)
    latitud = models.DecimalField(max_digits=15, decimal_places=10, blank=True, null=True)
    precision = models.DecimalField(max_digits=15, decimal_places=10, blank=True, null=True)
    altitud = models.DecimalField(max_digits=15, decimal_places=10, blank=True, null=True)

    # -------------------------------------------------------------

    # Caracteristicas generales

    sexo = models.CharField(max_length=100,blank=True,null=True)
    tiene_libreta = models.BooleanField(default=False,blank=True,null=True)
    numero_libreta = models.CharField(max_length=100, blank=True, null=True)

    identidad_genero = models.CharField(max_length=100, blank=True, null=True)
    condicion_sexual = models.CharField(max_length=100, blank=True, null=True)
    estado_civil = models.CharField(max_length=100,blank=True,null=True)
    etnia = models.CharField(max_length=100,blank=True,null=True)

    pueblo_indigena = models.ForeignKey(PueblosIndigenas, on_delete=models.DO_NOTHING, blank=True, null=True)#aparece si se selecciona "INDIGENA" en la etnia
    resguardo_indigena = models.ForeignKey(ResguardosIndigenas, on_delete=models.DO_NOTHING, blank=True, null=True)  # aparece si se selecciona "INDIGENA" en la etnia
    comunidad_indigena = models.ForeignKey(ComunidadesIndigenas, on_delete=models.DO_NOTHING, blank=True, null=True)  # aparece si se selecciona "INDIGENA" en la etnia
    lengua_nativa_indigena = models.BooleanField(blank=True,null=True)  # aparece si se selecciona "INDIGENA" en la etnia
    cual_lengua_indigena = models.ForeignKey(LenguasNativas, related_name='lengua_indigena', on_delete=models.DO_NOTHING,blank=True,null=True)# aparece si se selecciona "INDIGENA" en la etnia y si se activa lengua nativa

    consejo_afro = models.ForeignKey(ConsejosAfro, on_delete=models.DO_NOTHING,blank=True,null=True)#aparece si se selecciona "AFROCOLOMBIANO" en la etnia
    comunidad_afro = models.ForeignKey(ComunidadesAfro, on_delete=models.DO_NOTHING, blank=True,null=True)  # aparece si se selecciona "AFROCOLOMBIANO" en la etnia
    lengua_nativa_afro = models.BooleanField(blank=True,null=True)  # aparece si se selecciona "AFROCOLOMBIANO" en la etnia
    cual_lengua_afro = models.ForeignKey(LenguasNativas, related_name='lengua_afro', on_delete=models.DO_NOTHING, blank=True,null=True)  # aparece si se selecciona "AFROCOLOMBIANO" en la etnia y si se activa lengua nativa

    discapacidad = models.BooleanField(blank=True,null=True)

    registro_discapacidad = models.CharField(max_length=100,blank=True,null=True) #aparece si hay discapacidad
    categoria_discapacidad = models.ManyToManyField(CategoriaDiscapacidad, blank=True) #aparece si hay discapacidad
    dificultades_permanentes = models.ManyToManyField(DificultadesPermanentesDiscapacidad, blank=True) #aparece si hay discapacidad
    utiliza_actualmente = models.ManyToManyField(ElementosDiscapacidad, blank=True) #aparece si hay discapacidad
    rehabilitacion = models.ManyToManyField(TiposRehabilitacionDiscapacidad, blank=True) #aparece si hay discapacidad
    tiene_cuidador = models.BooleanField(blank=True,null=True) #aparece si hay discapacidad
    cuidador = models.CharField(max_length=100,blank=True,null=True)

    parentezco = models.CharField(max_length=100,blank=True,null=True)
    es_jefe = models.BooleanField(blank=True,null=True)
    es_representante_hogar = models.BooleanField(blank=True,null=True)

    bancarizacion = models.BooleanField(blank=True,null=True)
    banco = models.ForeignKey(Bancos,on_delete=models.DO_NOTHING,blank=True,null=True) # se activa si hay bancarizacion
    tipo_cuenta = models.CharField(max_length=100, blank=True, null=True) # se activa si hay bancarizacion
    numero_cuenta = models.CharField(max_length=100, blank=True, null=True) # se activa si hay bancarizacion

    nivel_escolaridad = models.CharField(max_length=100,blank=True,null=True)
    grado_titulo = models.CharField(max_length=100,blank=True,null=True)
    sabe_leer = models.BooleanField(blank=True,null=True)
    sabe_sumar_restar = models.BooleanField(blank=True,null=True)
    actualmente_estudia = models.BooleanField(blank=True,null=True)
    recibe_alimentos = models.BooleanField(blank=True,null=True)

    razon_no_estudia = models.CharField(max_length=100, blank=True, null=True)  # se activa si no estudia
    razon_no_estudia_otra = models.CharField(max_length=100, blank=True, null=True)  # se activa si no estudia y hay otra razon
    regimen_seguridad_social = models.CharField(max_length=100,blank=True,null=True)


    def get_str_list_categoria_discapacidad(self):
        string = ', '

        for categoria in self.categoria_discapacidad.all():
            string += categoria.nombre + ', '

        return string[:-2]

    def get_str_list_dificultades_permanentes(self):
        string = ', '

        for categoria in self.dificultades_permanentes.all():
            string += categoria.nombre + ', '

        return string[:-2]

    def get_str_list_utiliza_actualmente(self):
        string = ', '

        for categoria in self.utiliza_actualmente.all():
            string += categoria.nombre + ', '

        return string[:-2]

    def get_str_list_rehabilitacion(self):
        string = ', '

        for categoria in self.rehabilitacion.all():
            string += categoria.nombre + ', '

        return string[:-2]




class Contactos(models.Model):

    id = models.UUIDField(primary_key=True, default=uuid.uuid4, unique=True, editable=False)
    municipio = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING, related_name='hcontactos_municipio')
    nombres = models.CharField(max_length=100)
    apellidos = models.CharField(max_length=100)
    cargo = models.CharField(max_length=100)
    celular = models.CharField(max_length=100)
    email = models.EmailField(max_length=100)
    resguardo = models.CharField(max_length=100)
    comunidad = models.CharField(max_length=100)
    lenguas = models.CharField(max_length=100,null=True,blank=True)


    def __str__(self):
        return self.nombres



class GeoreferenciacionApi(models.Model):
    creation = models.DateTimeField(auto_now_add=True)
    json = JSONField(default=dict)

    def __str__(self):
        try:
            name = f"Gestor: {self.json['documento']}"
        except:
            name = str(self.id)
        return name



def upload_dinamic_ficha_proyecto(instance, filename):
    return '/'.join(['IRACA', 'Proyectos', str(instance.id), filename])


class ProyectosApi(models.Model):
    json = JSONField(default=dict)
    creation = models.DateTimeField(auto_now_add=True)
    valor = MoneyField(max_digits=20, decimal_places=2, default_currency='COP', default=0)

    flujo_caja = JSONField(default=dict)

    estado = models.CharField(max_length=200, default='Cargado')

    actualizar_app = models.BooleanField(default=True)

    problematica_1_1 = models.TextField(blank=True, null=True)
    problematica_2_1 = models.TextField(blank=True, null=True)
    problematica_3_1 = models.TextField(blank=True, null=True)
    problematica_4_1 = models.TextField(blank=True, null=True)
    problematica_5_1 = models.TextField(blank=True, null=True)
    problematica_6_1 = models.TextField(blank=True, null=True)

    acciones_1_1 = models.TextField(blank=True, null=True)
    acciones_2_1 = models.TextField(blank=True, null=True)
    acciones_3_1 = models.TextField(blank=True, null=True)
    acciones_4_1 = models.TextField(blank=True, null=True)
    acciones_5_1 = models.TextField(blank=True, null=True)
    acciones_6_1 = models.TextField(blank=True, null=True)

    proyectos_potenciales_1_1 = models.TextField(blank=True, null=True)
    proyectos_potenciales_2_1 = models.TextField(blank=True, null=True)
    proyectos_potenciales_3_1 = models.TextField(blank=True, null=True)
    proyectos_potenciales_4_1 = models.TextField(blank=True, null=True)
    proyectos_potenciales_5_1 = models.TextField(blank=True, null=True)
    proyectos_potenciales_6_1 = models.TextField(blank=True, null=True)

    priorizacion_1_1 = models.TextField(blank=True, null=True)
    priorizacion_2_1 = models.TextField(blank=True, null=True)

    problematica_1_2 = models.TextField(blank=True, null=True)
    problematica_2_2 = models.TextField(blank=True, null=True)
    problematica_3_2 = models.TextField(blank=True, null=True)
    problematica_4_2 = models.TextField(blank=True, null=True)
    problematica_5_2 = models.TextField(blank=True, null=True)
    problematica_6_2 = models.TextField(blank=True, null=True)

    acciones_1_2 = models.TextField(blank=True, null=True)
    acciones_2_2 = models.TextField(blank=True, null=True)
    acciones_3_2 = models.TextField(blank=True, null=True)
    acciones_4_2 = models.TextField(blank=True, null=True)
    acciones_5_2 = models.TextField(blank=True, null=True)
    acciones_6_2 = models.TextField(blank=True, null=True)

    proyectos_potenciales_1_2 = models.TextField(blank=True, null=True)
    proyectos_potenciales_2_2 = models.TextField(blank=True, null=True)
    proyectos_potenciales_3_2 = models.TextField(blank=True, null=True)
    proyectos_potenciales_4_2 = models.TextField(blank=True, null=True)
    proyectos_potenciales_5_2 = models.TextField(blank=True, null=True)
    proyectos_potenciales_6_2 = models.TextField(blank=True, null=True)

    priorizacion_1_2 = models.TextField(blank=True, null=True)
    priorizacion_2_2 = models.TextField(blank=True, null=True)



    problematica_1_3 = models.TextField(blank=True, null=True)
    problematica_2_3 = models.TextField(blank=True, null=True)
    problematica_3_3 = models.TextField(blank=True, null=True)
    problematica_4_3 = models.TextField(blank=True, null=True)
    problematica_5_3 = models.TextField(blank=True, null=True)
    problematica_6_3 = models.TextField(blank=True, null=True)

    acciones_1_3 = models.TextField(blank=True, null=True)
    acciones_2_3 = models.TextField(blank=True, null=True)
    acciones_3_3 = models.TextField(blank=True, null=True)
    acciones_4_3 = models.TextField(blank=True, null=True)
    acciones_5_3 = models.TextField(blank=True, null=True)
    acciones_6_3 = models.TextField(blank=True, null=True)

    proyectos_potenciales_1_3 = models.TextField(blank=True, null=True)
    proyectos_potenciales_2_3 = models.TextField(blank=True, null=True)
    proyectos_potenciales_3_3 = models.TextField(blank=True, null=True)
    proyectos_potenciales_4_3 = models.TextField(blank=True, null=True)
    proyectos_potenciales_5_3 = models.TextField(blank=True, null=True)
    proyectos_potenciales_6_3 = models.TextField(blank=True, null=True)

    priorizacion_1_3 = models.TextField(blank=True, null=True)
    priorizacion_2_3 = models.TextField(blank=True, null=True)

    problematica_1_4 = models.TextField(blank=True, null=True)
    problematica_2_4 = models.TextField(blank=True, null=True)
    problematica_3_4 = models.TextField(blank=True, null=True)
    problematica_4_4 = models.TextField(blank=True, null=True)
    problematica_5_4 = models.TextField(blank=True, null=True)
    problematica_6_4 = models.TextField(blank=True, null=True)

    acciones_1_4 = models.TextField(blank=True, null=True)
    acciones_2_4 = models.TextField(blank=True, null=True)
    acciones_3_4 = models.TextField(blank=True, null=True)
    acciones_4_4 = models.TextField(blank=True, null=True)
    acciones_5_4 = models.TextField(blank=True, null=True)
    acciones_6_4 = models.TextField(blank=True, null=True)

    proyectos_potenciales_1_4 = models.TextField(blank=True, null=True)
    proyectos_potenciales_2_4 = models.TextField(blank=True, null=True)
    proyectos_potenciales_3_4 = models.TextField(blank=True, null=True)
    proyectos_potenciales_4_4 = models.TextField(blank=True, null=True)
    proyectos_potenciales_5_4 = models.TextField(blank=True, null=True)
    proyectos_potenciales_6_4 = models.TextField(blank=True, null=True)

    priorizacion_1_4 = models.TextField(blank=True, null=True)
    priorizacion_2_4 = models.TextField(blank=True, null=True)



    convenio = models.TextField(default='213-19')
    codigo_proyecto = models.TextField(blank=True, null=True)
    fecha_elaboracion = models.DateField(blank=True, null=True)
    municipio = models.ForeignKey(Municipios, on_delete=models.DO_NOTHING, related_name='municipio_ficha_proyecto',
                                  blank=True, null=True)
    resguado_indigena_consejo_comunitario = models.ForeignKey(ConsejosResguardosProyectosIraca,
                                                              on_delete=models.DO_NOTHING, blank=True, null=True)
    nombre_comunidad = models.ManyToManyField(ComunidadesProyectosIraca, blank=True)
    nombre_representante = models.TextField(blank=True, null=True)
    numero_hogares = models.IntegerField(blank=True, null=True)

    nombre_proyecto = models.TextField(blank=True, null=True)
    tipo_proyecto = models.TextField(blank=True, null=True)
    linea = models.TextField(blank=True, null=True)
    duracion = models.TextField(blank=True, null=True)
    ubicacion_proyecto = models.TextField(blank=True, null=True)
    producto_servicio = models.TextField(blank=True, null=True)
    problema = models.TextField(blank=True, null=True)
    justificacion = models.TextField(blank=True, null=True)
    criterios_socioculturales = models.TextField(blank=True, null=True)

    objetivo_general = models.TextField(blank=True, null=True)

    objetivo_especifico_1 = models.TextField(blank=True, null=True)
    objetivo_especifico_2 = models.TextField(blank=True, null=True)
    objetivo_especifico_3 = models.TextField(blank=True, null=True)

    actividad_1 = models.TextField(blank=True, null=True)
    mes_1_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_2_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_3_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_4_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_5_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_6_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_7_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_8_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_9_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_10_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_11_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_12_1 = models.PositiveIntegerField(default=0, blank=True, null=True)
    indicador_1 = models.TextField(blank=True, null=True)
    unidad_medida_1 = models.TextField(blank=True, null=True)
    meta_1 = models.IntegerField(blank=True, null=True)
    medio_verificacion_1 = models.TextField(blank=True, null=True)
    observaciones_1 = models.TextField(blank=True, null=True)

    actividad_2 = models.TextField(blank=True, null=True)
    mes_1_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_2_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_3_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_4_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_5_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_6_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_7_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_8_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_9_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_10_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_11_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_12_2 = models.PositiveIntegerField(default=0, blank=True, null=True)
    indicador_2 = models.TextField(blank=True, null=True)
    unidad_medida_2 = models.TextField(blank=True, null=True)
    meta_2 = models.IntegerField(blank=True, null=True)
    medio_verificacion_2 = models.TextField(blank=True, null=True)
    observaciones_2 = models.TextField(blank=True, null=True)

    actividad_3 = models.TextField(blank=True, null=True)
    mes_1_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_2_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_3_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_4_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_5_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_6_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_7_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_8_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_9_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_10_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_11_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_12_3 = models.PositiveIntegerField(default=0, blank=True, null=True)
    indicador_3 = models.TextField(blank=True, null=True)
    unidad_medida_3 = models.TextField(blank=True, null=True)
    meta_3 = models.IntegerField(blank=True, null=True)
    medio_verificacion_3 = models.TextField(blank=True, null=True)
    observaciones_3 = models.TextField(blank=True, null=True)

    actividad_4 = models.TextField(blank=True, null=True)
    mes_1_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_2_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_3_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_4_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_5_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_6_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_7_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_8_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_9_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_10_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_11_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_12_4 = models.PositiveIntegerField(default=0, blank=True, null=True)
    indicador_4 = models.TextField(blank=True, null=True)
    unidad_medida_4 = models.TextField(blank=True, null=True)
    meta_4 = models.IntegerField(blank=True, null=True)
    medio_verificacion_4 = models.TextField(blank=True, null=True)
    observaciones_4 = models.TextField(blank=True, null=True)

    actividad_5 = models.TextField(blank=True, null=True)
    mes_1_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_2_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_3_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_4_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_5_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_6_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_7_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_8_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_9_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_10_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_11_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_12_5 = models.PositiveIntegerField(default=0, blank=True, null=True)
    indicador_5 = models.TextField(blank=True, null=True)
    unidad_medida_5 = models.TextField(blank=True, null=True)
    meta_5 = models.IntegerField(blank=True, null=True)
    medio_verificacion_5 = models.TextField(blank=True, null=True)
    observaciones_5 = models.TextField(blank=True, null=True)

    actividad_6 = models.TextField(blank=True, null=True)
    mes_1_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_2_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_3_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_4_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_5_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_6_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_7_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_8_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_9_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_10_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_11_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_12_6 = models.PositiveIntegerField(default=0, blank=True, null=True)
    indicador_6 = models.TextField(blank=True, null=True)
    unidad_medida_6 = models.TextField(blank=True, null=True)
    meta_6 = models.IntegerField(blank=True, null=True)
    medio_verificacion_6 = models.TextField(blank=True, null=True)
    observaciones_6 = models.TextField(blank=True, null=True)

    actividad_7 = models.TextField(blank=True, null=True)
    mes_1_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_2_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_3_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_4_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_5_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_6_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_7_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_8_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_9_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_10_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_11_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_12_7 = models.PositiveIntegerField(default=0, blank=True, null=True)
    indicador_7 = models.TextField(blank=True, null=True)
    unidad_medida_7 = models.TextField(blank=True, null=True)
    meta_7 = models.IntegerField(blank=True, null=True)
    medio_verificacion_7 = models.TextField(blank=True, null=True)
    observaciones_7 = models.TextField(blank=True, null=True)

    actividad_8 = models.TextField(blank=True, null=True)
    mes_1_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_2_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_3_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_4_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_5_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_6_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_7_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_8_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_9_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_10_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_11_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_12_8 = models.PositiveIntegerField(default=0, blank=True, null=True)
    indicador_8 = models.TextField(blank=True, null=True)
    unidad_medida_8 = models.TextField(blank=True, null=True)
    meta_8 = models.IntegerField(blank=True, null=True)
    medio_verificacion_8 = models.TextField(blank=True, null=True)
    observaciones_8 = models.TextField(blank=True, null=True)

    actividad_9 = models.TextField(blank=True, null=True)
    mes_1_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_2_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_3_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_4_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_5_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_6_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_7_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_8_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_9_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_10_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_11_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_12_9 = models.PositiveIntegerField(default=0, blank=True, null=True)
    indicador_9 = models.TextField(blank=True, null=True)
    unidad_medida_9 = models.TextField(blank=True, null=True)
    meta_9 = models.IntegerField(blank=True, null=True)
    medio_verificacion_9 = models.TextField(blank=True, null=True)
    observaciones_9 = models.TextField(blank=True, null=True)

    actividad_10 = models.TextField(blank=True, null=True)
    mes_1_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_2_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_3_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_4_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_5_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_6_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_7_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_8_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_9_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_10_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_11_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    mes_12_10 = models.PositiveIntegerField(default=0, blank=True, null=True)
    indicador_10 = models.TextField(blank=True, null=True)
    unidad_medida_10 = models.TextField(blank=True, null=True)
    meta_10 = models.IntegerField(blank=True, null=True)
    medio_verificacion_10 = models.TextField(blank=True, null=True)
    observaciones_10 = models.TextField(blank=True, null=True)

    conservacion_manejo_ambiental = models.TextField(blank=True, null=True)
    sustentabilidad = models.TextField(blank=True, null=True)
    riesgos_acciones = models.TextField(blank=True, null=True)

    aliado_1 = models.TextField(blank=True, null=True)
    aporte_aliado_1 = models.TextField(blank=True, null=True)
    nombre_aliado_1 = models.TextField(blank=True, null=True)
    datos_contacto_aliado_1 = models.TextField(blank=True, null=True)

    aliado_2 = models.TextField(blank=True, null=True)
    aporte_aliado_2 = models.TextField(blank=True, null=True)
    nombre_aliado_2 = models.TextField(blank=True, null=True)
    datos_contacto_aliado_2 = models.TextField(blank=True, null=True)

    aliado_3 = models.TextField(blank=True, null=True)
    aporte_aliado_3 = models.TextField(blank=True, null=True)
    nombre_aliado_3 = models.TextField(blank=True, null=True)
    datos_contacto_aliado_3 = models.TextField(blank=True, null=True)

    aliado_4 = models.TextField(blank=True, null=True)
    aporte_aliado_4 = models.TextField(blank=True, null=True)
    nombre_aliado_4 = models.TextField(blank=True, null=True)
    datos_contacto_aliado_4 = models.TextField(blank=True, null=True)

    concepto_tecnico = models.TextField(blank=True, null=True)

    anexo_1 = models.TextField(blank=True, null=True)
    anexo_2 = models.TextField(blank=True, null=True)
    anexo_3 = models.TextField(blank=True, null=True)
    anexo_4 = models.TextField(blank=True, null=True)

    nombre_representante_consejo = models.TextField(blank=True, null=True)
    cedula_representante_consejo = models.BigIntegerField(blank=True, null=True)

    nombre_representante_comite = models.TextField(blank=True, null=True)
    cedula_representante_comite = models.BigIntegerField(blank=True, null=True)

    nombre_funcionario = models.TextField(blank=True, null=True)
    cedula_funcionario = models.BigIntegerField(blank=True, null=True)

    file = ContentTypeRestrictedFileField(
        upload_to=upload_dinamic_ficha_proyecto,
        content_types=[
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ],
        max_upload_size=20485760,
        max_length=255,
        blank=True, null=True
    )

    file2 = models.FileField(
        upload_to=upload_dinamic_ficha_proyecto,
        max_length=255,
        blank=True, null=True
    )

    file3 = models.FileField(
        upload_to=upload_dinamic_ficha_proyecto,
        max_length=255,
        blank=True, null=True
    )

    file4 = models.FileField(
        upload_to=upload_dinamic_ficha_proyecto,
        max_length=255,
        blank=True, null=True
    )

    file5 = models.FileField(
        upload_to=upload_dinamic_ficha_proyecto,
        max_length=255,
        blank=True, null=True
    )


    def get_hogares(self):
        managers = []

        for manager in self.json['data']['managers']:

            verificado = False
            nombre_verificado = ''

            if Hogares.objects.filter(documento = manager['document']).count() > 0:
                verificado = True
                nombre_verificado = Hogares.objects.filter(documento = manager['document'])[0].get_full_name()

            managers.append({
                'name': manager['name'],
                'document': manager['document'],
                'verificado': verificado,
                'nombre_verificado': nombre_verificado,
            })

        return managers


    def url_file(self):
        url = None
        try:
            url = self.file.url
        except:
            pass
        return url

    def url_file2(self):
        url = None
        try:
            url = self.file2.url
        except:
            pass
        return url

    def url_file3(self):
        url = None
        try:
            url = self.file3.url
        except:
            pass
        return url

    def url_file4(self):
        url = None
        try:
            url = self.file4.url
        except:
            pass
        return url

    def url_file5(self):
        url = None
        try:
            url = self.file5.url
        except:
            pass
        return url


    def __str__(self):
        try:
            name = f"Gestor: {self.json['documento']}"
        except:
            name = str(self.id)
        return name


    def get_documentos_managers(self):
        documentos = []

        for manager in self.json['data']['managers']:
            documentos.append(manager['document'])

        return documentos

    def get_comunidades(self):
        nombre = ''

        for comunidad in self.nombre_comunidad.all():
            nombre += f'{comunidad.nombre}, '

        return nombre[:-2]


    def get_observaciones(self):
        return ObservacionesProyectosApi.objects.filter(proyecto = self).order_by("-creation")


    def agregar_observacion(self, user, estado, descripcion):
        obs = ObservacionesProyectosApi.objects.create(proyecto=self, user = user, estado = estado, descripcion = descripcion)
        return obs


class ObservacionesProyectosApi(models.Model):
    proyecto = models.ForeignKey(ProyectosApi, on_delete=models.DO_NOTHING)
    user = models.ForeignKey(User, on_delete=models.DO_NOTHING, blank=True, null=True)
    creation = models.DateTimeField(auto_now_add=True)
    estado = models.CharField(max_length=200)
    descripcion = models.TextField()

    def get_day(self):
        return self.creation.astimezone(settings_time_zone).strftime('%d')

    def get_hora(self):
        return self.creation.astimezone(settings_time_zone).strftime('%I:%M:%S %p')

    def get_mes(self):
        meses = ['ENE','FEB','MAR','ABR','MAY','JUN','JUL','AGO','SEP','OCT','NOV','DIC']
        mes = int(self.creation.astimezone(settings_time_zone).strftime('%m')) - 1
        return meses[mes]



class PermisosMisProyectos(models.Model):
    user = models.ForeignKey(User, on_delete=models.DO_NOTHING)


    def __str__(self):
        return str(self.user)



@receiver(post_save, sender=ProyectosApi)
def ProyectosApiPostSave(sender, instance, **kwargs):

    try:

        valor = 0
        numero_hogares = 0

        for linea in instance.json['data']['lines']:


            for producto in linea['supplies']:

                valor += float(producto["price"]) * (float(producto["count"]["countIRACA"]) + float(producto["count"]["countOthers"]) + float(producto["count"]["countCommunity"]))



            try:
                homes_linea = int(linea['homes'])
            except:
                homes_linea = 0

            numero_hogares += homes_linea


        #valor = int(instance.json['data']['budgetIRACAUsed']) + int(instance.json['data']['budgedCommunity']) + int(instance.json['data']['budgedOthers'])
        convenio = instance.json['data']['agreement']
        codigo_proyecto = instance.json['data']['projectCode']
        fecha_elaboracion = dateutil.parser.parse(instance.json['data']['creationDate'])
        #fecha_elaboracion = datetime.datetime.fromtimestamp(instance.json['data']['createdAt'] / 2e3)
        nombre_representante = instance.json['data']['legalRepresentative']

        nombre_proyecto = instance.json['data']['projectName']
        linea = instance.json['data']['line']
        duracion = instance.json['data']['duration']
        ubicacion_proyecto = instance.json['data']['projectLocation']
        producto_servicio = instance.json['data']['productService']
        problema = instance.json['data']['problem']
        justificacion = instance.json['data']['justification']
        tipo_proyecto = instance.json['data']['project_type']
        criterios_socioculturales = instance.json['data']['criterion']
        objetivo_general = instance.json['data']['objective']
        objetivo_especifico_1 = instance.json['data']['specificObjectives'][0]
        objetivo_especifico_2 = instance.json['data']['specificObjectives'][1]
        objetivo_especifico_3 = instance.json['data']['specificObjectives'][2]

        conservacion_manejo_ambiental = instance.json['data']['environmentalManagement']
        sustentabilidad = instance.json['data']['sustainability']
        riesgos_acciones = instance.json['data']['risks']

        #concepto_tecnico = instance.json['data']['technicalConcept']

        nombre_representante_consejo = instance.json['data']['RepresentativeCouncil']['name']
        cedula_representante_consejo = instance.json['data']['RepresentativeCouncil']['document']

        nombre_representante_comite = instance.json['data']['RepresentativeCommittee']['name']
        cedula_representante_comite = instance.json['data']['RepresentativeCommittee']['document']

        nombre_funcionario = instance.json['data']['Official']['name']
        cedula_funcionario = instance.json['data']['Official']['document']

    except:
        pass

    else:

        post_save.disconnect(ProyectosApiPostSave, sender=ProyectosApi)

        if instance.actualizar_app:

            instance.valor = valor
            instance.convenio = convenio
            instance.tipo_proyecto = tipo_proyecto
            instance.codigo_proyecto = codigo_proyecto
            instance.fecha_elaboracion = fecha_elaboracion
            instance.nombre_representante = nombre_representante
            instance.numero_hogares = numero_hogares
            instance.nombre_proyecto = nombre_proyecto
            instance.linea = linea
            instance.duracion = duracion
            instance.ubicacion_proyecto = ubicacion_proyecto
            instance.producto_servicio = producto_servicio
            instance.problema = problema
            instance.justificacion = justificacion
            instance.criterios_socioculturales = criterios_socioculturales
            instance.objetivo_general = objetivo_general
            instance.objetivo_especifico_1 = objetivo_especifico_1
            instance.objetivo_especifico_2 = objetivo_especifico_2
            instance.objetivo_especifico_3 = objetivo_especifico_3

            instance.conservacion_manejo_ambiental = conservacion_manejo_ambiental
            instance.sustentabilidad = sustentabilidad
            instance.riesgos_acciones = riesgos_acciones

            instance.nombre_representante_consejo = nombre_representante_consejo
            instance.cedula_representante_consejo = cedula_representante_consejo

            instance.nombre_representante_comite = nombre_representante_comite
            instance.cedula_representante_comite = cedula_representante_comite

            instance.nombre_funcionario = nombre_funcionario
            instance.cedula_funcionario = cedula_funcionario

            instance.save()


        output = BytesIO()

        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0] + '/documentos/ficha_proyecto.xlsx')

        ws = wb.get_sheet_by_name('IDENTIFICACION PROYECTOS')


        ws['B7'] = instance.problematica_1_1
        ws['B8'] = instance.problematica_2_1
        ws['B9'] = instance.problematica_3_1
        ws['B10'] = instance.problematica_4_1
        ws['B11'] = instance.problematica_5_1
        ws['B12'] = instance.problematica_6_1

        ws['C7'] = instance.acciones_1_1
        ws['C8'] = instance.acciones_2_1
        ws['C9'] = instance.acciones_3_1
        ws['C10'] = instance.acciones_4_1
        ws['C11'] = instance.acciones_5_1
        ws['C12'] = instance.acciones_6_1

        ws['D7'] = instance.proyectos_potenciales_1_1
        ws['D8'] = instance.proyectos_potenciales_2_1
        ws['D9'] = instance.proyectos_potenciales_3_1
        ws['D10'] = instance.proyectos_potenciales_4_1
        ws['D11'] = instance.proyectos_potenciales_5_1
        ws['D12'] = instance.proyectos_potenciales_6_1

        ws['E7'] = instance.priorizacion_1_1
        ws['E10'] = instance.priorizacion_2_1



        ws['B13'] = instance.problematica_1_2
        ws['B14'] = instance.problematica_2_2
        ws['B15'] = instance.problematica_3_2
        ws['B16'] = instance.problematica_4_2
        ws['B17'] = instance.problematica_5_2
        ws['B18'] = instance.problematica_6_2

        ws['C13'] = instance.acciones_1_2
        ws['C14'] = instance.acciones_2_2
        ws['C15'] = instance.acciones_3_2
        ws['C16'] = instance.acciones_4_2
        ws['C17'] = instance.acciones_5_2
        ws['C18'] = instance.acciones_6_2

        ws['D13'] = instance.proyectos_potenciales_1_2
        ws['D14'] = instance.proyectos_potenciales_2_2
        ws['D15'] = instance.proyectos_potenciales_3_2
        ws['D16'] = instance.proyectos_potenciales_4_2
        ws['D17'] = instance.proyectos_potenciales_5_2
        ws['D18'] = instance.proyectos_potenciales_6_2

        ws['E13'] = instance.priorizacion_1_2
        ws['E16'] = instance.priorizacion_2_2



        ws['B19'] = instance.problematica_1_3
        ws['B20'] = instance.problematica_2_3
        ws['B21'] = instance.problematica_3_3
        ws['B22'] = instance.problematica_4_3
        ws['B23'] = instance.problematica_5_3
        ws['B24'] = instance.problematica_6_3

        ws['C19'] = instance.acciones_1_3
        ws['C20'] = instance.acciones_2_3
        ws['C21'] = instance.acciones_3_3
        ws['C22'] = instance.acciones_4_3
        ws['C23'] = instance.acciones_5_3
        ws['C24'] = instance.acciones_6_3

        ws['D19'] = instance.proyectos_potenciales_1_3
        ws['D20'] = instance.proyectos_potenciales_2_3
        ws['D21'] = instance.proyectos_potenciales_3_3
        ws['D22'] = instance.proyectos_potenciales_4_3
        ws['D23'] = instance.proyectos_potenciales_5_3
        ws['D24'] = instance.proyectos_potenciales_6_3

        ws['E19'] = instance.priorizacion_1_3
        ws['E22'] = instance.priorizacion_2_3

        ws['B25'] = instance.problematica_1_4
        ws['B26'] = instance.problematica_2_4
        ws['B27'] = instance.problematica_3_4
        ws['B28'] = instance.problematica_4_4
        ws['B29'] = instance.problematica_5_4
        ws['B30'] = instance.problematica_6_4

        ws['C25'] = instance.acciones_1_4
        ws['C26'] = instance.acciones_2_4
        ws['C27'] = instance.acciones_3_4
        ws['C28'] = instance.acciones_4_4
        ws['C29'] = instance.acciones_5_4
        ws['C30'] = instance.acciones_6_4

        ws['D25'] = instance.proyectos_potenciales_1_4
        ws['D26'] = instance.proyectos_potenciales_2_4
        ws['D27'] = instance.proyectos_potenciales_3_4
        ws['D28'] = instance.proyectos_potenciales_4_4
        ws['D29'] = instance.proyectos_potenciales_5_4
        ws['D30'] = instance.proyectos_potenciales_6_4

        ws['E25'] = instance.priorizacion_1_4
        ws['E28'] = instance.priorizacion_2_4



        ws = wb.get_sheet_by_name('Ficha de proyecto')

        ws['G6'] = instance.convenio
        ws['R6'] = instance.codigo_proyecto
        ws['AI6'] = instance.fecha_elaboracion.strftime('%d/%m/%Y')

        if instance.municipio != None:
            ws['G8'] = instance.municipio.departamento.nombre
            ws['P8'] = instance.municipio.nombre

        if instance.resguado_indigena_consejo_comunitario != None:
            ws['Z8'] = instance.resguado_indigena_consejo_comunitario.nombre

        ws['K9'] = instance.get_comunidades()

        ws['Z9'] = instance.nombre_representante
        ws['AM9'] = instance.numero_hogares
        ws['K11'] = instance.nombre_proyecto
        ws['N12'] = instance.linea

        if instance.duracion == '1':
            ws['L13'] = f'{instance.duracion} mes'
        else:
            ws['L13'] = f'{instance.duracion} meses'

        ws['Z13'] = instance.ubicacion_proyecto
        ws['AO13'] = instance.producto_servicio
        ws['C15'] = instance.problema
        ws['C17'] = instance.justificacion
        ws['C19'] = instance.criterios_socioculturales
        ws['C21'] = instance.objetivo_general
        ws['D23'] = instance.objetivo_especifico_1
        ws['D24'] = instance.objetivo_especifico_2
        ws['D25'] = instance.objetivo_especifico_3

        ws['D31'] = instance.actividad_1
        if instance.mes_1_1 != 0:
            ws['N31'] = instance.mes_1_1
        if instance.mes_2_1 != 0:
            ws['O31'] = instance.mes_2_1
        if instance.mes_3_1 != 0:
            ws['P31'] = instance.mes_3_1
        if instance.mes_4_1 != 0:
            ws['Q31'] = instance.mes_4_1
        if instance.mes_5_1 != 0:
            ws['R31'] = instance.mes_5_1
        if instance.mes_6_1 != 0:
            ws['S31'] = instance.mes_6_1
        if instance.mes_7_1 != 0:
            ws['T31'] = instance.mes_7_1
        if instance.mes_8_1 != 0:
            ws['U31'] = instance.mes_8_1
        if instance.mes_9_1 != 0:
            ws['V31'] = instance.mes_9_1
        if instance.mes_10_1 != 0:
            ws['W31'] = instance.mes_10_1
        if instance.mes_11_1 != 0:
            ws['X31'] = instance.mes_11_1
        if instance.mes_12_1 != 0:
            ws['Y31'] = instance.mes_12_1
        ws['Z31'] = instance.indicador_1
        ws['AF31'] = instance.unidad_medida_1
        ws['AI31'] = instance.meta_1
        ws['AL31'] = instance.medio_verificacion_1
        ws['AQ31'] = instance.observaciones_1


        ws['D32'] = instance.actividad_2
        if instance.mes_1_2 != 0:
            ws['N32'] = instance.mes_1_2
        if instance.mes_2_2 != 0:
            ws['O32'] = instance.mes_2_2
        if instance.mes_3_2 != 0:
            ws['P32'] = instance.mes_3_2
        if instance.mes_4_2 != 0:
            ws['Q32'] = instance.mes_4_2
        if instance.mes_5_2 != 0:
            ws['R32'] = instance.mes_5_2
        if instance.mes_6_2 != 0:
            ws['S32'] = instance.mes_6_2
        if instance.mes_7_2 != 0:
            ws['T32'] = instance.mes_7_2
        if instance.mes_8_2 != 0:
            ws['U32'] = instance.mes_8_2
        if instance.mes_9_2 != 0:
            ws['V32'] = instance.mes_9_2
        if instance.mes_10_2 != 0:
            ws['W32'] = instance.mes_10_2
        if instance.mes_11_2 != 0:
            ws['X32'] = instance.mes_11_2
        if instance.mes_12_2 != 0:
            ws['Y32'] = instance.mes_12_2
        ws['Z32'] = instance.indicador_2
        ws['AF32'] = instance.unidad_medida_2
        ws['AI32'] = instance.meta_2
        ws['AL32'] = instance.medio_verificacion_2
        ws['AQ32'] = instance.observaciones_2



        ws['D33'] = instance.actividad_3
        if instance.mes_1_3 != 0:
            ws['N33'] = instance.mes_1_3
        if instance.mes_2_3 != 0:
            ws['O33'] = instance.mes_2_3
        if instance.mes_3_3 != 0:
            ws['P33'] = instance.mes_3_3
        if instance.mes_4_3 != 0:
            ws['Q33'] = instance.mes_4_3
        if instance.mes_5_3 != 0:
            ws['R33'] = instance.mes_5_3
        if instance.mes_6_3 != 0:
            ws['S33'] = instance.mes_6_3
        if instance.mes_7_3 != 0:
            ws['T33'] = instance.mes_7_3
        if instance.mes_8_3 != 0:
            ws['U33'] = instance.mes_8_3
        if instance.mes_9_3 != 0:
            ws['V33'] = instance.mes_9_3
        if instance.mes_10_3 != 0:
            ws['W33'] = instance.mes_10_3
        if instance.mes_11_3 != 0:
            ws['X33'] = instance.mes_11_3
        if instance.mes_12_3 != 0:
            ws['Y33'] = instance.mes_12_3
        ws['Z33'] = instance.indicador_3
        ws['AF33'] = instance.unidad_medida_3
        ws['AI33'] = instance.meta_3
        ws['AL33'] = instance.medio_verificacion_3
        ws['AQ33'] = instance.observaciones_3




        ws['D34'] = instance.actividad_4
        if instance.mes_1_4 != 0:
            ws['N34'] = instance.mes_1_4
        if instance.mes_2_4 != 0:
            ws['O34'] = instance.mes_2_4
        if instance.mes_3_4 != 0:
            ws['P34'] = instance.mes_3_4
        if instance.mes_4_4 != 0:
            ws['Q34'] = instance.mes_4_4
        if instance.mes_5_4 != 0:
            ws['R34'] = instance.mes_5_4
        if instance.mes_6_4 != 0:
            ws['S34'] = instance.mes_6_4
        if instance.mes_7_4 != 0:
            ws['T34'] = instance.mes_7_4
        if instance.mes_8_4 != 0:
            ws['U34'] = instance.mes_8_4
        if instance.mes_9_4 != 0:
            ws['V34'] = instance.mes_9_4
        if instance.mes_10_4 != 0:
            ws['W34'] = instance.mes_10_4
        if instance.mes_11_4 != 0:
            ws['X34'] = instance.mes_11_4
        if instance.mes_12_4 != 0:
            ws['Y34'] = instance.mes_12_4
        ws['Z34'] = instance.indicador_4
        ws['AF34'] = instance.unidad_medida_4
        ws['AI34'] = instance.meta_4
        ws['AL34'] = instance.medio_verificacion_4
        ws['AQ34'] = instance.observaciones_4



        ws['D35'] = instance.actividad_5
        if instance.mes_1_5 != 0:
            ws['N35'] = instance.mes_1_5
        if instance.mes_2_5 != 0:
            ws['O35'] = instance.mes_2_5
        if instance.mes_3_5 != 0:
            ws['P35'] = instance.mes_3_5
        if instance.mes_4_5 != 0:
            ws['Q35'] = instance.mes_4_5
        if instance.mes_5_5 != 0:
            ws['R35'] = instance.mes_5_5
        if instance.mes_6_5 != 0:
            ws['S35'] = instance.mes_6_5
        if instance.mes_7_5 != 0:
            ws['T35'] = instance.mes_7_5
        if instance.mes_8_5 != 0:
            ws['U35'] = instance.mes_8_5
        if instance.mes_9_5 != 0:
            ws['V35'] = instance.mes_9_5
        if instance.mes_10_5 != 0:
            ws['W35'] = instance.mes_10_5
        if instance.mes_11_5 != 0:
            ws['X35'] = instance.mes_11_5
        if instance.mes_12_5 != 0:
            ws['Y35'] = instance.mes_12_5
        ws['Z35'] = instance.indicador_5
        ws['AF35'] = instance.unidad_medida_5
        ws['AI35'] = instance.meta_5
        ws['AL35'] = instance.medio_verificacion_5
        ws['AQ35'] = instance.observaciones_5




        ws['D36'] = instance.actividad_6
        if instance.mes_1_6 != 0:
            ws['N36'] = instance.mes_1_6
        if instance.mes_2_6 != 0:
            ws['O36'] = instance.mes_2_6
        if instance.mes_3_6 != 0:
            ws['P36'] = instance.mes_3_6
        if instance.mes_4_6 != 0:
            ws['Q36'] = instance.mes_4_6
        if instance.mes_5_6 != 0:
            ws['R36'] = instance.mes_5_6
        if instance.mes_6_6 != 0:
            ws['S36'] = instance.mes_6_6
        if instance.mes_7_6 != 0:
            ws['T36'] = instance.mes_7_6
        if instance.mes_8_6 != 0:
            ws['U36'] = instance.mes_8_6
        if instance.mes_9_6 != 0:
            ws['V36'] = instance.mes_9_6
        if instance.mes_10_6 != 0:
            ws['W36'] = instance.mes_10_6
        if instance.mes_11_6 != 0:
            ws['X36'] = instance.mes_11_6
        if instance.mes_12_6 != 0:
            ws['Y36'] = instance.mes_12_6
        ws['Z36'] = instance.indicador_6
        ws['AF36'] = instance.unidad_medida_6
        ws['AI36'] = instance.meta_6
        ws['AL36'] = instance.medio_verificacion_6
        ws['AQ36'] = instance.observaciones_6




        ws['D37'] = instance.actividad_7
        if instance.mes_1_7 != 0:
            ws['N37'] = instance.mes_1_7
        if instance.mes_2_7 != 0:
            ws['O37'] = instance.mes_2_7
        if instance.mes_3_7 != 0:
            ws['P37'] = instance.mes_3_7
        if instance.mes_4_7 != 0:
            ws['Q37'] = instance.mes_4_7
        if instance.mes_5_7 != 0:
            ws['R37'] = instance.mes_5_7
        if instance.mes_6_7 != 0:
            ws['S37'] = instance.mes_6_7
        if instance.mes_7_7 != 0:
            ws['T37'] = instance.mes_7_7
        if instance.mes_8_7 != 0:
            ws['U37'] = instance.mes_8_7
        if instance.mes_9_7 != 0:
            ws['V37'] = instance.mes_9_7
        if instance.mes_10_7 != 0:
            ws['W37'] = instance.mes_10_7
        if instance.mes_11_7 != 0:
            ws['X37'] = instance.mes_11_7
        if instance.mes_12_7 != 0:
            ws['Y37'] = instance.mes_12_7
        ws['Z37'] = instance.indicador_7
        ws['AF37'] = instance.unidad_medida_7
        ws['AI37'] = instance.meta_7
        ws['AL37'] = instance.medio_verificacion_7
        ws['AQ37'] = instance.observaciones_7




        ws['D38'] = instance.actividad_8
        if instance.mes_1_8 != 0:
            ws['N38'] = instance.mes_1_8
        if instance.mes_2_8 != 0:
            ws['O38'] = instance.mes_2_8
        if instance.mes_3_8 != 0:
            ws['P38'] = instance.mes_3_8
        if instance.mes_4_8 != 0:
            ws['Q38'] = instance.mes_4_8
        if instance.mes_5_8 != 0:
            ws['R38'] = instance.mes_5_8
        if instance.mes_6_8 != 0:
            ws['S38'] = instance.mes_6_8
        if instance.mes_7_8 != 0:
            ws['T38'] = instance.mes_7_8
        if instance.mes_8_8 != 0:
            ws['U38'] = instance.mes_8_8
        if instance.mes_9_8 != 0:
            ws['V38'] = instance.mes_9_8
        if instance.mes_10_8 != 0:
            ws['W38'] = instance.mes_10_8
        if instance.mes_11_8 != 0:
            ws['X38'] = instance.mes_11_8
        if instance.mes_12_8 != 0:
            ws['Y38'] = instance.mes_12_8
        ws['Z38'] = instance.indicador_8
        ws['AF38'] = instance.unidad_medida_8
        ws['AI38'] = instance.meta_8
        ws['AL38'] = instance.medio_verificacion_8
        ws['AQ38'] = instance.observaciones_8




        ws['D39'] = instance.actividad_9
        if instance.mes_1_9 != 0:
            ws['N39'] = instance.mes_1_9
        if instance.mes_2_9 != 0:
            ws['O39'] = instance.mes_2_9
        if instance.mes_3_9 != 0:
            ws['P39'] = instance.mes_3_9
        if instance.mes_4_9 != 0:
            ws['Q39'] = instance.mes_4_9
        if instance.mes_5_9 != 0:
            ws['R39'] = instance.mes_5_9
        if instance.mes_6_9 != 0:
            ws['S39'] = instance.mes_6_9
        if instance.mes_7_9 != 0:
            ws['T39'] = instance.mes_7_9
        if instance.mes_8_9 != 0:
            ws['U39'] = instance.mes_8_9
        if instance.mes_9_9 != 0:
            ws['V39'] = instance.mes_9_9
        if instance.mes_10_9 != 0:
            ws['W39'] = instance.mes_10_9
        if instance.mes_11_9 != 0:
            ws['X39'] = instance.mes_11_9
        if instance.mes_12_9 != 0:
            ws['Y39'] = instance.mes_12_9
        ws['Z39'] = instance.indicador_9
        ws['AF39'] = instance.unidad_medida_9
        ws['AI39'] = instance.meta_9
        ws['AL39'] = instance.medio_verificacion_9
        ws['AQ39'] = instance.observaciones_9




        ws['D40'] = instance.actividad_10
        if instance.mes_1_10 != 0:
            ws['N40'] = instance.mes_1_10
        if instance.mes_2_10 != 0:
            ws['O40'] = instance.mes_2_10
        if instance.mes_3_10 != 0:
            ws['P40'] = instance.mes_3_10
        if instance.mes_4_10 != 0:
            ws['Q40'] = instance.mes_4_10
        if instance.mes_5_10 != 0:
            ws['R40'] = instance.mes_5_10
        if instance.mes_6_10 != 0:
            ws['S40'] = instance.mes_6_10
        if instance.mes_7_10 != 0:
            ws['T40'] = instance.mes_7_10
        if instance.mes_8_10 != 0:
            ws['U40'] = instance.mes_8_10
        if instance.mes_9_10 != 0:
            ws['V40'] = instance.mes_9_10
        if instance.mes_10_10 != 0:
            ws['W40'] = instance.mes_10_10
        if instance.mes_11_10 != 0:
            ws['X40'] = instance.mes_11_10
        if instance.mes_12_10 != 0:
            ws['Y40'] = instance.mes_12_10
        ws['Z40'] = instance.indicador_10
        ws['AF40'] = instance.unidad_medida_10
        ws['AI40'] = instance.meta_10
        ws['AL40'] = instance.medio_verificacion_10
        ws['AQ40'] = instance.observaciones_10


        ws['C50'] = instance.aliado_1
        ws['K50'] = instance.aporte_aliado_1
        ws['AA50'] = instance.nombre_aliado_1
        ws['AK50'] = instance.datos_contacto_aliado_1

        ws['C51'] = instance.aliado_2
        ws['K51'] = instance.aporte_aliado_2
        ws['AA51'] = instance.nombre_aliado_2
        ws['AK51'] = instance.datos_contacto_aliado_2

        ws['C52'] = instance.aliado_3
        ws['K52'] = instance.aporte_aliado_3
        ws['AA52'] = instance.nombre_aliado_3
        ws['AK52'] = instance.datos_contacto_aliado_3

        ws['C53'] = instance.aliado_4
        ws['K53'] = instance.aporte_aliado_4
        ws['AA53'] = instance.nombre_aliado_4
        ws['AK53'] = instance.datos_contacto_aliado_4



        ws['C43'] = instance.conservacion_manejo_ambiental
        ws['C45'] = instance.sustentabilidad
        ws['C47'] = instance.riesgos_acciones

        ws['C55'] = instance.concepto_tecnico

        ws['E62'] = instance.nombre_representante_consejo
        ws['E63'] = instance.cedula_representante_consejo

        ws['S62'] = instance.nombre_representante_comite
        ws['S63'] = instance.cedula_representante_comite

        ws['AJ62'] = instance.nombre_funcionario
        ws['AJ63'] = instance.cedula_funcionario


        ws['C57'] = f"Anexo 1: {instance.anexo_1}"
        ws['C58'] = f"Anexo 2: {instance.anexo_2}"
        ws['AA57'] = f"Anexo 3: {instance.anexo_3}"
        ws['AA58'] = f"Anexo 4: {instance.anexo_4}"


        index = 1

        for linea in instance.json['data']['lines']:
            ws = wb.get_sheet_by_name('Plan de inversion')
            ws = wb.copy_worksheet(ws)
            ws.title = f'Linea {index}'

            ws.add_image(Image(settings.STATICFILES_DIRS[0] + '/img/logo_prosperidad_PI.png'), 'A1')
            ws.add_image(Image(settings.STATICFILES_DIRS[0] + '/img/logo_ut_PI.png'), 'A2')


            if instance.municipio != None:
                ws['O6'] = instance.municipio.departamento.nombre
                ws['O7'] = instance.municipio.nombre

            ws['E7'] = instance.codigo_proyecto

            if instance.resguado_indigena_consejo_comunitario != None:
                ws['S6'] = instance.resguado_indigena_consejo_comunitario.nombre

            ws['S7'] = instance.get_comunidades()

            ws['V6'] = int(linea['homes'])

            try:
                productos_list = linea['supplies']
            except:
                pass
            else:

                i = 1

                if len(productos_list) > 1:
                    ws.insert_rows(11, amount=len(productos_list) - 1)

                for product in productos_list:

                    ws.row_dimensions[9 + i].height = 50

                    ws[f'A{9 + i}'].value = i
                    ws[f'A{9 + i}']._style = copy(ws[f'A10']._style)

                    try:
                        ws.merge_cells(f'B{9 + i}:F{9 + i}')
                        ws[f'B{9 + i}'].value = product['name']
                        ws[f'B{9 + i}']._style = copy(ws[f'B10']._style)
                        ws[f'C{9 + i}']._style = copy(ws[f'C10']._style)
                        ws[f'D{9 + i}']._style = copy(ws[f'D10']._style)
                        ws[f'E{9 + i}']._style = copy(ws[f'E10']._style)
                        ws[f'F{9 + i}']._style = copy(ws[f'F10']._style)

                    except:
                        pass

                    try:
                        ws.merge_cells(f'G{9 + i}:L{9 + i}')
                        ws[f'G{9 + i}'].value = product['description']
                        ws[f'G{9 + i}']._style = copy(ws[f'G10']._style)
                        ws[f'H{9 + i}']._style = copy(ws[f'H10']._style)
                        ws[f'I{9 + i}']._style = copy(ws[f'I10']._style)
                        ws[f'J{9 + i}']._style = copy(ws[f'J10']._style)
                        ws[f'K{9 + i}']._style = copy(ws[f'K10']._style)
                        ws[f'L{9 + i}']._style = copy(ws[f'L10']._style)

                    except:
                        pass

                    try:
                        ws[f'M{9 + i}'].value = product['unitOfMeasurement']
                        ws[f'M{9 + i}']._style = copy(ws[f'M10']._style)
                    except:
                        pass

                    try:
                        ws[f'N{9 + i}'].value = int(product['count']['countIRACA']) + int(product['count']['countOthers']) + int(product['count']['countCommunity'])
                        ws[f'N{9 + i}']._style = copy(ws[f'N10']._style)
                    except:
                        ws[f'N{9 + i}']._style = copy(ws[f'N10']._style)

                    try:
                        ws[f'O{9 + i}'].value = int(product['price'])
                        ws[f'O{9 + i}']._style = copy(ws[f'O10']._style)
                    except:
                        pass

                    try:
                        ws[f'P{9 + i}'].value = int(product['price']) * (int(product['count']['countIRACA']) + int(product['count']['countOthers']) + int(product['count']['countCommunity']))
                        ws[f'P{9 + i}']._style = copy(ws[f'P10']._style)
                    except:
                        ws[f'P{9 + i}']._style = copy(ws[f'P10']._style)

                    try:
                        ws[f'Q{9 + i}'].value = int(product['count']['countIRACA'])
                        ws[f'Q{9 + i}']._style = copy(ws[f'Q10']._style)
                    except:
                        ws[f'Q{9 + i}']._style = copy(ws[f'Q10']._style)

                    try:
                        ws[f'R{9 + i}'].value = int(product['price']) * int(product['count']['countIRACA'])
                        ws[f'R{9 + i}']._style = copy(ws[f'R10']._style)
                    except:
                        ws[f'R{9 + i}']._style = copy(ws[f'R10']._style)

                    try:
                        ws[f'S{9 + i}'].value = int(product['count']['countCommunity'])
                        ws[f'S{9 + i}']._style = copy(ws[f'S10']._style)
                    except:
                        ws[f'S{9 + i}']._style = copy(ws[f'S10']._style)

                    try:
                        ws[f'T{9 + i}'].value = int(product['price']) * int(product['count']['countCommunity'])
                        ws[f'T{9 + i}']._style = copy(ws[f'T10']._style)
                    except:
                        ws[f'T{9 + i}']._style = copy(ws[f'T10']._style)

                    try:
                        ws[f'U{9 + i}'].value = int(product['count']['countOthers'])
                        ws[f'U{9 + i}']._style = copy(ws[f'U10']._style)
                    except:
                        ws[f'U{9 + i}']._style = copy(ws[f'U10']._style)

                    try:
                        ws[f'V{9 + i}'].value = int(product['price']) * int(product['count']['countOthers'])
                        ws[f'V{9 + i}']._style = copy(ws[f'V10']._style)
                    except:
                        ws[f'V{9 + i}']._style = copy(ws[f'V10']._style)

                    i += 1

                ws.row_dimensions[9 + i].height = 50

                ws[f'P{9 + i}'] = f"=SUM(P10:P{8 + i})"
                ws[f'R{9 + i}'] = f"=SUM(R10:R{8 + i})"
                ws[f'T{9 + i}'] = f"=SUM(T10:T{8 + i})"
                ws[f'V{9 + i}'] = f"=SUM(V10:V{8 + i})"

                ws.row_dimensions[10 + i].height = 20

                ws.row_dimensions[11 + i].height = 50
                ws[f'R{11 + i}'] = f"=R{9 + i}/V6"


            index += 1



        ws = wb.get_sheet_by_name('Plan de inversion')
        wb.remove(ws)

        index = 1

        for flujo in instance.flujo_caja:
            ws = wb.get_sheet_by_name('Flujo de caja')
            ws = wb.copy_worksheet(ws)
            ws.title = f'Flujo de caja {index}'

            ws.add_image(Image(settings.STATICFILES_DIRS[0] + '/img/logo_prosperidad_FC.png'), 'B1')
            ws.add_image(Image(settings.STATICFILES_DIRS[0] + '/img/logo_ut_FC.png'), 'B3')


            ws['D5'] = f"COMUNIDAD(ES): {instance.get_comunidades()}"

            try:
                ingresos = flujo['ingresos']
            except:
                pass
            else:

                i = 1
                cantidad_ingresos = len(ingresos)

                if len(ingresos) > 1:
                    ws.insert_rows(16, amount=len(ingresos) - 1)

                for ingreso in ingresos:

                    # ws.row_dimensions[14 + i].height = 50

                    ws[f'B{14 + i}'].value = ingreso['description']
                    ws[f'B{14 + i}']._style = copy(ws[f'B15']._style)

                    try:
                        ws[f'C{14 + i}'].value = ingreso['meses'][0]['value']
                    except:
                        ws[f'C{14 + i}'].value = 0

                    ws[f'C{14 + i}']._style = copy(ws[f'C15']._style)

                    try:
                        ws[f'D{14 + i}'].value = ingreso['meses'][1]['value']
                    except:
                        ws[f'D{14 + i}'].value = 0

                    ws[f'D{14 + i}']._style = copy(ws[f'D15']._style)

                    try:
                        ws[f'E{14 + i}'].value = ingreso['meses'][2]['value']
                    except:
                        ws[f'E{14 + i}'].value = 0

                    ws[f'E{14 + i}']._style = copy(ws[f'E15']._style)

                    try:
                        ws[f'F{14 + i}'].value = ingreso['meses'][3]['value']
                    except:
                        ws[f'F{14 + i}'].value = 0
                    ws[f'F{14 + i}']._style = copy(ws[f'F15']._style)

                    try:
                        ws[f'G{14 + i}'].value = ingreso['meses'][4]['value']
                    except:
                        ws[f'G{14 + i}'].value = 0
                    ws[f'G{14 + i}']._style = copy(ws[f'G15']._style)

                    try:
                        ws[f'H{14 + i}'].value = ingreso['meses'][5]['value']
                    except:
                        ws[f'H{14 + i}'].value = 0
                    ws[f'H{14 + i}']._style = copy(ws[f'H15']._style)

                    try:
                        ws[f'I{14 + i}'].value = ingreso['meses'][6]['value']
                    except:
                        ws[f'I{14 + i}'].value = 0
                    ws[f'I{14 + i}']._style = copy(ws[f'I15']._style)

                    try:
                        ws[f'J{14 + i}'].value = ingreso['meses'][7]['value']
                    except:
                        ws[f'J{14 + i}'].value = 0
                    ws[f'J{14 + i}']._style = copy(ws[f'J15']._style)

                    try:
                        ws[f'K{14 + i}'].value = ingreso['meses'][8]['value']
                    except:
                        ws[f'K{14 + i}'].value = 0
                    ws[f'K{14 + i}']._style = copy(ws[f'K15']._style)

                    try:
                        ws[f'L{14 + i}'].value = ingreso['meses'][9]['value']
                    except:
                        ws[f'L{14 + i}'].value = 0
                    ws[f'L{14 + i}']._style = copy(ws[f'L15']._style)

                    try:
                        ws[f'M{14 + i}'].value = ingreso['meses'][10]['value']
                    except:
                        ws[f'M{14 + i}'].value = 0
                    ws[f'M{14 + i}']._style = copy(ws[f'M15']._style)

                    try:
                        ws[f'N{14 + i}'].value = ingreso['meses'][11]['value']
                    except:
                        ws[f'N{14 + i}'].value = 0
                    ws[f'N{14 + i}']._style = copy(ws[f'N15']._style)

                    i += 1

                if len(ingresos) >= 1:
                    celda_ingresos = 14 + i

                    ws[f'C{14 + i}'] = f"=SUM(C15:C{13 + i})"
                    ws[f'D{14 + i}'] = f"=SUM(D15:D{13 + i})"
                    ws[f'E{14 + i}'] = f"=SUM(E15:E{13 + i})"
                    ws[f'F{14 + i}'] = f"=SUM(F15:F{13 + i})"
                    ws[f'G{14 + i}'] = f"=SUM(G15:G{13 + i})"
                    ws[f'H{14 + i}'] = f"=SUM(H15:H{13 + i})"
                    ws[f'I{14 + i}'] = f"=SUM(I15:I{13 + i})"
                    ws[f'J{14 + i}'] = f"=SUM(J15:J{13 + i})"
                    ws[f'K{14 + i}'] = f"=SUM(K15:K{13 + i})"
                    ws[f'L{14 + i}'] = f"=SUM(L15:L{13 + i})"
                    ws[f'M{14 + i}'] = f"=SUM(M15:M{13 + i})"
                    ws[f'N{14 + i}'] = f"=SUM(N15:N{13 + i})"

                else:
                    celda_ingresos = 16

            try:
                egresos = flujo['egresos']
            except:
                pass
            else:

                i = cantidad_ingresos
                j = cantidad_ingresos

                if cantidad_ingresos > 1:
                    inicio = 19 + cantidad_ingresos
                else:
                    inicio = 20

                if len(egresos) > 1:
                    ws.insert_rows(inicio, amount=len(egresos) - 1)

                for egreso in egresos:

                    # ws.row_dimensions[14 + i].height = 50

                    ws[f'B{19 + i}'].value = egreso['description']
                    ws[f'B{19 + i}']._style = copy(ws[f'B{19 + j}']._style)

                    try:
                        ws[f'C{19 + i}'].value = egreso['meses'][0]['value']
                    except:
                        ws[f'C{19 + i}'].value = 0

                    ws[f'C{19 + i}']._style = copy(ws[f'C{19 + j}']._style)

                    try:
                        ws[f'D{19 + i}'].value = egreso['meses'][1]['value']
                    except:
                        ws[f'D{19 + i}'].value = 0

                    ws[f'D{19 + i}']._style = copy(ws[f'D{19 + j}']._style)

                    try:
                        ws[f'E{19 + i}'].value = egreso['meses'][2]['value']
                    except:
                        ws[f'E{19 + i}'].value = 0

                    ws[f'E{19 + i}']._style = copy(ws[f'E{19 + j}']._style)

                    try:
                        ws[f'F{19 + i}'].value = egreso['meses'][3]['value']
                    except:
                        ws[f'F{19 + i}'].value = 0
                    ws[f'F{19 + i}']._style = copy(ws[f'F{19 + j}']._style)

                    try:
                        ws[f'G{19 + i}'].value = egreso['meses'][4]['value']
                    except:
                        ws[f'G{19 + i}'].value = 0
                    ws[f'G{19 + i}']._style = copy(ws[f'G{19 + j}']._style)

                    try:
                        ws[f'H{19 + i}'].value = egreso['meses'][5]['value']
                    except:
                        ws[f'H{19 + i}'].value = 0
                    ws[f'H{19 + i}']._style = copy(ws[f'H{19 + j}']._style)

                    try:
                        ws[f'I{19 + i}'].value = egreso['meses'][6]['value']
                    except:
                        ws[f'I{19 + i}'].value = 0
                    ws[f'I{19 + i}']._style = copy(ws[f'I{19 + j}']._style)

                    try:
                        ws[f'J{19 + i}'].value = egreso['meses'][7]['value']
                    except:
                        ws[f'J{19 + i}'].value = 0
                    ws[f'J{19 + i}']._style = copy(ws[f'J{19 + j}']._style)

                    try:
                        ws[f'K{19 + i}'].value = egreso['meses'][8]['value']
                    except:
                        ws[f'K{19 + i}'].value = 0
                    ws[f'K{19 + i}']._style = copy(ws[f'K{19 + j}']._style)

                    try:
                        ws[f'L{19 + i}'].value = egreso['meses'][9]['value']
                    except:
                        ws[f'L{19 + i}'].value = 0
                    ws[f'L{19 + i}']._style = copy(ws[f'L{19 + j}']._style)

                    try:
                        ws[f'M{19 + i}'].value = egreso['meses'][10]['value']
                    except:
                        ws[f'M{19 + i}'].value = 0
                    ws[f'M{19 + i}']._style = copy(ws[f'M{19 + j}']._style)

                    try:
                        ws[f'N{19 + i}'].value = egreso['meses'][11]['value']
                    except:
                        ws[f'N{19 + i}'].value = 0
                    ws[f'N{19 + i}']._style = copy(ws[f'N{19 + j}']._style)

                    i += 1

                if len(egresos) >= 1:

                    celda_egresos = 19 + i

                    ws[f'C{19 + i}'] = f"=SUM(C{19 + j}:C{18 + i})"
                    ws[f'D{19 + i}'] = f"=SUM(D{19 + j}:D{18 + i})"
                    ws[f'E{19 + i}'] = f"=SUM(E{19 + j}:E{18 + i})"
                    ws[f'F{19 + i}'] = f"=SUM(F{19 + j}:F{18 + i})"
                    ws[f'G{19 + i}'] = f"=SUM(G{19 + j}:G{18 + i})"
                    ws[f'H{19 + i}'] = f"=SUM(H{19 + j}:H{18 + i})"
                    ws[f'I{19 + i}'] = f"=SUM(I{19 + j}:I{18 + i})"
                    ws[f'J{19 + i}'] = f"=SUM(J{19 + j}:J{18 + i})"
                    ws[f'K{19 + i}'] = f"=SUM(K{19 + j}:K{18 + i})"
                    ws[f'L{19 + i}'] = f"=SUM(L{19 + j}:L{18 + i})"
                    ws[f'M{19 + i}'] = f"=SUM(M{19 + j}:M{18 + i})"
                    ws[f'N{19 + i}'] = f"=SUM(N{19 + j}:N{18 + i})"

                    ws[f'C{20 + i}'] = f"=+(C{celda_ingresos}-C{celda_egresos})+C12"
                    ws[f'D{20 + i}'] = f"=+(D{celda_ingresos}-D{celda_egresos})+D12"
                    ws[f'E{20 + i}'] = f"=+(E{celda_ingresos}-E{celda_egresos})+E12"
                    ws[f'F{20 + i}'] = f"=+(F{celda_ingresos}-F{celda_egresos})+F12"
                    ws[f'G{20 + i}'] = f"=+(G{celda_ingresos}-G{celda_egresos})+G12"
                    ws[f'H{20 + i}'] = f"=+(H{celda_ingresos}-H{celda_egresos})+H12"
                    ws[f'I{20 + i}'] = f"=+(I{celda_ingresos}-I{celda_egresos})+I12"
                    ws[f'J{20 + i}'] = f"=+(J{celda_ingresos}-J{celda_egresos})+J12"
                    ws[f'K{20 + i}'] = f"=+(K{celda_ingresos}-K{celda_egresos})+K12"
                    ws[f'L{20 + i}'] = f"=+(L{celda_ingresos}-L{celda_egresos})+L12"
                    ws[f'M{20 + i}'] = f"=+(M{celda_ingresos}-M{celda_egresos})+M12"
                    ws[f'N{20 + i}'] = f"=+(N{celda_ingresos}-N{celda_egresos})+N12"

                    celda_disponible = 20 + i

                else:
                    celda_disponible = 22


                valor_dict = instance.json['data']['lines'][index - 1]

                valor_linea = 0


                for producto in valor_dict['supplies']:
                    valor_linea += float(producto["price"]) * (float(producto["count"]["countIRACA"]) + float(producto["count"]["countOthers"]) + float(producto["count"]["countCommunity"]))




                ws[f'C12'] = valor_linea
                ws[f'D12'] = f"=+C{celda_disponible}"
                ws[f'E12'] = f"=+D{celda_disponible}"
                ws[f'F12'] = f"=+E{celda_disponible}"
                ws[f'G12'] = f"=+F{celda_disponible}"
                ws[f'H12'] = f"=+G{celda_disponible}"
                ws[f'I12'] = f"=+H{celda_disponible}"
                ws[f'J12'] = f"=+I{celda_disponible}"
                ws[f'K12'] = f"=+J{celda_disponible}"
                ws[f'L12'] = f"=+K{celda_disponible}"
                ws[f'M12'] = f"=+L{celda_disponible}"
                ws[f'N12'] = f"=+M{celda_disponible}"






            index += 1

        ws = wb.get_sheet_by_name('Flujo de caja')
        wb.remove(ws)

        """
        ws = wb.get_sheet_by_name('Plan de inversion')
        #logo_sican = Image(settings.STATICFILES_DIRS[0] + '/img/logo_prosperidad_2.png') #size=(400, 85))

        #logo_sican.width = 420
        #logo_sican.height = 82

        #ws.add_image(logo_sican)


        if instance.municipio != None:
            ws['O6'] = instance.municipio.departamento.nombre
            ws['O7'] = instance.municipio.nombre


        ws['E7'] = instance.codigo_proyecto



        if instance.resguado_indigena_consejo_comunitario != None:
            ws['S6'] = instance.resguado_indigena_consejo_comunitario.nombre

        ws['S7'] = instance.get_comunidades()

        ws['V6'] = instance.numero_hogares


        try:
            productos_list = instance.json['data']['supplies']
        except:
            pass
        else:

            i = 1

            if len(productos_list) > 1:
                ws.insert_rows(11, amount=len(productos_list)-1)


            for product in productos_list:

                ws.row_dimensions[9 + i].height = 50

                ws[f'A{9+i}'].value = i
                ws[f'A{9+i}']._style = copy(ws[f'A10']._style)

                try:
                    ws.merge_cells(f'B{9+i}:F{9+i}')
                    ws[f'B{9+i}'].value = product['name']
                    ws[f'B{9+i}']._style = copy(ws[f'B10']._style)
                    ws[f'C{9+i}']._style = copy(ws[f'C10']._style)
                    ws[f'D{9+i}']._style = copy(ws[f'D10']._style)
                    ws[f'E{9+i}']._style = copy(ws[f'E10']._style)
                    ws[f'F{9+i}']._style = copy(ws[f'F10']._style)

                except:
                    pass


                try:
                    ws.merge_cells(f'G{9 + i}:L{9 + i}')
                    ws[f'G{9+i}'].value = product['description']
                    ws[f'G{9+i}']._style = copy(ws[f'G10']._style)
                    ws[f'H{9 + i}']._style = copy(ws[f'H10']._style)
                    ws[f'I{9 + i}']._style = copy(ws[f'I10']._style)
                    ws[f'J{9 + i}']._style = copy(ws[f'J10']._style)
                    ws[f'K{9 + i}']._style = copy(ws[f'K10']._style)
                    ws[f'L{9 + i}']._style = copy(ws[f'L10']._style)

                except:
                    pass



                try:
                    ws[f'M{9+i}'].value = product['unit_of_measurement']
                    ws[f'M{9 + i}']._style = copy(ws[f'M10']._style)
                except:
                    pass


                try:
                    ws[f'N{9+i}'].value = int(product['count']['count_IRACA']) + int(product['count']['count_Others']) + int(product['count']['count_Community'])
                    ws[f'N{9 + i}']._style = copy(ws[f'N10']._style)
                except:
                    pass


                try:
                    ws[f'O{9+i}'].value = int(product['price'])
                    ws[f'O{9 + i}']._style = copy(ws[f'O10']._style)
                except:
                    pass


                try:
                    ws[f'P{9+i}'].value = int(product['price']) * (int(product['count']['count_IRACA']) + int(product['count']['count_Others']) + int(product['count']['count_Community']))
                    ws[f'P{9 + i}']._style = copy(ws[f'P10']._style)
                except:
                    pass


                try:
                    ws[f'Q{9+i}'].value = int(product['count']['count_IRACA'])
                    ws[f'Q{9 + i}']._style = copy(ws[f'Q10']._style)
                except:
                    pass


                try:
                    ws[f'R{9+i}'].value = int(product['price']) * int(product['count']['count_IRACA'])
                    ws[f'R{9 + i}']._style = copy(ws[f'R10']._style)
                except:
                    pass





                try:
                    ws[f'S{9+i}'].value = int(product['count']['count_Community'])
                    ws[f'S{9 + i}']._style = copy(ws[f'S10']._style)
                except:
                    pass


                try:
                    ws[f'T{9+i}'].value = int(product['price']) * int(product['count']['count_Community'])
                    ws[f'T{9 + i}']._style = copy(ws[f'T10']._style)
                except:
                    pass



                try:
                    ws[f'U{9+i}'].value = int(product['count']['count_Others'])
                    ws[f'U{9 + i}']._style = copy(ws[f'U10']._style)
                except:
                    pass


                try:
                    ws[f'V{9+i}'].value = int(product['price']) * int(product['count']['count_Others'])
                    ws[f'V{9 + i}']._style = copy(ws[f'V10']._style)
                except:
                    pass

                i += 1


            ws.row_dimensions[9 + i].height = 50

            ws[f'P{9+i}'] = f"=SUM(P10:P{8+i})"
            ws[f'R{9+i}'] = f"=SUM(R10:R{8+i})"
            ws[f'T{9+i}'] = f"=SUM(T10:T{8+i})"
            ws[f'V{9+i}'] = f"=SUM(V10:V{8+i})"

            ws.row_dimensions[10 + i].height = 20

            ws.row_dimensions[11 + i].height = 50
            ws[f'R{11+i}'] = f"=R{9+i}/V6"

        ws = wb.get_sheet_by_name('Flujo de caja')

        ws['D5'] = f"COMUNIDAD(ES): {instance.get_comunidades()}"

        try:
            ingresos = instance.flujo_caja['ingresos']
        except:
            pass
        else:

            i = 1
            cantidad_ingresos = len(ingresos)

            if len(ingresos) > 1:
                ws.insert_rows(16, amount=len(ingresos)-1)


            for ingreso in ingresos:

                #ws.row_dimensions[14 + i].height = 50

                ws[f'B{14+i}'].value = ingreso['description']
                ws[f'B{14+i}']._style = copy(ws[f'B15']._style)


                try:
                    ws[f'C{14 + i}'].value = ingreso['meses'][0]['value']
                except:
                    ws[f'C{14 + i}'].value = 0

                ws[f'C{14 + i}']._style = copy(ws[f'C15']._style)


                try:
                    ws[f'D{14 + i}'].value = ingreso['meses'][1]['value']
                except:
                    ws[f'D{14 + i}'].value = 0

                ws[f'D{14 + i}']._style = copy(ws[f'D15']._style)


                try:
                    ws[f'E{14 + i}'].value = ingreso['meses'][2]['value']
                except:
                    ws[f'E{14 + i}'].value = 0

                ws[f'E{14 + i}']._style = copy(ws[f'E15']._style)


                try:
                    ws[f'F{14 + i}'].value = ingreso['meses'][3]['value']
                except:
                    ws[f'F{14 + i}'].value = 0
                ws[f'F{14 + i}']._style = copy(ws[f'F15']._style)


                try:
                    ws[f'G{14 + i}'].value = ingreso['meses'][4]['value']
                except:
                    ws[f'G{14 + i}'].value = 0
                ws[f'G{14 + i}']._style = copy(ws[f'G15']._style)


                try:
                    ws[f'H{14 + i}'].value = ingreso['meses'][5]['value']
                except:
                    ws[f'H{14 + i}'].value = 0
                ws[f'H{14 + i}']._style = copy(ws[f'H15']._style)


                try:
                    ws[f'I{14 + i}'].value = ingreso['meses'][6]['value']
                except:
                    ws[f'I{14 + i}'].value = 0
                ws[f'I{14 + i}']._style = copy(ws[f'I15']._style)


                try:
                    ws[f'J{14 + i}'].value = ingreso['meses'][7]['value']
                except:
                    ws[f'J{14 + i}'].value = 0
                ws[f'J{14 + i}']._style = copy(ws[f'J15']._style)


                try:
                    ws[f'K{14 + i}'].value = ingreso['meses'][8]['value']
                except:
                    ws[f'K{14 + i}'].value = 0
                ws[f'K{14 + i}']._style = copy(ws[f'K15']._style)



                try:
                    ws[f'L{14 + i}'].value = ingreso['meses'][9]['value']
                except:
                    ws[f'L{14 + i}'].value = 0
                ws[f'L{14 + i}']._style = copy(ws[f'L15']._style)


                try:
                    ws[f'M{14 + i}'].value = ingreso['meses'][10]['value']
                except:
                    ws[f'M{14 + i}'].value = 0
                ws[f'M{14 + i}']._style = copy(ws[f'M15']._style)


                try:
                    ws[f'N{14 + i}'].value = ingreso['meses'][11]['value']
                except:
                    ws[f'N{14 + i}'].value = 0
                ws[f'N{14 + i}']._style = copy(ws[f'N15']._style)


                i += 1

            if len(ingresos) >= 1:
                celda_ingresos = 14 + i


                ws[f'C{14 + i}'] = f"=SUM(C15:C{13 + i})"
                ws[f'D{14 + i}'] = f"=SUM(D15:D{13 + i})"
                ws[f'E{14 + i}'] = f"=SUM(E15:E{13 + i})"
                ws[f'F{14 + i}'] = f"=SUM(F15:F{13 + i})"
                ws[f'G{14 + i}'] = f"=SUM(G15:G{13 + i})"
                ws[f'H{14 + i}'] = f"=SUM(H15:H{13 + i})"
                ws[f'I{14 + i}'] = f"=SUM(I15:I{13 + i})"
                ws[f'J{14 + i}'] = f"=SUM(J15:J{13 + i})"
                ws[f'K{14 + i}'] = f"=SUM(K15:K{13 + i})"
                ws[f'L{14 + i}'] = f"=SUM(L15:L{13 + i})"
                ws[f'M{14 + i}'] = f"=SUM(M15:M{13 + i})"
                ws[f'N{14 + i}'] = f"=SUM(N15:N{13 + i})"

            else:
                celda_ingresos = 16


        try:
            egresos = instance.flujo_caja['egresos']
        except:
            pass
        else:

            i = cantidad_ingresos
            j = cantidad_ingresos

            if cantidad_ingresos > 1:
                inicio = 19 + cantidad_ingresos
            else:
                inicio = 20

            if len(egresos) > 1:

                ws.insert_rows(inicio, amount=len(egresos)-1)


            for egreso in egresos:

                #ws.row_dimensions[14 + i].height = 50

                ws[f'B{18+i}'].value = egreso['description']
                ws[f'B{18+i}']._style = copy(ws[f'B{18+j}']._style)


                try:
                    ws[f'C{18 + i}'].value = egreso['meses'][0]['value']
                except:
                    ws[f'C{18 + i}'].value = 0

                ws[f'C{18 + i}']._style = copy(ws[f'C{18+j}']._style)


                try:
                    ws[f'D{18 + i}'].value = egreso['meses'][1]['value']
                except:
                    ws[f'D{18 + i}'].value = 0

                ws[f'D{18 + i}']._style = copy(ws[f'D{18+j}']._style)


                try:
                    ws[f'E{18 + i}'].value = egreso['meses'][2]['value']
                except:
                    ws[f'E{18 + i}'].value = 0

                ws[f'E{18 + i}']._style = copy(ws[f'E{18+j}']._style)


                try:
                    ws[f'F{18 + i}'].value = egreso['meses'][3]['value']
                except:
                    ws[f'F{18 + i}'].value = 0
                ws[f'F{18 + i}']._style = copy(ws[f'F{18+j}']._style)


                try:
                    ws[f'G{18 + i}'].value = egreso['meses'][4]['value']
                except:
                    ws[f'G{18 + i}'].value = 0
                ws[f'G{18 + i}']._style = copy(ws[f'G{18+j}']._style)


                try:
                    ws[f'H{18 + i}'].value = egreso['meses'][5]['value']
                except:
                    ws[f'H{18 + i}'].value = 0
                ws[f'H{18 + i}']._style = copy(ws[f'H{18+j}']._style)


                try:
                    ws[f'I{18 + i}'].value = egreso['meses'][6]['value']
                except:
                    ws[f'I{18 + i}'].value = 0
                ws[f'I{18 + i}']._style = copy(ws[f'I{18+j}']._style)


                try:
                    ws[f'J{18 + i}'].value = egreso['meses'][7]['value']
                except:
                    ws[f'J{18 + i}'].value = 0
                ws[f'J{18 + i}']._style = copy(ws[f'J{18+j}']._style)


                try:
                    ws[f'K{18 + i}'].value = egreso['meses'][8]['value']
                except:
                    ws[f'K{18 + i}'].value = 0
                ws[f'K{18 + i}']._style = copy(ws[f'K{18+j}']._style)



                try:
                    ws[f'L{18 + i}'].value = egreso['meses'][9]['value']
                except:
                    ws[f'L{18 + i}'].value = 0
                ws[f'L{18 + i}']._style = copy(ws[f'L{18+j}']._style)


                try:
                    ws[f'M{18 + i}'].value = egreso['meses'][10]['value']
                except:
                    ws[f'M{18 + i}'].value = 0
                ws[f'M{18 + i}']._style = copy(ws[f'M{18+j}']._style)


                try:
                    ws[f'N{18 + i}'].value = egreso['meses'][11]['value']
                except:
                    ws[f'N{18 + i}'].value = 0
                ws[f'N{18 + i}']._style = copy(ws[f'N{18+j}']._style)


                i += 1


            if len(egresos) >= 1:

                celda_egresos = 18 + i

                ws[f'C{18 + i}'] = f"=SUM(C{18+j}:C{17 + i})"
                ws[f'D{18 + i}'] = f"=SUM(D{18+j}:D{17 + i})"
                ws[f'E{18 + i}'] = f"=SUM(E{18+j}:E{17 + i})"
                ws[f'F{18 + i}'] = f"=SUM(F{18+j}:F{17 + i})"
                ws[f'G{18 + i}'] = f"=SUM(G{18+j}:G{17 + i})"
                ws[f'H{18 + i}'] = f"=SUM(H{18+j}:H{17 + i})"
                ws[f'I{18 + i}'] = f"=SUM(I{18+j}:I{17 + i})"
                ws[f'J{18 + i}'] = f"=SUM(J{18+j}:J{17 + i})"
                ws[f'K{18 + i}'] = f"=SUM(K{18+j}:K{17 + i})"
                ws[f'L{18 + i}'] = f"=SUM(L{18+j}:L{17 + i})"
                ws[f'M{18 + i}'] = f"=SUM(M{18+j}:M{17 + i})"
                ws[f'N{18 + i}'] = f"=SUM(N{18+j}:N{17 + i})"

                ws[f'C{19 + i}'] = f"=+(C{celda_ingresos}-C{celda_egresos})+C12"
                ws[f'D{19 + i}'] = f"=+(D{celda_ingresos}-D{celda_egresos})+D12"
                ws[f'E{19 + i}'] = f"=+(E{celda_ingresos}-E{celda_egresos})+E12"
                ws[f'F{19 + i}'] = f"=+(F{celda_ingresos}-F{celda_egresos})+F12"
                ws[f'G{19 + i}'] = f"=+(G{celda_ingresos}-G{celda_egresos})+G12"
                ws[f'H{19 + i}'] = f"=+(H{celda_ingresos}-H{celda_egresos})+H12"
                ws[f'I{19 + i}'] = f"=+(I{celda_ingresos}-I{celda_egresos})+I12"
                ws[f'J{19 + i}'] = f"=+(J{celda_ingresos}-J{celda_egresos})+J12"
                ws[f'K{19 + i}'] = f"=+(K{celda_ingresos}-K{celda_egresos})+K12"
                ws[f'L{19 + i}'] = f"=+(L{celda_ingresos}-L{celda_egresos})+L12"
                ws[f'M{19 + i}'] = f"=+(M{celda_ingresos}-M{celda_egresos})+M12"
                ws[f'N{19 + i}'] = f"=+(N{celda_ingresos}-N{celda_egresos})+N12"

                celda_disponible = 19 + i

            else:
                celda_egresos = 20
                celda_disponible = 21


            ws[f'C12'] = instance.valor.amount
            ws[f'D12'] = f"=+C{celda_disponible}"
            ws[f'E12'] = f"=+D{celda_disponible}"
            ws[f'F12'] = f"=+E{celda_disponible}"
            ws[f'G12'] = f"=+F{celda_disponible}"
            ws[f'H12'] = f"=+G{celda_disponible}"
            ws[f'I12'] = f"=+H{celda_disponible}"
            ws[f'J12'] = f"=+I{celda_disponible}"
            ws[f'K12'] = f"=+J{celda_disponible}"
            ws[f'L12'] = f"=+K{celda_disponible}"
            ws[f'M12'] = f"=+L{celda_disponible}"
            ws[f'N12'] = f"=+M{celda_disponible}"

        """

        wb.save(output)
        filename = str(instance.id) + '.xlsx'
        instance.file.save(filename, File(output))

        post_save.connect(ProyectosApiPostSave, sender=ProyectosApi)

from __future__ import absolute_import, unicode_literals
from config.celery import app
import openpyxl
from io import BytesIO
from usuarios.models import User
from django.conf import settings
from openpyxl.drawing.image import Image
from django.core.files import File
from mail_templated import send_mail
from direccion_financiera import models
from config.functions import construir_reporte
from reportes import models as models_reportes
from usuarios.models import Notifications
from formacion import models
from six.moves.urllib import parse as urlparse
import os
import ftplib
import io

@app.task
def build_resultado_actualizacion_sedes(id):

    location = os.getenv('FTP_SICAN_2018')
    splitted_url = urlparse.urlparse(location)
    actualizacion = models.ActualizacionSedes.objects.get(id=id)

    ftp = ftplib.FTP()

    input = io.BytesIO()

    try:
        ftp.connect(splitted_url.hostname,int(splitted_url.port))
        ftp.login(splitted_url.username,splitted_url.password)
        ftp.retrbinary('RETR /' + actualizacion.file.name, input.write)
        input.seek(0)
    except:
        return 'Error FTP'
    else:
        ftp.quit()
        wb = openpyxl.load_workbook(input, use_iterators=True)
        ws = wb.get_sheet_by_name('ACTUALIZACION SEDES')

        nuevos = 0
        modificados = 0
        rechazados = 0

        proceso = "SICAN-ACTUALIZACIÓN-SEDES"

        titulos = [
            'Código DANE SEDE', 'Nombre de la sede', 'Código DANE Institución', 'Nombre Institución', 'Código municipio', 'Resultado']

        formatos = ['0', 'General', '0', 'General', '0', 'General']

        ancho_columnas = [20, 40, 40, 40, 40, 40]

        contenidos = []

        for fila in ws.iter_rows(row_offset=1):

            resultado = ''

            if fila[0].value != '' and fila[0].value != None:

                try:
                    municipio = models.Municipios.objects.get(numero=fila[4].value)
                except:
                    resultado = 'No existen municipios con el código ingresado'
                    rechazados += 1
                else:
                    if fila[0].value != '' and fila[1].value != '' and fila[2].value != '' and fila[3].value != '' \
                            and fila[0].value != None and fila[1].value != None and fila[2].value != None and fila[3].value != None:

                        cantidad = models.Sedes.objects.filter(dane_sede = fila[0].value).count()
                        if not cantidad > 0:
                            sede = models.Sedes.objects.create(
                                municipio = municipio,
                                dane_sede = fila[0].value,
                                nombre_sede=fila[1].value,
                                dane_ie=fila[2].value,
                                nombre_ie=fila[3].value
                            )
                            resultado = 'Sede creada'
                            nuevos += 1
                        else:
                            sede = models.Sedes.objects.filter(dane_sede = fila[0].value)
                            sede.update(
                                municipio=municipio,
                                nombre_sede=fila[1].value,
                                dane_ie=fila[2].value,
                                nombre_ie=fila[3].value
                            )
                            resultado = 'Sede actualizada'
                            modificados += 1

                    else:
                        resultado = 'Información obligatoria incompleta'
                        rechazados += 1


            contenidos.append([
                fila[0].value,
                fila[1].value,
                fila[2].value,
                fila[3].value,
                fila[4].value,
                resultado
            ])

        output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, 'Actualiación de sedes', actualizacion.creation,
                                   actualizacion.usuario_creacion, proceso)

        filename = str(actualizacion.id) + '.xlsx'
        actualizacion.result.save(filename, File(output))

        actualizacion.nuevos = nuevos
        actualizacion.modificados = modificados
        actualizacion.rechazados = rechazados
        actualizacion.save()

        return "Resultado actualizacion completo"


@app.task
def build_resultado_actualizacion_docentes(id):

    location = os.getenv('FTP_SICAN_2018')
    splitted_url = urlparse.urlparse(location)
    actualizacion = models.ActualizacionDocentes.objects.get(id=id)

    ftp = ftplib.FTP()

    input = io.BytesIO()

    try:
        ftp.connect(splitted_url.hostname,int(splitted_url.port))
        ftp.login(splitted_url.username,splitted_url.password)
        ftp.retrbinary('RETR /' + actualizacion.file.name, input.write)
        input.seek(0)
    except:
        return 'Error FTP'
    else:
        ftp.quit()
        wb = openpyxl.load_workbook(input, use_iterators=True)
        ws = wb.get_sheet_by_name('ACTUALIZACION DE DOCENTES')

        nuevos = 0
        modificados = 0
        rechazados = 0

        proceso = "SICAN-ACTUALIZACIÓN-DOCENTES"

        titulos = [
            'Código DANE SEDE', 'Nombres del docente', 'Apellidos del docente', 'Cedula del docente', 'Vigencia de formación',
            'Diplomado', 'Resultado']

        formatos = ['0', 'General', 'General', 'General', '0', 'General',
                    'General', 'General']

        ancho_columnas = [20, 40, 40, 40, 40, 40,
                          40, 40]

        contenidos = []

        for fila in ws.iter_rows(row_offset=1):

            resultado = ''

            if fila[0].value != '' and fila[0].value != None:

                try:
                    sede = models.Sedes.objects.get(dane_sede=fila[0].value)
                except:
                    resultado = 'No existe el código dane de la sede educativa'
                    rechazados += 1
                else:
                    if fila[1].value != '' and fila[2].value != '' and fila[3].value != '' and fila[4].value != '' and fila[5].value != ''\
                            and fila[1].value != None and fila[2].value != None and fila[3].value != None and fila[4].value != None and fila[5].value != None:

                        cantidad = models.DocentesFormados.objects.filter(cedula = fila[3].value).count()
                        if not cantidad > 0:
                            docente = models.DocentesFormados.objects.create(
                                sede = sede,
                                nombres=fila[1].value,
                                apellidos=fila[2].value,
                                cedula=fila[3].value,
                                vigencia=fila[4].value,
                                diplomado=fila[5].value,
                            )
                            resultado = 'Docente creado'
                            nuevos += 1
                        else:
                            sede = models.DocentesFormados.objects.filter(cedula = fila[3].value)
                            sede.update(
                                sede=sede,
                                nombres=fila[1].value,
                                apellidos=fila[2].value,
                                vigencia=fila[4].value,
                                diplomado=fila[5].value,
                            )
                            resultado = 'Docente actualizado'
                            modificados += 1

                    else:
                        resultado = 'Información obligatoria incompleta'
                        rechazados += 1


            contenidos.append([
                fila[0].value,
                fila[1].value,
                fila[2].value,
                fila[3].value,
                fila[4].value,
                fila[5].value,
                resultado
            ])

        output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, 'Actualiación de docentes', actualizacion.creation,
                                   actualizacion.usuario_creacion, proceso)

        filename = str(actualizacion.id) + '.xlsx'
        actualizacion.result.save(filename, File(output))

        actualizacion.nuevos = nuevos
        actualizacion.modificados = modificados
        actualizacion.rechazados = rechazados
        actualizacion.save()

        return "Resultado actualizacion completo"
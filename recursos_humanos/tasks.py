from __future__ import absolute_import, unicode_literals
from config.celery import app
from mail_templated import send_mail
import openpyxl
from django.conf import settings
from io import BytesIO
from recursos_humanos import models
from django.core.files import File
from reportes import models as models_reportes
from config.functions import construir_reporte

@app.task
def send_mail_templated_certificacion(template,dictionary,from_email,list_to_email,attachments):
    send_mail(template, dictionary, from_email, list_to_email,attachments = attachments)
    return 'Email enviado'

@app.task
def build_formato_hv(id):

    hv = models.Hv.objects.get(id = id)
    output = BytesIO()

    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/formato_hv.xlsx')
    ws = wb.get_sheet_by_name('Formato Hoja de vida')


    ws['C2'] = 'Asociación Nacional Para el Desarrollo Social - ANDES'.upper()
    ws['C3'] = hv.region.upper()
    ws['C5'] = hv.contratista.get_full_name().upper()
    ws['C6'] = hv.contratista.cedula
    ws['C7'] = hv.contratista.birthday.strftime('%d/%m/%Y') if hv.contratista.birthday != None else ''
    ws['C8'] = hv.cargo.upper()

    ws['B13'] = hv.titulo_1.upper()
    ws['C13'] = hv.institucion_1.upper()
    ws['E13'] = hv.nivel_1.upper()
    ws['F13'] = hv.grado_1.strftime('%d/%m/%Y') if hv.grado_1 != None else ''
    ws['G13'] = hv.folio_1

    ws['B14'] = hv.titulo_2.upper()
    ws['C14'] = hv.institucion_2.upper()
    ws['E14'] = hv.nivel_2.upper()
    ws['F14'] = hv.grado_2.strftime('%d/%m/%Y') if hv.grado_2 != None else ''
    ws['G14'] = hv.folio_2

    ws['B15'] = hv.titulo_3.upper()
    ws['C15'] = hv.institucion_3.upper()
    ws['E15'] = hv.nivel_3.upper()
    ws['F15'] = hv.grado_3.strftime('%d/%m/%Y') if hv.grado_3 != None else ''
    ws['G15'] = hv.folio_3

    ws['B16'] = hv.titulo_4.upper()
    ws['C16'] = hv.institucion_4.upper()
    ws['E16'] = hv.nivel_4.upper()
    ws['F16'] = hv.grado_4.strftime('%d/%m/%Y') if hv.grado_4 != None else ''
    ws['G16'] = hv.folio_4

    ws['B17'] = hv.titulo_5.upper()
    ws['C17'] = hv.institucion_5.upper()
    ws['E17'] = hv.nivel_5.upper()
    ws['F17'] = hv.grado_5.strftime('%d/%m/%Y') if hv.grado_5 != None else ''
    ws['G17'] = hv.folio_5

    ws['B18'] = hv.titulo_6.upper()
    ws['C18'] = hv.institucion_6.upper()
    ws['E18'] = hv.nivel_6.upper()
    ws['F18'] = hv.grado_6.strftime('%d/%m/%Y') if hv.grado_6 != None else ''
    ws['G18'] = hv.folio_6

    ws['B19'] = hv.titulo_7.upper()
    ws['C19'] = hv.institucion_7.upper()
    ws['E19'] = hv.nivel_7.upper()
    ws['F19'] = hv.grado_7.strftime('%d/%m/%Y') if hv.grado_7 != None else ''
    ws['G19'] = hv.folio_7

    ws['B24'] = hv.numero_tarjeta.upper()
    ws['C24'] = hv.fecha_expedicion.strftime('%d/%m/%Y') if hv.fecha_expedicion != None else ''
    ws['E24'] = hv.folio

    ws['B30'] = hv.empresa_1.upper()
    ws['C30'] = hv.fecha_inicio_1.strftime('%d/%m/%Y') if hv.fecha_inicio_1 != None else ''
    ws['D30'] = hv.fecha_fin_1.strftime('%d/%m/%Y') if hv.fecha_fin_1 != None else ''
    ws['F30'] = hv.cargo_1.upper()
    ws['I30'] = hv.folio_empresa_1
    ws['K30'] = hv.observaciones_1.upper()

    ws['B31'] = hv.empresa_2.upper()
    ws['C31'] = hv.fecha_inicio_2.strftime('%d/%m/%Y') if hv.fecha_inicio_2 != None else ''
    ws['D31'] = hv.fecha_fin_2.strftime('%d/%m/%Y') if hv.fecha_fin_2 != None else ''
    ws['F31'] = hv.cargo_2.upper()
    ws['I31'] = hv.folio_empresa_2
    ws['K31'] = hv.observaciones_2.upper()

    ws['B32'] = hv.empresa_3.upper()
    ws['C32'] = hv.fecha_inicio_3.strftime('%d/%m/%Y') if hv.fecha_inicio_3 != None else ''
    ws['D32'] = hv.fecha_fin_3.strftime('%d/%m/%Y') if hv.fecha_fin_3 != None else ''
    ws['F32'] = hv.cargo_3.upper()
    ws['I32'] = hv.folio_empresa_3
    ws['K32'] = hv.observaciones_3.upper()

    ws['B33'] = hv.empresa_4.upper()
    ws['C33'] = hv.fecha_inicio_4.strftime('%d/%m/%Y') if hv.fecha_inicio_4 != None else ''
    ws['D33'] = hv.fecha_fin_4.strftime('%d/%m/%Y') if hv.fecha_fin_4 != None else ''
    ws['F33'] = hv.cargo_4.upper()
    ws['I33'] = hv.folio_empresa_4
    ws['K33'] = hv.observaciones_4.upper()

    ws['B34'] = hv.empresa_5.upper()
    ws['C34'] = hv.fecha_inicio_5.strftime('%d/%m/%Y') if hv.fecha_inicio_5 != None else ''
    ws['D34'] = hv.fecha_fin_5.strftime('%d/%m/%Y') if hv.fecha_fin_5 != None else ''
    ws['F34'] = hv.cargo_5.upper()
    ws['I34'] = hv.folio_empresa_5
    ws['K34'] = hv.observaciones_5.upper()

    ws['B35'] = hv.empresa_6.upper()
    ws['C35'] = hv.fecha_inicio_6.strftime('%d/%m/%Y') if hv.fecha_inicio_6 != None else ''
    ws['D35'] = hv.fecha_fin_6.strftime('%d/%m/%Y') if hv.fecha_fin_6 != None else ''
    ws['F35'] = hv.cargo_6.upper()
    ws['I35'] = hv.folio_empresa_6
    ws['K35'] = hv.observaciones_6.upper()

    ws['B36'] = hv.empresa_7.upper()
    ws['C36'] = hv.fecha_inicio_7.strftime('%d/%m/%Y') if hv.fecha_inicio_7 != None else ''
    ws['D36'] = hv.fecha_fin_7.strftime('%d/%m/%Y') if hv.fecha_fin_7 != None else ''
    ws['F36'] = hv.cargo_7.upper()
    ws['I36'] = hv.folio_empresa_7
    ws['K36'] = hv.observaciones_7.upper()

    ws['B37'] = hv.empresa_8.upper()
    ws['C37'] = hv.fecha_inicio_8.strftime('%d/%m/%Y') if hv.fecha_inicio_8 != None else ''
    ws['D37'] = hv.fecha_fin_8.strftime('%d/%m/%Y') if hv.fecha_fin_8 != None else ''
    ws['F37'] = hv.cargo_8.upper()
    ws['I37'] = hv.folio_empresa_8
    ws['K37'] = hv.observaciones_8.upper()

    ws['B38'] = hv.empresa_9.upper()
    ws['C38'] = hv.fecha_inicio_9.strftime('%d/%m/%Y') if hv.fecha_inicio_9 != None else ''
    ws['D38'] = hv.fecha_fin_9.strftime('%d/%m/%Y') if hv.fecha_fin_9 != None else ''
    ws['F38'] = hv.cargo_9.upper()
    ws['I38'] = hv.folio_empresa_9
    ws['K38'] = hv.observaciones_9.upper()

    ws['B39'] = hv.empresa_10.upper()
    ws['C39'] = hv.fecha_inicio_10.strftime('%d/%m/%Y') if hv.fecha_inicio_10 != None else ''
    ws['D39'] = hv.fecha_fin_10.strftime('%d/%m/%Y') if hv.fecha_fin_10 != None else ''
    ws['F39'] = hv.cargo_10.upper()
    ws['I39'] = hv.folio_empresa_10
    ws['K39'] = hv.observaciones_10.upper()

    ws['B40'] = hv.empresa_11.upper()
    ws['C40'] = hv.fecha_inicio_11.strftime('%d/%m/%Y') if hv.fecha_inicio_11 != None else ''
    ws['D40'] = hv.fecha_fin_11.strftime('%d/%m/%Y') if hv.fecha_fin_11 != None else ''
    ws['F40'] = hv.cargo_11.upper()
    ws['I40'] = hv.folio_empresa_11
    ws['K40'] = hv.observaciones_11.upper()



    wb.save(output)
    filename = 'Formato.xlsx'
    hv.excel.save(filename, File(output))
    hv.save()

    return "Resultado actualizacion completo"

@app.task
def build_listado_contratos(id):
    from recursos_humanos.models import Contratos, Soportes, SoportesContratos
    reporte = models_reportes.Reportes.objects.get(id = id)
    proceso = "SICAN-LST-CONTRATOS"


    titulos = ['Consecutivo', 'Código', 'Estado contrato','Proyecto', 'Cargo' ,'Fecha de creación','Contratista', 'Cédula', 'Lugar de Expedición', 'Dirección','Residencia','Celular','Tipo de sangre', 'Fecha de nacimiento' ,'Correo','Fecha inicio', 'Fecha finalización', 'Tipo contrato', 'Valor',
               'Fecha de legalización']

    formatos = ['0', 'General', 'General','General', 'General', 'General', 'General', 'General', '0', 'General', 'General','General','General','dd/mm/yyyy','General', 'dd/mm/yyyy', 'dd/mm/yyyy', 'General', '"$"#,##0_);("$"#,##0)',
                'dd/mm/yyyy']

    ancho_columnas = [20, 30, 30, 30, 30, 30, 40, 25, 25, 25, 25,25,25,25,25,25, 35, 40, 20,
                      40]

    contenidos = []

    for soporte in Soportes.objects.all().order_by('numero'):
        titulos.append(soporte.nombre)
        formatos.append('General')
        ancho_columnas.append(30)

    i = 0
    for contrato in Contratos.objects.filter(visible = True).order_by('-creation'):
        i += 1

        soportes = []

        for soporte in Soportes.objects.all().order_by('numero'):
            try:
                s = SoportesContratos.objects.get(soporte = soporte, contrato = contrato)
            except:
                soportes.append('')
            else:
                url = s.url_file()
                if url == None:
                    soportes.append('')
                else:
                    soportes.append(('Link','https://sican.asoandes.org' + s.url_file()))


        contenidos.append([
            int(i),
            contrato.nombre,
            contrato.get_estado_contrato(),
            contrato.get_proyecto(),
            contrato.get_cargo(),
            contrato.pretty_creation_datetime(),
            contrato.contratista.get_full_name(),
            contrato.contratista.cedula,
            contrato.contratista.get_lugar_expedicion(),
            contrato.get_direccion(),
            contrato.contratista.get_lugar_residencia(),
            contrato.contratista.get_celular(),
            contrato.get_tipo_sangre(),
            contrato.get_fecha_nacimiento(),
            contrato.get_email(),
            contrato.inicio,
            contrato.fin,
            contrato.tipo_contrato,
            contrato.valor.amount,
            contrato.fecha_legalizacion
        ]+soportes)

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))


    return "Archivo paquete ID: " + filename

@app.task
def build_reporte_hv(id):
    from recursos_humanos.models import Hv
    reporte = models_reportes.Reportes.objects.get(id = id)
    proceso = "SICAN-LST-HV"


    titulos = ['Consecutivo', 'Contratista', 'Cedula', 'Región', 'Envio', 'Consecutivo', 'Cargo', 'Estado', 'Observaciones', 'Hv', 'Excel']

    formatos = ['0', 'General', 'General', 'General', 'General', 'General', 'General', 'General', 'General', 'General', 'General']

    ancho_columnas = [20, 40, 30, 30, 30, 30, 30, 30, 40, 30, 30]

    contenidos = []


    i = 0
    for hv in Hv.objects.all().order_by('-creation'):
        i += 1

        contenidos.append([
            int(i),
            hv.contratista.get_full_name(),
            hv.contratista.cedula,
            hv.region,
            hv.envio,
            hv.consecutivo_cargo,
            hv.cargo,
            hv.estado,
            hv.observacion,
            ('LINK','https://sican.asoandes.org/'+hv.url_file()),
            ('LINK', 'https://sican.asoandes.org/' + hv.url_excel())
        ])

    output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, reporte.nombre, reporte.creation, reporte.usuario, proceso)

    filename = str(reporte.id) + '.xlsx'
    reporte.file.save(filename, File(output))


    return "Archivo paquete ID: " + filename